import ast
import pathlib

import pytest
from PIL import Image

import app_paths
from clients_db import ClientsDB
from gui import _crop_logo_preview_image, _safe_make_logo_photo
from models import Client, Supplier, normalize_rgb_color
from orders import (
    DEFAULT_FOOTER_NOTE,
    REPORTLAB_OK,
    generate_pdf_order_platypus,
    write_order_excel,
)


def test_clients_db_persists_logo_fields(tmp_path):
    client = Client(
        name="ACME",
        website="https://acme.example",
        accent_color="#0A64C8",
        logo_path="client_logos/acme.png",
        logo_crop={"left": 5, "top": 10, "right": 105, "bottom": 210},
    )
    db = ClientsDB([client])
    path = tmp_path / "clients.json"
    db.save(path)

    loaded = ClientsDB.load(path)
    assert loaded.clients
    saved = loaded.clients[0]
    assert saved.website == "https://acme.example"
    assert saved.accent_color == "#0A64C8"
    assert saved.logo_path == "client_logos/acme.png"
    assert saved.logo_crop == {"left": 5, "top": 10, "right": 105, "bottom": 210}


def test_client_normalizes_accent_color_formats():
    client_rgb = Client.from_any({"name": "ACME", "accent_color": "12, 34, 56"})
    assert client_rgb.accent_color == "#0C2238"

    client_hex = Client.from_any({"name": "ACME", "kleur": "#ff77ff"})
    assert client_hex.accent_color == "#FF77FF"

    client_website = Client.from_any({"name": "ACME", "website": "https://acme.example"})
    assert client_website.website == "https://acme.example"

    assert normalize_rgb_color("255,119,255") == "#FF77FF"
    assert normalize_rgb_color("#0c2238") == "#0C2238"
    assert normalize_rgb_color("ongeldig") is None


def test_crop_logo_preview_image_ignores_invalid_crop_data():
    img = Image.new("RGBA", (120, 60), "red")

    cropped = _crop_logo_preview_image(
        img,
        {"left": "niet-numeriek", "top": 0, "right": 60, "bottom": 40},
    )

    assert cropped is not None
    assert cropped.size == (120, 60)


def test_safe_make_logo_photo_returns_none_when_photoimage_fails():
    img = Image.new("RGBA", (120, 60), "red")

    class BrokenImageTk:
        @staticmethod
        def PhotoImage(_img):
            raise RuntimeError("kan image niet renderen")

    assert _safe_make_logo_photo(img, BrokenImageTk, None, (80, 40)) is None


def test_client_logo_order_preview_uses_shared_doc_number_formatter():
    source = pathlib.Path("gui.py").read_text(encoding="utf-8")
    mod = ast.parse(source)
    start = next(
        node for node in mod.body if isinstance(node, ast.FunctionDef) and node.name == "start_gui"
    )
    client_cls = next(
        node
        for node in start.body
        if isinstance(node, ast.ClassDef) and node.name == "ClientsManagerFrame"
    )
    open_dialog = next(
        node
        for node in client_cls.body
        if isinstance(node, ast.FunctionDef) and node.name == "_open_edit_dialog"
    )

    assert not any(
        isinstance(node, ast.Attribute) and node.attr == "_format_document_display_number"
        for node in ast.walk(open_dialog)
    )
    assert any(
        isinstance(node, ast.Name) and node.id == "format_document_number_for_display"
        for node in ast.walk(open_dialog)
    )


def test_client_dialog_includes_website_field():
    source = pathlib.Path("gui.py").read_text(encoding="utf-8")
    assert '("Website", "website")' in source


@pytest.mark.skipif(not REPORTLAB_OK, reason="ReportLab niet beschikbaar")
def test_generate_pdf_order_includes_logo(tmp_path):
    logo_path = tmp_path / "logo.png"
    Image.new("RGB", (400, 200), "red").save(logo_path)

    pdf_path = tmp_path / "order.pdf"
    company = {
        "name": "ACME",
        "address": "Example Street 1",
        "vat": "BE0123456789",
        "email": "info@example.com",
        "logo_path": str(logo_path),
        "logo_crop": {"left": 50, "top": 0, "right": 350, "bottom": 200},
    }
    supplier = Supplier(supplier="Supplier BV")
    items = [
        {
            "PartNumber": "PN-1",
            "Description": "Onderdeel",
            "Materiaal": "",
            "Aantal": 1,
            "Oppervlakte": "",
            "Gewicht": "",
        }
    ]

    generate_pdf_order_platypus(
        str(pdf_path),
        company,
        supplier,
        production="PROD-1",
        items=items,
    )

    from PyPDF2 import PdfReader

    reader = PdfReader(str(pdf_path))
    page = reader.pages[0]
    resources = page.get("/Resources")
    assert resources is not None
    xobjects = resources.get("/XObject")
    assert xobjects is not None
    dims = []
    for obj in xobjects.values():
        xobj = obj.get_object()
        if xobj.get("/Subtype") == "/Image":
            dims.append((int(xobj.get("/Width")), int(xobj.get("/Height"))))
    assert dims
    assert (300, 200) in dims


@pytest.mark.skipif(not REPORTLAB_OK, reason="ReportLab niet beschikbaar")
def test_generate_pdf_order_resolves_runtime_relative_logo_path(tmp_path, monkeypatch):
    logo_dir = tmp_path / "client_logos"
    logo_dir.mkdir()
    logo_path = logo_dir / "runtime-logo.png"
    Image.new("RGB", (320, 120), "navy").save(logo_path)

    monkeypatch.setattr(app_paths, "bundle_root", lambda: tmp_path)
    monkeypatch.setattr(app_paths, "storage_dir", lambda: tmp_path)

    workdir = tmp_path / "other-working-dir"
    workdir.mkdir()
    monkeypatch.chdir(workdir)

    pdf_path = workdir / "order.pdf"
    company = {
        "name": "ACME",
        "address": "Example Street 1",
        "vat": "BE0123456789",
        "email": "info@example.com",
        "logo_path": "client_logos/runtime-logo.png",
        "logo_crop": None,
    }
    supplier = Supplier(supplier="Supplier BV")
    items = [
        {
            "PartNumber": "PN-1",
            "Description": "Onderdeel",
            "Materiaal": "",
            "Aantal": 1,
            "Oppervlakte": "",
            "Gewicht": "",
        }
    ]

    generate_pdf_order_platypus(
        str(pdf_path),
        company,
        supplier,
        production="PROD-1",
        items=items,
    )

    from PyPDF2 import PdfReader

    reader = PdfReader(str(pdf_path))
    page = reader.pages[0]
    resources = page.get("/Resources")
    assert resources is not None
    xobjects = resources.get("/XObject")
    assert xobjects is not None
    dims = []
    for obj in xobjects.values():
        xobj = obj.get_object()
        if xobj.get("/Subtype") == "/Image":
            dims.append((int(xobj.get("/Width")), int(xobj.get("/Height"))))
    assert dims
    assert (320, 120) in dims


@pytest.mark.skipif(not REPORTLAB_OK, reason="ReportLab niet beschikbaar")
def test_generate_pdf_footer_note_only_for_bestelbon(tmp_path):
    supplier = Supplier(supplier="Supplier BV")
    items = [
        {
            "PartNumber": "PN-1",
            "Description": "Onderdeel",
            "Materiaal": "",
            "Aantal": 1,
            "Oppervlakte": "",
            "Gewicht": "",
        }
    ]

    bestelbon_pdf = tmp_path / "bestelbon.pdf"
    standaard_pdf = tmp_path / "standaard-bon.pdf"

    generate_pdf_order_platypus(
        str(bestelbon_pdf),
        {},
        supplier,
        production="PROD-1",
        items=items,
        doc_type="Bestelbon",
        footer_note=DEFAULT_FOOTER_NOTE,
    )
    generate_pdf_order_platypus(
        str(standaard_pdf),
        {},
        supplier,
        production="PROD-1",
        items=items,
        doc_type="Standaard bon",
        footer_note=DEFAULT_FOOTER_NOTE,
    )

    from PyPDF2 import PdfReader

    bestelbon_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(str(bestelbon_pdf)).pages
    )
    standaard_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(str(standaard_pdf)).pages
    )

    assert "Gelieve afwijkingen schriftelijk te bevestigen." in bestelbon_text
    assert "Gelieve afwijkingen schriftelijk te bevestigen." not in standaard_text


@pytest.mark.skipif(not REPORTLAB_OK, reason="ReportLab niet beschikbaar")
def test_generate_pdf_order_includes_client_website(tmp_path):
    supplier = Supplier(supplier="Supplier BV")
    items = [
        {
            "PartNumber": "PN-1",
            "Description": "Onderdeel",
            "Materiaal": "",
            "Aantal": 1,
            "Oppervlakte": "",
            "Gewicht": "",
        }
    ]
    pdf_path = tmp_path / "website.pdf"

    generate_pdf_order_platypus(
        str(pdf_path),
        {
            "name": "ACME",
            "address": "Example Street 1",
            "vat": "BE0123456789",
            "email": "info@example.com",
            "website": "https://acme.example",
        },
        supplier,
        production="PROD-1",
        items=items,
    )

    from PyPDF2 import PdfReader

    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)
    assert "Website: https://acme.example" in text


def test_write_order_excel_includes_client_website(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl

    path = tmp_path / "website.xlsx"
    write_order_excel(
        str(path),
        [
            {
                "PartNumber": "PN-1",
                "Description": "Onderdeel",
                "Materiaal": "",
                "Aantal": 1,
                "Oppervlakte": "",
                "Gewicht": "",
            }
        ],
        company_info={
            "name": "ACME",
            "address": "Example Street 1",
            "vat": "BE0123456789",
            "email": "info@example.com",
            "website": "https://acme.example",
        },
        supplier=Supplier(supplier="Supplier BV"),
        context_label="Laser",
        context_kind="productie",
    )

    ws = openpyxl.load_workbook(path).active
    col_a = [ws[f"A{i}"].value for i in range(1, 20)]
    assert "Website" in col_a
    row_idx = col_a.index("Website") + 1
    assert ws[f"B{row_idx}"].value == "https://acme.example"
