import datetime
import pandas as pd
import openpyxl
import pytest
from PyPDF2 import PdfReader

import cli
from cli import build_parser, cli_copy_per_prod
from models import Supplier
from suppliers_db import SuppliersDB
from clients_db import ClientsDB
from delivery_addresses_db import DeliveryAddressesDB


def test_project_info_in_documents(tmp_path, monkeypatch):
    reportlab = pytest.importorskip("reportlab")

    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}
    ])
    dst = tmp_path / "dst"
    dst.mkdir()

    parser = build_parser()
    args = parser.parse_args([
        "copy-per-prod",
        "--source",
        str(src),
        "--dest",
        str(dst),
        "--bom",
        str(tmp_path / "bom.xlsx"),
        "--exts",
        "pdf",
        "--project-number",
        "PRJ123",
        "--project-name",
        "New Project",
    ])

    monkeypatch.setattr(cli, "load_bom", lambda path: bom_df)
    monkeypatch.setattr(SuppliersDB, "load", classmethod(lambda cls, path: db))
    monkeypatch.setattr(ClientsDB, "load", classmethod(lambda cls, path: ClientsDB([])))
    monkeypatch.setattr(
        DeliveryAddressesDB, "load", classmethod(lambda cls, path: DeliveryAddressesDB([]))
    )

    cli_copy_per_prod(args)

    prod_folder = dst / "Laser"
    today = datetime.date.today().strftime("%Y-%m-%d")

    xlsx_path = prod_folder / f"Bestelbon_Laser_{today}.xlsx"
    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws["A1"].value == "Datum"
    assert ws["B1"].value == today
    assert ws["A2"].value == "Projectnr."
    assert ws["B2"].value == "PRJ123"
    assert ws["A3"].value == "Projectnaam"
    assert ws["B3"].value == "New Project"

    pdf_path = prod_folder / f"Bestelbon_Laser_{today}.pdf"
    assert pdf_path.exists()
    reader = PdfReader(pdf_path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = text.splitlines()
    assert lines.index(f"Datum: {today}") < lines.index("Projectnr.: PRJ123") < lines.index("Projectnaam: New Project")
