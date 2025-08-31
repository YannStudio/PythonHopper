from __future__ import annotations

import os
import tempfile

import pandas as pd
import openpyxl

from models import Supplier, Client
from suppliers_db import SuppliersDB
from clients_db import ClientsDB
from bom import load_bom
from orders import copy_per_production_and_orders, DEFAULT_FOOTER_NOTE


def run_tests() -> int:
    print("Running self-tests...")
    db = SuppliersDB()
    db.upsert(Supplier.from_any({
        "supplier": "ACME",
        "description": "Snijwerk",
        "btw": "BE123",
        "adress_1": "Teststraat 1 bus 2",
        "address_2": "BE-2000 Antwerpen",
        "land": "BE",
        "e-mail sales": "x@y.z",
        "tel. sales": "+32 123",
    }))
    assert db.suppliers and db.suppliers[0].adres_2 == "BE-2000 Antwerpen"
    db.toggle_fav("ACME")
    assert db.suppliers[0].favorite
    db.set_default("Laser", "ACME")
    assert db.get_default("Laser") == "ACME"

    cdb = ClientsDB()
    cdb.upsert(Client.from_any({
        "name": "TestClient",
        "address": "Straat 1, 1000 Brussel",
        "vat": "BE000",
        "email": "test@example.com",
    }))
    client = cdb.clients[0]

    with tempfile.TemporaryDirectory() as td:
        src = os.path.join(td, "src")
        dst = os.path.join(td, "dst")
        os.makedirs(src)
        os.makedirs(dst)
        open(os.path.join(src, "PN1.pdf"), "wb").write(b"%PDF-1.4")
        open(os.path.join(src, "PN1.stp"), "wb").write(b"step")
        bom = os.path.join(td, "bom.xlsx")
        df = pd.DataFrame([
            {
                "PartNumber": "PN1-THIS-IS-A-VERY-LONG-CODE-OVER-25CHARS",
                "Description": "Lange omschrijving die netjes moet wrappen.",
                "Production": "Laser",
                "Aantal": 2,
                "Materiaal": "S235JR",
                "Oppervlakte (m²)": "1,23",
                "Gewicht (kg)": "4,56",
            },
            {
                "PartNumber": "PN2",
                "Description": "Geen files",
                "Production": "Laser",
                "Aantal": 1000,
                "Material": "Alu 5754",
                "Area": "0.50",
                "Weight": "1.00",
            },
        ])
        df.to_excel(bom, index=False, engine="openpyxl")
        ldf = load_bom(bom)
        assert ldf["Aantal"].max() <= 999  # capped
        cnt, chosen = copy_per_production_and_orders(
            src,
            dst,
            ldf,
            [".pdf", ".stp"],
            db,
            {},
            {},
            {},
            True,
            client=client,
            delivery_map={},
            footer_note=DEFAULT_FOOTER_NOTE,
        )
        assert cnt == 2
        assert chosen.get("Laser") == "ACME"
        prod_folder = os.path.join(dst, "Laser")
        assert os.path.exists(os.path.join(prod_folder, "PN1.pdf"))
        assert os.path.exists(os.path.join(prod_folder, "PN1.stp"))
        xlsx = [f for f in os.listdir(prod_folder) if f.lower().endswith(".xlsx")]
        assert xlsx, "Excel bestelbon niet aangemaakt"
        wb = openpyxl.load_workbook(os.path.join(prod_folder, xlsx[0]))
        ws = wb.active
        assert ws["A1"].value == "Bedrijf" and ws["B1"].value == client.name
        assert ws["A6"].value == "Leverancier" and ws["B6"].value == "ACME"
        assert ws["A7"].value == "Adres"
        assert (
            ws["B7"].value == "Teststraat 1 bus 2, BE-2000 Antwerpen, BE"
        )
        assert ws["A8"].value == "BTW" and ws["B8"].value == "BE123"
        assert ws["A9"].value == "E-mail" and ws["B9"].value == "x@y.z"
        assert ws["A10"].value == "Tel" and ws["B10"].value == "+32 123"
        pdfs = [f for f in os.listdir(prod_folder) if f.lower().endswith(".pdf")]
        assert pdfs, "PDF bestelbon niet aangemaakt"
    print("All tests passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(run_tests())
