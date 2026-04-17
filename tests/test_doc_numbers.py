import datetime
import os
import zipfile

import openpyxl
import pandas as pd
import pytest
pytest.importorskip("openpyxl")
from PyPDF2 import PdfReader

from models import Supplier
from suppliers_db import SuppliersDB
from orders import (
    copy_per_production_and_orders,
    _prefix_for_doc_type,
    _normalize_doc_number,
    build_document_export_basename,
    format_document_number_for_display,
)


def test_doc_number_in_name_and_header(tmp_path):
    reportlab = pytest.importorskip("reportlab")

    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}
    ])
    dst = tmp_path / "dst"
    dst.mkdir()

    doc_num_map = {"Laser": "BB-123/45"}
    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        doc_num_map,
        False,
        client=None,
        delivery_map={},
    )

    prod_folder = dst / "Laser"
    today = datetime.date.today().strftime("%Y-%m-%d")

    xlsx_path = prod_folder / f"Bestelbon_BB-123_45_Laser_{today}.xlsx"
    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws["A1"].value == "Nummer"
    assert ws["B1"].value == "BB-123/45"
    assert ws["A2"].value == "Datum"
    assert ws["B2"].value == today

    pdf_path = prod_folder / f"Bestelbon_BB-123_45_Laser_{today}.pdf"
    assert pdf_path.exists()
    reader = PdfReader(pdf_path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = text.splitlines()
    assert f"Nummer: BB-123/45" in text
    assert f"Datum: {today}" in text
    assert "BB-123" not in lines[0]


def test_prefix_helper():
    assert _prefix_for_doc_type("Bestelbon") == "BB-"
    assert _prefix_for_doc_type("Standaard bon") == "BOM-"
    assert _prefix_for_doc_type("Offerteaanvraag") == "OFF-"
    assert _prefix_for_doc_type("Onbekend") == ""


def test_normalize_doc_number_removes_duplicate_prefix():
    assert _normalize_doc_number("BB-BB64646", "Bestelbon") == "BB-64646"
    assert _normalize_doc_number("BB64646", "Bestelbon") == "BB-64646"
    assert _normalize_doc_number("64646", "Bestelbon") == "64646"
    assert _normalize_doc_number("BB-", "Bestelbon") == ""
    assert _normalize_doc_number(None, "Bestelbon") == ""


def test_build_document_export_basename_profiles():
    export_date = "2026-04-15"

    assert (
        build_document_export_basename(
            "Bestelbon",
            "BB-123/45",
            "Laser",
            export_date,
        )
        == "Bestelbon_BB-123_45_Laser_2026-04-15"
    )
    assert (
        build_document_export_basename(
            "Bestelbon",
            "BB-123/45",
            "Laser",
            export_date,
            profile="short",
        )
        == "BB-123_45"
    )
    assert (
        build_document_export_basename(
            "Bestelbon",
            "BB-123/45",
            "Laser",
            export_date,
            profile="compact",
        )
        == "BB123_45"
    )
    assert (
        build_document_export_basename(
            "Bestelbon",
            "BB-123/45",
            "Laser",
            export_date,
            profile="custom",
            show_doc_type=False,
            show_doc_number=True,
            show_context=False,
            show_date=False,
            compact_doc_number=True,
            separator="none",
        )
        == "BB123_45"
    )
    assert (
        build_document_export_basename(
            "Bestelbon",
            "",
            "Laser",
            export_date,
            profile="short",
        )
        == "Bestelbon_Laser_2026-04-15"
    )
    assert (
        build_document_export_basename(
            "Bestelbon",
            "BB-",
            "Laser",
            export_date,
        )
        == "Bestelbon_Laser_2026-04-15"
    )


def test_format_document_number_for_display():
    assert (
        format_document_number_for_display("BB-5487", "Bestelbon", compact=False)
        == "BB-5487"
    )
    assert (
        format_document_number_for_display("BB-5487", "Bestelbon", compact=True)
        == "BB5487"
    )
    assert (
        format_document_number_for_display("OFF-12/34", "Offerteaanvraag", compact=True)
        == "OFF12/34"
    )


def test_offerte_prefix_in_output(tmp_path):
    reportlab = pytest.importorskip("reportlab")

    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}
    ])
    dst = tmp_path / "dst"
    dst.mkdir()

    doc_num_map = {"Laser": "OFF-42"}
    doc_type_map = {"Laser": "Offerteaanvraag"}
    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        doc_type_map,
        doc_num_map,
        False,
        client=None,
        delivery_map={},
    )

    prod_folder = dst / "Laser"
    today = datetime.date.today().strftime("%Y-%m-%d")

    xlsx_path = prod_folder / f"Offerteaanvraag_OFF-42_Laser_{today}.xlsx"
    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws["A1"].value == "Nummer"
    assert ws["B1"].value == "OFF-42"
    assert ws["A2"].value == "Datum"
    assert ws["B2"].value == today

    pdf_path = prod_folder / f"Offerteaanvraag_OFF-42_Laser_{today}.pdf"
    assert pdf_path.exists()
    reader = PdfReader(pdf_path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = text.splitlines()
    assert f"Nummer: OFF-42" in text
    assert f"Datum: {today}" in text
    assert "OFF-42" not in lines[0]


def test_missing_doc_number_omits_prefix_and_header(tmp_path):
    reportlab = pytest.importorskip("reportlab")

    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}
    ])
    dst = tmp_path / "dst"
    dst.mkdir()

    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        {},
        False,
        client=None,
        delivery_map={},
    )

    prod_folder = dst / "Laser"
    today = datetime.date.today().strftime("%Y-%m-%d")

    xlsx_path = prod_folder / f"Bestelbon_Laser_{today}.xlsx"
    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    # Without document number the header starts with the date
    assert ws["A1"].value == "Datum"
    assert ws["B1"].value == today
    rows = list(ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True))
    assert all(r[0] != "Nummer" for r in rows)

    pdf_path = prod_folder / f"Bestelbon_Laser_{today}.pdf"
    assert pdf_path.exists()
    reader = PdfReader(pdf_path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Nummer:" not in text
    assert f"Datum: {today}" in text


@pytest.mark.parametrize(
    ("profile", "expected_base"),
    [
        ("short", "BB-123_45"),
        ("compact", "BB123_45"),
    ],
)
def test_document_filename_profile_controls_export_name(
    tmp_path,
    profile,
    expected_base,
):
    pytest.importorskip("reportlab")

    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))

    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")

    bom_df = pd.DataFrame(
        [{"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}]
    )

    dst = tmp_path / f"dst_{profile}"
    dst.mkdir()

    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        {"Laser": "BB-123/45"},
        False,
        client=None,
        delivery_map={},
        document_filename_profile=profile,
    )

    prod_folder = dst / "Laser"
    assert (prod_folder / f"{expected_base}.xlsx").exists()
    assert (prod_folder / f"{expected_base}.pdf").exists()


def test_document_display_compact_doc_number_changes_pdf_and_excel_header(tmp_path):
    pytest.importorskip("reportlab")

    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))

    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")

    bom_df = pd.DataFrame(
        [{"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}]
    )

    dst = tmp_path / "dst_display"
    dst.mkdir()

    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        {"Laser": "BB-5487"},
        False,
        client=None,
        delivery_map={},
        document_display_compact_doc_number=True,
    )

    today = datetime.date.today().strftime("%Y-%m-%d")
    prod_folder = dst / "Laser"

    xlsx_path = prod_folder / f"Bestelbon_BB-5487_Laser_{today}.xlsx"
    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws["A1"].value == "Nummer"
    assert ws["B1"].value == "BB5487"

    pdf_path = prod_folder / f"Bestelbon_BB-5487_Laser_{today}.pdf"
    assert pdf_path.exists()
    reader = PdfReader(pdf_path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Nummer: BB5487" in text
    assert "Nummer: BB-5487" not in text


def test_doc_number_applied_to_zip_filename(tmp_path):
    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))

    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")

    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}
    ])

    dst = tmp_path / "dst_zip"
    dst.mkdir()

    doc_num_map = {"Laser": "BB-123/45"}

    cnt, _ = copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        doc_num_map,
        False,
        client=None,
        delivery_map={},
        zip_parts=True,
    )

    assert cnt == 1

    prod_folder = dst / "Laser"
    zip_files = sorted(prod_folder.glob("Laser*.zip"))
    assert len(zip_files) == 1
    zip_path = zip_files[0]
    assert zip_path.name == "Laser_BB-123_45.zip"

    with zipfile.ZipFile(zip_path) as zf:
        assert "PN1.pdf" in zf.namelist()
        info = zf.getinfo("PN1.pdf")
        if getattr(zipfile, "zlib", None):
            assert info.compress_type == zipfile.ZIP_DEFLATED
        else:
            assert info.compress_type == zipfile.ZIP_STORED

