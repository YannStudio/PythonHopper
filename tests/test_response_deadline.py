import os
import pandas as pd
import openpyxl
from models import Supplier
from suppliers_db import SuppliersDB
from orders import copy_per_production_and_orders


def test_response_deadline_in_excel(tmp_path):
    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    dst = tmp_path / "dst"
    src.mkdir(); dst.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1},
    ])
    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        False,
        doc_type="offerteaanvraag",
        response_deadline="2024-12-31",
    )
    prod_folder = dst / "Laser"
    xlsx = [f for f in os.listdir(prod_folder) if f.lower().endswith(".xlsx")][0]
    wb = openpyxl.load_workbook(prod_folder / xlsx)
    ws = wb.active
    assert ws["A1"].value == "Antwoord tegen"
    assert ws["B1"].value == "2024-12-31"
