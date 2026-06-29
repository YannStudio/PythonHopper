import pandas as pd
from openpyxl import load_workbook

from models import Supplier
from orders import (
    copy_per_production_and_orders,
    make_spare_part_default_key,
    make_spare_part_selection_key,
    parse_selection_key,
)
from spare_parts import (
    SPARE_PARTS_CUSTOM_SOURCE,
    SPARE_PARTS_FULL_LIST_DISPLAY_LABEL,
    SPARE_PARTS_FULL_LIST_DOCUMENT_LABEL,
    SPARE_PARTS_FULL_LIST_KEY,
    SPARE_PARTS_UNASSIGNED_KEY,
    build_spare_part_groups,
    collect_spare_part_groups,
    collect_spare_part_items,
    is_spare_parts_production,
    make_custom_spare_part_group_key,
    match_spare_part_group_overrides,
    spare_part_document_readiness,
    spare_part_identity_match_key,
    summarize_spare_part_warnings,
)
from suppliers_db import SuppliersDB


def test_spare_parts_detection_is_tolerant():
    assert is_spare_parts_production("Spare Parts") is True
    assert is_spare_parts_production(" spare-parts ") is True
    assert is_spare_parts_production("Laser") is False


def test_collect_spare_parts_reads_supplier_and_manufacturer_fields():
    df = pd.DataFrame(
        [
            {
                "PartNumber": "458",
                "Description": "Hangkastje",
                "Production": "Spare Parts",
                "Aantal": 2,
                "Supplier": "Herbaroof",
                "Supplier code": "ND SM-25",
                "Manufacturer": "Herbaroof",
                "Manufacturer code": "ND SM-25",
            },
            {
                "PartNumber": "999",
                "Production": "Laser",
                "Supplier": "Ignored",
            },
        ]
    )

    items = collect_spare_part_items(df)

    assert len(items) == 1
    assert items[0].part_number == "458"
    assert items[0].supplier == "Herbaroof"
    assert items[0].manufacturer_code == "ND SM-25"
    assert items[0].status == "OK"


def test_spare_part_identity_key_keeps_empty_field_positions():
    item = collect_spare_part_items(
        [
            {
                "PartNumber": "PN1",
                "Production": "Spare Parts",
                "Supplier code": "RS-204",
            }
        ]
    )[0]

    assert item.identity_key == "sparepart:0|PN1|||RS-204"
    assert spare_part_identity_match_key(item.identity_key) == item.match_key


def test_spare_part_groups_prefer_supplier_and_keep_full_list():
    items = collect_spare_part_items(
        [
            {
                "PartNumber": "A",
                "Production": "Spare Parts",
                "Supplier": "Herbaroof",
                "Supplier code": "H-1",
                "Manufacturer": "Maker A",
                "Manufacturer code": "M-1",
            },
            {
                "PartNumber": "B",
                "Production": "Spare Parts",
                "Manufacturer": "Maker B",
                "Manufacturer code": "M-2",
            },
        ]
    )

    groups = build_spare_part_groups(items)

    assert groups[0].key == SPARE_PARTS_FULL_LIST_KEY
    assert groups[0].is_full_list is True
    assert groups[0].default_doc_type == "Standaard bon"
    assert groups[0].display_label == SPARE_PARTS_FULL_LIST_DISPLAY_LABEL
    assert groups[0].document_label == SPARE_PARTS_FULL_LIST_DOCUMENT_LABEL
    assert groups[0].to_mapping()["document_label"] == SPARE_PARTS_FULL_LIST_DOCUMENT_LABEL
    assert groups[0].item_count == 2
    route_groups = {group.key: group for group in groups[1:]}
    assert route_groups["supplier--herbaroof"].default_supplier == "Herbaroof"
    assert route_groups["manufacturer--maker-b"].default_supplier == ""


def test_spare_part_groups_track_unassigned_and_missing_codes():
    items = collect_spare_part_items(
        [
            {
                "PartNumber": "A",
                "Production": "Spare Parts",
                "Supplier": "Electro",
            },
            {"PartNumber": "B", "Production": "Spare Parts"},
        ]
    )

    groups = build_spare_part_groups(items)
    by_key = {group.key: group for group in groups}

    assert by_key["supplier--electro"].missing_count == 1
    assert by_key[SPARE_PARTS_UNASSIGNED_KEY].label == "Nog toe te wijzen"
    assert by_key[SPARE_PARTS_UNASSIGNED_KEY].missing_count == 1


def test_spare_part_warnings_summarize_open_data_and_supplier_gaps():
    items = collect_spare_part_items(
        [
            {
                "PartNumber": "A",
                "Production": "Spare Parts",
                "Supplier": "Electro",
            },
            {
                "PartNumber": "B",
                "Production": "Spare Parts",
                "Manufacturer": "Maker",
                "Manufacturer code": "M-1",
            },
            {"PartNumber": "C", "Production": "Spare Parts"},
        ]
    )
    groups = build_spare_part_groups(items)

    warnings = summarize_spare_part_warnings(groups)

    assert "1 nog toe te wijzen" in warnings
    assert "1 zonder Supplier/Manufacturer" in warnings
    assert "2 zonder Supplier code/Manufacturer code" in warnings
    assert "2 groep(en) zonder standaardleverancier" in warnings


def test_spare_part_document_readiness_reports_open_checks():
    items = collect_spare_part_items(
        [
            {
                "PartNumber": "A",
                "Production": "Spare Parts",
                "Supplier": "Electro",
            },
            {
                "PartNumber": "B",
                "Production": "Spare Parts",
                "Manufacturer": "Maker",
                "Manufacturer code": "M-1",
            },
            {"PartNumber": "C", "Production": "Spare Parts"},
        ]
    )

    checks = {
        check["label"]: check for check in spare_part_document_readiness(
            build_spare_part_groups(items)
        )
    }

    assert checks["Klaarleglijst"]["ok"] is True
    assert checks["Groepen"]["ok"] is False
    assert checks["Groepen"]["detail"] == "1 open"
    assert checks["Leverancier/fabrikant"]["ok"] is False
    assert checks["Codes"]["ok"] is False
    assert checks["Groepsleverancier"]["ok"] is False


def test_spare_part_document_readiness_accepts_manual_groups():
    df = pd.DataFrame(
        [
            {
                "PartNumber": "A",
                "Production": "Spare Parts",
                "Supplier": "RS",
                "Supplier code": "RS-1",
            },
            {
                "PartNumber": "B",
                "Production": "Spare Parts",
                "Supplier": "RS",
                "Supplier code": "RS-2",
            },
        ]
    )
    items = collect_spare_part_items(df)
    groups = build_spare_part_groups(
        items,
        group_overrides={
            items[0].identity_key: "Electro",
            items[1].identity_key: "Electro",
        },
    )

    checks = {
        check["label"]: check for check in spare_part_document_readiness(groups)
    }

    assert checks["Klaarleglijst"]["ok"] is True
    assert checks["Groepen"]["ok"] is True
    assert checks["Leverancier/fabrikant"]["ok"] is True
    assert checks["Codes"]["ok"] is True
    assert checks["Groepsleverancier"]["ok"] is False
    assert checks["Groepsleverancier"]["detail"] == "1 kiezen in documentflow"


def test_spare_part_groups_accept_manual_overrides_without_mutating_bom():
    df = pd.DataFrame(
        [
            {
                "PartNumber": "A",
                "Production": "Spare Parts",
                "Supplier": "Herbaroof",
                "Supplier code": "H-1",
            },
            {
                "PartNumber": "B",
                "Production": "Spare Parts",
                "Manufacturer": "Maker B",
                "Manufacturer code": "M-2",
            },
        ]
    )
    items = collect_spare_part_items(df)
    custom_key = make_custom_spare_part_group_key("Electro")
    overrides = {
        items[0].identity_key: "Electro",
        items[1].identity_key: "Nog toe te wijzen",
    }

    groups = build_spare_part_groups(items, group_overrides=overrides)
    by_key = {group.key: group for group in groups}

    assert by_key[custom_key].route_source == SPARE_PARTS_CUSTOM_SOURCE
    assert by_key[custom_key].label == "Electro"
    assert by_key[custom_key].items == [items[0]]
    assert by_key[SPARE_PARTS_UNASSIGNED_KEY].items == [items[1]]
    full_items = by_key[SPARE_PARTS_FULL_LIST_KEY].to_mapping()["items"]
    assert [item["Bestelgroep"] for item in full_items] == [
        "Electro",
        "Nog toe te wijzen",
    ]
    assert list(df["Production"]) == ["Spare Parts", "Spare Parts"]


def test_spare_part_override_matching_accepts_shifted_row_index():
    old_items = collect_spare_part_items(
        pd.DataFrame(
            [
                {
                    "PartNumber": "PN1",
                    "Description": "Sensor kabel",
                    "Production": "Spare Parts",
                    "Supplier code": "RS-204",
                    "Manufacturer code": "M-42",
                }
            ],
            index=[10],
        )
    )
    new_items = collect_spare_part_items(
        pd.DataFrame(
            [
                {
                    "PartNumber": "PN1",
                    "Description": "Sensor kabel",
                    "Production": "Spare Parts",
                    "Supplier code": "RS-204",
                    "Manufacturer code": "M-42",
                }
            ],
            index=[22],
        )
    )

    matched = match_spare_part_group_overrides(
        new_items,
        {old_items[0].identity_key: "Electro"},
    )

    assert spare_part_identity_match_key(old_items[0].identity_key)
    assert matched == {new_items[0].identity_key: "Electro"}


def test_spare_part_override_matching_accepts_legacy_keys_with_empty_fields():
    new_items = collect_spare_part_items(
        pd.DataFrame(
            [
                {
                    "PartNumber": "PN1",
                    "Production": "Spare Parts",
                    "Supplier code": "RS-204",
                    "Manufacturer code": "M-42",
                }
            ],
            index=[22],
        )
    )
    legacy_key = "sparepart:10|PN1|M-42|RS-204"

    matched = match_spare_part_group_overrides(new_items, {legacy_key: "Electro"})

    assert matched == {new_items[0].identity_key: "Electro"}


def test_spare_part_override_matching_skips_ambiguous_shifted_rows():
    old_items = collect_spare_part_items(
        pd.DataFrame(
            [
                {
                    "PartNumber": "PN1",
                    "Description": "Sensor kabel",
                    "Production": "Spare Parts",
                    "Supplier code": "RS-204",
                    "Manufacturer code": "M-42",
                }
            ],
            index=[10],
        )
    )
    new_items = collect_spare_part_items(
        pd.DataFrame(
            [
                {
                    "PartNumber": "PN1",
                    "Description": "Sensor kabel",
                    "Production": "Spare Parts",
                    "Supplier code": "RS-204",
                    "Manufacturer code": "M-42",
                },
                {
                    "PartNumber": "PN1",
                    "Description": "Sensor kabel",
                    "Production": "Spare Parts",
                    "Supplier code": "RS-204",
                    "Manufacturer code": "M-42",
                },
            ],
            index=[22, 23],
        )
    )

    matched = match_spare_part_group_overrides(
        new_items,
        {old_items[0].identity_key: "Electro"},
    )

    assert matched == {}


def test_spare_part_selection_key_roundtrip():
    key = make_spare_part_selection_key("supplier--herbaroof")

    assert parse_selection_key(key) == ("sparepart", "supplier--herbaroof")
    assert make_spare_part_default_key("supplier--herbaroof").endswith(
        "supplier--herbaroof"
    )


def test_spare_part_groups_export_full_list_and_supplier_order(tmp_path):
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("one", encoding="utf-8")
    (src / "PN2.pdf").write_text("two", encoding="utf-8")
    db = SuppliersDB(storage_path=tmp_path / "suppliers_db.json")
    db.upsert(Supplier.from_any({"supplier": "Herbaroof"}))
    bom_df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Description": "Spare rail",
                "Production": "Spare Parts",
                "Aantal": 2,
                "Supplier": "Herbaroof",
                "Supplier code": "ND SM-25",
                "Manufacturer": "Herbaroof",
                "Manufacturer code": "MF-25",
            },
            {
                "PartNumber": "PN2",
                "Description": "Loose cover",
                "Production": "Spare Parts",
                "Aantal": 1,
                "Supplier": "Herbaroof",
                "Supplier code": "ND SM-30",
                "Manufacturer": "Maker",
                "Manufacturer code": "MF-30",
            },
        ]
    )
    groups = [group.to_mapping() for group in collect_spare_part_groups(bom_df)]

    generated_documents = []

    copied, chosen = copy_per_production_and_orders(
        str(src),
        str(dest),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        {},
        True,
        spare_part_groups=groups,
        spare_part_override_map={
            SPARE_PARTS_FULL_LIST_KEY: "Herbaroof",
            "supplier--herbaroof": "Herbaroof",
        },
        spare_part_doc_type_map={SPARE_PARTS_FULL_LIST_KEY: "Bestelbon"},
        generated_documents=generated_documents,
    )

    assert copied == 2
    assert chosen[make_spare_part_selection_key(SPARE_PARTS_FULL_LIST_KEY)] == ""
    assert chosen[make_spare_part_selection_key("supplier--herbaroof")] == "Herbaroof"
    assert (
        db.get_default(make_spare_part_default_key("supplier--herbaroof"))
        == "Herbaroof"
    )
    full_docs = list((dest / "Spare Parts").glob("Standaard*Spare*Parts*.xlsx"))
    supplier_docs = list(
        (dest / "Spare Parts-Herbaroof").glob("Bestelbon*Spare*Parts-Herbaroof*.xlsx")
    )
    assert full_docs
    assert supplier_docs
    assert any("Spare Parts klaarleglijst" in path.name for path in full_docs)
    full_records = [
        record
        for record in generated_documents
        if record.get("selection_key") == make_spare_part_selection_key(SPARE_PARTS_FULL_LIST_KEY)
    ]
    assert full_records
    assert {record.get("context_label") for record in full_records} == {"Klaarleglijst"}
    assert {record.get("doc_type") for record in full_records} == {"Standaard bon"}
    assert all(not record.get("supplier") for record in full_records)

    workbook = load_workbook(full_docs[0])
    values = [
        value
        for row in workbook.active.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    assert "Supplier code" in values
    assert "Manufacturer code" in values
    assert "Bestelgroep" not in values
    assert "Status" not in values
    assert "ND SM-25" in values
    assert "MF-30" in values


def test_spare_part_manual_group_is_used_for_export_without_mutating_bom(tmp_path):
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("one", encoding="utf-8")
    db = SuppliersDB(storage_path=tmp_path / "suppliers_db.json")
    db.upsert(Supplier.from_any({"supplier": "RS Components"}))
    bom_df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Description": "Spare cable",
                "Production": "Spare Parts",
                "Aantal": 3,
                "Supplier": "RS Components",
                "Supplier code": "RS-1",
                "Manufacturer": "Maker",
                "Manufacturer code": "M-1",
            }
        ]
    )
    items = collect_spare_part_items(bom_df)
    custom_group_key = make_custom_spare_part_group_key("Electro")
    groups = [
        group.to_mapping()
        for group in collect_spare_part_groups(
            bom_df,
            group_overrides={items[0].identity_key: "Electro"},
        )
    ]

    _copied, chosen = copy_per_production_and_orders(
        str(src),
        str(dest),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        {},
        True,
        spare_part_groups=groups,
        spare_part_override_map={custom_group_key: "RS Components"},
    )

    assert chosen[make_spare_part_selection_key(custom_group_key)] == "RS Components"
    assert list(bom_df["Production"]) == ["Spare Parts"]
    custom_docs = list(
        (dest / "Spare Parts-Electro").glob("Bestelbon*Spare*Parts-Electro*.xlsx")
    )
    assert custom_docs
    workbook = load_workbook(custom_docs[0])
    values = [
        value
        for row in workbook.active.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    assert "Bestelgroep" in values
    assert "Electro" in values
