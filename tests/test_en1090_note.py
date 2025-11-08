import pandas as pd
from openpyxl import load_workbook

from en1090 import EN1090_NOTE_TEXT
from models import Supplier
from orders import copy_per_production_and_orders
from suppliers_db import SuppliersDB


def _make_db() -> SuppliersDB:
    return SuppliersDB([Supplier.from_any({"supplier": "ACME"})])


def test_en1090_note_present_in_excel(tmp_path):
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()

    # Provide a matching export file to exercise the copy logic.
    export_file = src / "123.pdf"
    export_file.write_text("dummy")

    bom_df = pd.DataFrame(
        [
            {
                "PartNumber": "123",
                "Description": "Test onderdeel",
                "Production": "LaserCutting",
                "Materiaal": "S235",
                "Aantal": 1,
                "Oppervlakte": 0,
                "Gewicht": 0,
            }
        ]
    )

    count, _ = copy_per_production_and_orders(
        str(src),
        str(dest),
        bom_df,
        [".pdf"],
        _make_db(),
        {"LaserCutting": "ACME"},
        {},
        {},
        remember_defaults=False,
        en1090_overrides={"LaserCutting": True},
    )

    assert count >= 1
    production_dir = dest / "LaserCutting"
    excel_files = sorted(production_dir.glob("*.xlsx"))
    assert excel_files, "Bestelbon Excel niet gevonden"

    workbook = load_workbook(excel_files[0])
    sheet = workbook.active
    note_found = any(
        EN1090_NOTE_TEXT in str(cell)
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell
    )
    assert note_found, "EN 1090 vermelding niet gevonden in Excel"
