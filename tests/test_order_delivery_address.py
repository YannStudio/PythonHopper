import pandas as pd
from openpyxl import load_workbook
import pytest

from models import Supplier, Client
from suppliers_db import SuppliersDB
from orders import copy_per_production_and_orders


def test_delivery_address_used_in_order(tmp_path, monkeypatch):
    """The selected delivery address should appear in the order document."""
    # operate within temporary directory to avoid side effects
    monkeypatch.chdir(tmp_path)

    # supplier database with one supplier
    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])

    # create source and destination directories
    src = tmp_path / "src"
    dst = tmp_path / "dst"
    src.mkdir()
    dst.mkdir()

    # dummy source file
    (src / "PN1.pdf").write_text("dummy")

    # BOM with a single production entry
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1},
    ])

    supplier_map = {"Laser": "ACME"}
    delivery_map = {"Laser": "Custom Street 5"}

    client = Client.from_any({"name": "Client", "address": "Base Addr"})

    cnt, chosen = copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        supplier_map,
        delivery_map,
        False,
        client=client,
    )

    assert cnt == 1

    # verify that the generated Excel order contains invoice and delivery address
    prod_dir = dst / "Laser"
    excel_files = list(prod_dir.glob("Bestelbon_Laser_*.xlsx"))
    assert excel_files, "Order Excel file not created"

    wb = load_workbook(excel_files[0])
    ws = wb.active
    # row 2 should contain the invoice address, row 3 the chosen delivery address
    assert ws.cell(row=2, column=2).value == "Base Addr"
    assert ws.cell(row=3, column=2).value == "Custom Street 5"


def test_pdf_delivery_address_in_right_column(tmp_path, monkeypatch):
    reportlab = pytest.importorskip("reportlab")
    from PyPDF2 import PdfReader

    monkeypatch.chdir(tmp_path)

    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])

    src = tmp_path / "src"
    dst = tmp_path / "dst"
    src.mkdir()
    dst.mkdir()

    (src / "PN1.pdf").write_text("dummy")

    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1},
    ])

    supplier_map = {"Laser": "ACME"}
    delivery_map = {"Laser": "Custom Street 5"}

    client = Client.from_any({"name": "Client", "address": "Base Addr"})

    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        supplier_map,
        delivery_map,
        False,
        client=client,
    )

    prod_dir = dst / "Laser"
    pdf_files = list(prod_dir.glob("Bestelbon_Laser_*.pdf"))
    assert pdf_files, "Order PDF file not created"

    reader = PdfReader(str(pdf_files[0]))
    page = reader.pages[0]
    positions = {}

    def visitor(text, cm, tm, *_):
        text = text.strip()
        if not text:
            return
        x = cm[0] * tm[4] + cm[2] * tm[5] + cm[4]
        positions.setdefault(text, []).append(x)

    page.extract_text(visitor_text=visitor)

    inv_x = positions["Base Addr"][0]
    del_x = positions["Custom Street 5"][0]
    assert del_x > inv_x

