import datetime

import pandas as pd
import pytest
from openpyxl import load_workbook
from PyPDF2 import PdfReader

from models import Supplier
from orders import (
    _normalize_finish_folder,
    copy_per_production_and_orders,
    make_finish_selection_key,
    make_production_selection_key,
)
from suppliers_db import SuppliersDB


def _sheet_contains(ws, expected: str) -> bool:
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value == expected:
                return True
    return False


def _pdf_text(path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def test_grouped_production_orders_create_single_document_in_master_folder(tmp_path):
    pytest.importorskip("reportlab")
    pytest.importorskip("openpyxl")

    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("pdf", encoding="utf-8")
    (src / "PN2.pdf").write_text("pdf", encoding="utf-8")

    bom_df = pd.DataFrame(
        [
            {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1},
            {"PartNumber": "PN2", "Description": "", "Production": "Plasma", "Aantal": 1},
        ]
    )

    today = datetime.date.today().strftime("%Y-%m-%d")
    master_key = make_production_selection_key("Laser")
    follower_key = make_production_selection_key("Plasma")

    cnt, _chosen = copy_per_production_and_orders(
        str(src),
        str(dest),
        bom_df,
        [".pdf"],
        db,
        {"Laser": "ACME", "Plasma": "ACME"},
        {"Laser": "Bestelbon", "Plasma": "Bestelbon"},
        {"Laser": "BB-200", "Plasma": "BB-201"},
        False,
        client=None,
        delivery_map={},
        document_group_map={follower_key: master_key},
        export_bom=False,
    )

    assert cnt == 2

    master_dir = dest / "Laser"
    follower_dir = dest / "Plasma"
    excel_path = master_dir / f"Bestelbon_BB-200_Laser_Groep_{today}.xlsx"
    pdf_path = master_dir / f"Bestelbon_BB-200_Laser_Groep_{today}.pdf"

    assert excel_path.exists()
    assert pdf_path.exists()
    assert not (follower_dir / f"Bestelbon_BB-201_Plasma_{today}.xlsx").exists()
    assert not (follower_dir / f"Bestelbon_BB-201_Plasma_{today}.pdf").exists()

    ws = load_workbook(excel_path).active
    assert _sheet_contains(ws, "Gecombineerde bon voor")
    assert _sheet_contains(ws, "Laser, Plasma")
    assert _sheet_contains(ws, "Productie: Laser")
    assert _sheet_contains(ws, "Productie: Plasma")

    text = _pdf_text(pdf_path)
    assert "Gecombineerde bon voor: Laser, Plasma" in text
    assert "Productie: Laser" in text
    assert "Productie: Plasma" in text
    assert "Nummer: BB-200" in text


def test_grouped_finish_orders_create_single_document_in_master_folder(tmp_path):
    pytest.importorskip("reportlab")
    pytest.importorskip("openpyxl")

    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("pdf", encoding="utf-8")
    (src / "PN2.pdf").write_text("pdf", encoding="utf-8")

    bom_df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Description": "",
                "Production": "Laser",
                "Finish": "Poedercoating",
                "RAL color": "RAL 9005",
                "Aantal": 1,
            },
            {
                "PartNumber": "PN2",
                "Description": "",
                "Production": "Laser",
                "Finish": "Nattenlak",
                "RAL color": "RAL 9010",
                "Aantal": 1,
            },
        ]
    )

    finish_key_a = (
        "Finish-"
        + _normalize_finish_folder("Poedercoating")
        + "-"
        + _normalize_finish_folder("RAL 9005")
    )
    finish_key_b = (
        "Finish-"
        + _normalize_finish_folder("Nattenlak")
        + "-"
        + _normalize_finish_folder("RAL 9010")
    )
    master_key = make_finish_selection_key(finish_key_a)
    follower_key = make_finish_selection_key(finish_key_b)
    today = datetime.date.today().strftime("%Y-%m-%d")

    copy_per_production_and_orders(
        str(src),
        str(dest),
        bom_df,
        [".pdf"],
        db,
        {"Laser": "ACME"},
        {},
        {},
        False,
        client=None,
        delivery_map={},
        copy_finish_exports=True,
        finish_override_map={
            finish_key_a: "ACME",
            finish_key_b: "ACME",
        },
        finish_doc_num_map={
            finish_key_a: "BB-300",
            finish_key_b: "BB-301",
        },
        document_group_map={follower_key: master_key},
        export_bom=False,
    )

    master_dir = dest / finish_key_a
    follower_dir = dest / finish_key_b
    excel_matches = list(master_dir.glob(f"Bestelbon_BB-300_*_Groep_{today}.xlsx"))
    pdf_matches = list(master_dir.glob(f"Bestelbon_BB-300_*_Groep_{today}.pdf"))
    assert len(excel_matches) == 1
    assert len(pdf_matches) == 1
    assert not list(follower_dir.glob(f"Bestelbon_BB-301_*_{today}.xlsx"))
    assert not list(follower_dir.glob(f"Bestelbon_BB-301_*_{today}.pdf"))

    excel_path = excel_matches[0]
    pdf_path = pdf_matches[0]

    ws = load_workbook(excel_path).active
    assert _sheet_contains(ws, "Gecombineerde bon voor")
    assert _sheet_contains(ws, "Poedercoating – RAL 9005, Nattenlak – RAL 9010")
    assert _sheet_contains(ws, "Afwerking: Poedercoating – RAL 9005")
    assert _sheet_contains(ws, "Afwerking: Nattenlak – RAL 9010")

    text = _pdf_text(pdf_path)
    assert "Gecombineerde bon voor: Poedercoating – RAL 9005" in text
    assert "Nattenlak – RAL 9010" in text
    assert "Afwerking: Poedercoating – RAL 9005" in text
    assert "Afwerking: Nattenlak – RAL 9010" in text
    assert "Nummer: BB-300" in text


def test_grouped_finish_orders_report_master_folder_path(tmp_path):
    pytest.importorskip("reportlab")
    pytest.importorskip("openpyxl")

    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("pdf", encoding="utf-8")
    (src / "PN2.pdf").write_text("pdf", encoding="utf-8")

    bom_df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Description": "",
                "Production": "Laser",
                "Finish": "Poedercoating",
                "RAL color": "RAL 9005",
                "Aantal": 1,
            },
            {
                "PartNumber": "PN2",
                "Description": "",
                "Production": "Laser",
                "Finish": "Nattenlak",
                "RAL color": "RAL 9010",
                "Aantal": 1,
            },
        ]
    )

    finish_key_a = (
        "Finish-"
        + _normalize_finish_folder("Poedercoating")
        + "-"
        + _normalize_finish_folder("RAL 9005")
    )
    finish_key_b = (
        "Finish-"
        + _normalize_finish_folder("Nattenlak")
        + "-"
        + _normalize_finish_folder("RAL 9010")
    )
    master_key = make_finish_selection_key(finish_key_a)
    follower_key = make_finish_selection_key(finish_key_b)
    status_lines = []

    copy_per_production_and_orders(
        str(src),
        str(dest),
        bom_df,
        [".pdf"],
        db,
        {"Laser": "ACME"},
        {},
        {},
        False,
        client=None,
        delivery_map={},
        copy_finish_exports=True,
        finish_override_map={
            finish_key_a: "ACME",
            finish_key_b: "ACME",
        },
        finish_doc_num_map={
            finish_key_a: "BB-300",
            finish_key_b: "BB-301",
        },
        document_group_map={follower_key: master_key},
        export_bom=False,
        document_status_messages=status_lines,
    )

    finish_reports = [
        line for line in status_lines if "Samengestelde afwerkingsbon" in line
    ]
    assert len(finish_reports) == 1
    assert "Poedercoating" in finish_reports[0]
    assert finish_key_a in finish_reports[0]
    assert "Bestelbon_BB-300" in finish_reports[0]


def test_finish_export_filter_reports_skipped_finish(tmp_path):
    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("pdf", encoding="utf-8")

    bom_df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Description": "",
                "Production": "Laser",
                "Finish": "Poedercoating",
                "RAL color": "RAL 9005",
                "Aantal": 1,
            }
        ]
    )

    finish_key = (
        "Finish-"
        + _normalize_finish_folder("Poedercoating")
        + "-"
        + _normalize_finish_folder("RAL 9005")
    )
    status_lines = []

    copy_per_production_and_orders(
        str(src),
        str(dest),
        bom_df,
        [".pdf"],
        db,
        {"Laser": "ACME"},
        {},
        {},
        False,
        client=None,
        delivery_map={},
        copy_finish_exports=True,
        finish_override_map={finish_key: "ACME"},
        finish_export_filter={finish_key: False},
        export_bom=False,
        document_status_messages=status_lines,
    )

    assert len(status_lines) == 1
    assert "Afwerking 'Poedercoating" in status_lines[0]
    assert "RAL 9005" in status_lines[0]
    assert "overgeslagen: export uitgeschakeld." in status_lines[0]


def test_grouped_production_orders_reject_mixed_en1090(tmp_path):
    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("pdf", encoding="utf-8")
    (src / "PN2.pdf").write_text("pdf", encoding="utf-8")

    bom_df = pd.DataFrame(
        [
            {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1},
            {"PartNumber": "PN2", "Description": "", "Production": "Plasma", "Aantal": 1},
        ]
    )

    with pytest.raises(ValueError, match="EN 1090"):
        copy_per_production_and_orders(
            str(src),
            str(dest),
            bom_df,
            [".pdf"],
            db,
            {"Laser": "ACME", "Plasma": "ACME"},
            {},
            {"Laser": "BB-400", "Plasma": "BB-401"},
            False,
            client=None,
            delivery_map={},
            document_group_map={
                make_production_selection_key("Plasma"): make_production_selection_key(
                    "Laser"
                )
            },
            export_bom=False,
            en1090_overrides={"Laser": True, "Plasma": False},
        )
