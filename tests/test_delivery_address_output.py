import os
import pandas as pd
import openpyxl
import pytest
from PyPDF2 import PdfReader

from models import Supplier, DeliveryAddress
from suppliers_db import SuppliersDB
from orders import copy_per_production_and_orders


def _setup_basic(tmp_path):
    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}
    ])
    return db, src, bom_df


def test_delivery_address_present_absent(tmp_path):
    reportlab = pytest.importorskip("reportlab")
    db, src, bom_df = _setup_basic(tmp_path)

    delivery = DeliveryAddress(name="Magazijn", address="Straat 1", remarks="achterdeur")

    # With delivery address
    dst1 = tmp_path / "dst1"
    dst1.mkdir()
    copy_per_production_and_orders(
        str(src),
        str(dst1),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        False,
        client=None,
        delivery_map={"Laser": delivery},
    )
    prod_folder = dst1 / "Laser"
    xlsx = next(f for f in os.listdir(prod_folder) if f.endswith(".xlsx"))
    wb = openpyxl.load_workbook(prod_folder / xlsx)
    ws = wb.active
    col_a = [ws[f"A{i}"].value for i in range(1, 20)]
    assert "Leveradres" in col_a
    row = col_a.index("Leveradres") + 1
    # Name is on the next line with an empty label
    assert ws[f"B{row}"].value in (None, "")
    assert ws[f"B{row+1}"].value == "Magazijn"
    assert ws[f"B{row+2}"].value == "Straat 1"
    assert ws[f"B{row+3}"].value == "achterdeur"
    pdf = next(f for f in os.listdir(prod_folder) if f.endswith(".pdf"))
    reader = PdfReader(prod_folder / pdf)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Leveradres:\nMagazijn" in text

    # Without delivery address
    dst2 = tmp_path / "dst2"
    dst2.mkdir()
    copy_per_production_and_orders(
        str(src),
        str(dst2),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        False,
        client=None,
        delivery_map={},
    )
    prod_folder2 = dst2 / "Laser"
    xlsx2 = next(f for f in os.listdir(prod_folder2) if f.endswith(".xlsx"))
    wb2 = openpyxl.load_workbook(prod_folder2 / xlsx2)
    ws2 = wb2.active
    col_a2 = [ws2[f"A{i}"].value for i in range(1, 20)]
    assert "Leveradres" not in col_a2
    pdf2 = next(f for f in os.listdir(prod_folder2) if f.endswith(".pdf"))
    reader2 = PdfReader(prod_folder2 / pdf2)
    text2 = "\n".join(page.extract_text() or "" for page in reader2.pages)
    assert "Leveradres" not in text2


def test_delivery_address_per_production(tmp_path):
    reportlab = pytest.importorskip("reportlab")
    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    src.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    (src / "PN2.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1},
        {"PartNumber": "PN2", "Description": "", "Production": "Plasma", "Aantal": 1},
    ])

    d1 = DeliveryAddress(name="Magazijn", address="Straat 1")
    d2 = DeliveryAddress(name="Depot", address="Weg 2")
    dst = tmp_path / "dst"
    dst.mkdir()
    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        False,
        client=None,
        delivery_map={"Laser": d1, "Plasma": d2},
    )

    # Laser folder checks
    laser = dst / "Laser"
    xlsx1 = next(f for f in os.listdir(laser) if f.endswith(".xlsx"))
    wb1 = openpyxl.load_workbook(laser / xlsx1)
    ws1 = wb1.active
    col_a1 = [ws1[f"A{i}"].value for i in range(1, 20)]
    row1 = col_a1.index("Leveradres") + 1
    assert ws1[f"B{row1}"].value in (None, "")
    assert ws1[f"B{row1+1}"].value == "Magazijn"
    pdf1 = next(f for f in os.listdir(laser) if f.endswith(".pdf"))
    reader1 = PdfReader(laser / pdf1)
    text1 = "\n".join(page.extract_text() or "" for page in reader1.pages)
    assert "Leveradres:\nMagazijn" in text1

    # Plasma folder checks
    plasma = dst / "Plasma"
    xlsx2 = next(f for f in os.listdir(plasma) if f.endswith(".xlsx"))
    wb2 = openpyxl.load_workbook(plasma / xlsx2)
    ws2 = wb2.active
    col_a2 = [ws2[f"A{i}"].value for i in range(1, 20)]
    row2 = col_a2.index("Leveradres") + 1
    assert ws2[f"B{row2}"].value in (None, "")
    assert ws2[f"B{row2+1}"].value == "Depot"
    pdf2 = next(f for f in os.listdir(plasma) if f.endswith(".pdf"))
    reader2 = PdfReader(plasma / pdf2)
    text2 = "\n".join(page.extract_text() or "" for page in reader2.pages)
    assert "Leveradres:\nDepot" in text2


def test_delivery_address_placeholder_prints(tmp_path):
    reportlab = pytest.importorskip("reportlab")
    db, src, bom_df = _setup_basic(tmp_path)

    # Placeholder address should still render in the delivery block
    delivery = DeliveryAddress(name="Magazijn", address="Bestelling wordt opgehaald")

    dst = tmp_path / "dst_placeholder"
    dst.mkdir()
    copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        {},
        {},
        False,
        client=None,
        delivery_map={"Laser": delivery},
    )

    prod_folder = dst / "Laser"
    pdf = next(f for f in os.listdir(prod_folder) if f.endswith(".pdf"))
    reader = PdfReader(prod_folder / pdf)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Bestelling wordt opgehaald" in text
