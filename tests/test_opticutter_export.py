import datetime
import os
import openpyxl
import pandas as pd
import pytest

from opticutter import analyse_profiles
from orders import copy_per_production_and_orders
from suppliers_db import SuppliersDB


def test_opticutter_files_written(tmp_path):
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()

    df = pd.DataFrame(
        [
            {
                "PartNumber": "P-001",
                "Description": "Profiel A",
                "Production": "Prod1",
                "Profile": "U-80",
                "Length profile": 2500,
                "Materiaal": "Staal",
                "Aantal": 2,
                "Gewicht": 5.0,
            },
            {
                "PartNumber": "P-002",
                "Description": "Profiel B",
                "Production": "Prod1",
                "Profile": "U-80",
                "Length profile": 1500,
                "Materiaal": "Staal",
                "Aantal": 1,
                "Gewicht": 3.0,
            },
            {
                "PartNumber": "P-003",
                "Description": "Profiel C",
                "Production": "Prod2",
                "Profile": "L-50",
                "Length profile": 3200,
                "Materiaal": "Alu",
                "Aantal": 3,
                "Gewicht": 2.5,
            },
        ]
    )

    analysis = analyse_profiles(df)
    choices = {profile.key: profile.best_choice for profile in analysis.profiles}

    copy_per_production_and_orders(
        str(src),
        str(dest),
        df,
        [".pdf"],
        SuppliersDB([]),
        {},
        {},
        {},
        False,
        opticutter_analysis=analysis,
        opticutter_choices=choices,
        export_bom=False,
    )

    today = datetime.date.today().strftime("%Y-%m-%d")
    prod1_dir = dest / "Prod1"
    scenario_path = prod1_dir / f"Opticutter_Prod1_{today}.xlsx"
    order_path = prod1_dir / f"Bestelbon_brutemateriaal_Prod1_{today}.xlsx"

    assert scenario_path.exists(), "Opticutter scenario workbook ontbreekt"
    assert order_path.exists(), "Bestelbon brutemateriaal workbook ontbreekt"

    scenario_df = pd.read_excel(scenario_path, sheet_name="Scenario")
    assert "Keuze" in scenario_df.columns
    assert scenario_df["Keuze"].astype(str).str.len().max() > 0

    order_df = pd.read_excel(order_path)
    assert "Aantal staven" in order_df.columns
    assert order_df["Aantal staven"].notna().any()
    assert "Totaal gewicht (kg)" in order_df.columns
    assert order_df["Totaal gewicht (kg)"].fillna(0).sum() > 0

    raw_docs = [
        f
        for f in os.listdir(prod1_dir)
        if f.startswith("Bestelbon_")
        and "Brutemateriaal" in f
        and f.lower().endswith(".xlsx")
    ]
    assert raw_docs, "Bestelbon brutemateriaal XLS ontbreekt"

    raw_order_path = prod1_dir / sorted(raw_docs)[0]
    wb = openpyxl.load_workbook(raw_order_path)
    ws = wb.active
    header_row = None
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Profiel":
            header_row = row_idx
            break
    assert header_row is not None, "Profiel header niet gevonden"
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, 6)]
    assert headers == ["Profiel", "Materiaal", "Lengte", "St.", "kg"]

    total_row_idx = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Totaal":
            total_row_idx = row_idx
            break
    assert total_row_idx is not None, "Totaal gewicht rij ontbreekt"
    total_weight_cell = ws.cell(row=total_row_idx, column=5).value
    total_weight_value = str(total_weight_cell or "").replace(",", ".")
    assert total_weight_value, "Totaal gewicht waarde ontbreekt"
    expected_total = round(order_df["Totaal gewicht (kg)"].fillna(0).sum(), 2)
    assert pytest.approx(float(total_weight_value), abs=0.01) == expected_total
