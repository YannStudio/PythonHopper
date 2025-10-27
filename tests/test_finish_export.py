import datetime
import zipfile

import pandas as pd
import pytest
import zipfile

import cli
import orders
from app_settings import AppSettings
from bom import load_bom
from cli import build_parser, cli_copy_per_prod
from models import Supplier
from orders import (
    copy_per_production_and_orders,
    _normalize_finish_folder,
    make_finish_selection_key,
)
from suppliers_db import SuppliersDB
from clients_db import ClientsDB
from delivery_addresses_db import DeliveryAddressesDB


def _make_db() -> SuppliersDB:
    return SuppliersDB([Supplier.from_any({"supplier": "ACME"})])


def test_load_bom_finish_columns_and_copy(tmp_path):
    bom_path = tmp_path / "bom.csv"
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()

    pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Production": "Laser",
                "finish": "Poedercoating",
                "RAL COLOR": "RAL 9005",
                "Aantal": 2,
            },
            {
                "PartNumber": "PN2",
                "Production": "Laser",
                "finish": "Geanodiseerd",
                "RAL COLOR": None,
                "Aantal": 1,
            },
            {
                "PartNumber": "PN3",
                "Production": "Laser",
                "finish": "Gelakt",
                "RAL COLOR": 7016,
                "Aantal": 1,
            },
        ]
    ).to_csv(bom_path, index=False)

    for pn, content in {"PN1": "one", "PN2": "two", "PN3": "three"}.items():
        (src / f"{pn}.pdf").write_text(content, encoding="utf-8")

    loaded = load_bom(str(bom_path))

    assert list(loaded["Finish"]) == ["Poedercoating", "Geanodiseerd", "Gelakt"]
    assert list(loaded["RAL color"]) == ["RAL 9005", "", "7016"]

    cnt, _ = copy_per_production_and_orders(
        str(src),
        str(dest),
        loaded,
        [".pdf"],
        _make_db(),
        {"Laser": "ACME"},
        {},
        {},
        False,
        copy_finish_exports=True,
        zip_finish_exports=False,
    )

    assert cnt == 3

    finish_dir_1 = dest / (
        "Finish-"
        + _normalize_finish_folder("Poedercoating")
        + "-"
        + _normalize_finish_folder("RAL 9005")
    )
    finish_dir_2 = dest / ("Finish-" + _normalize_finish_folder("Geanodiseerd"))
    finish_dir_3 = dest / (
        "Finish-"
        + _normalize_finish_folder("Gelakt")
        + "-"
        + _normalize_finish_folder("7016")
    )

    assert (finish_dir_1 / "PN1.pdf").is_file()
    assert (finish_dir_2 / "PN2.pdf").is_file()
    assert (finish_dir_3 / "PN3.pdf").is_file()
    assert finish_dir_3.name.endswith("-7016")


@pytest.mark.parametrize("zip_parts", [False, True])
@pytest.mark.parametrize("zip_finish", [False, True])
def test_finish_exports_written(tmp_path, zip_parts, zip_finish):
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()

    (src / "PN1.pdf").write_text("one", encoding="utf-8")
    (src / "PN2.pdf").write_text("two", encoding="utf-8")
    (src / "PN3.pdf").write_text("three", encoding="utf-8")
    (src / "PN4.pdf").write_text("four", encoding="utf-8")

    df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Production": "Laser",
                "Finish": "Poedercoating",
                "RAL color": "RAL 9005",
                "Aantal": 1,
            },
            {
                "PartNumber": "PN2",
                "Production": "Laser",
                "Finish": "Anodized/Marine",
                "RAL color": "RAL-9010",
                "Aantal": 1,
            },
            {
                "PartNumber": "PN1",
                "Production": "Laser",
                "Finish": "Poedercoating",
                "RAL color": "RAL 9005",
                "Aantal": 3,
            },
            {
                "PartNumber": "PN3",
                "Production": "Laser",
                "Finish": " ",
                "RAL color": "RAL 9016",
                "Aantal": 2,
            },
            {
                "PartNumber": "PN4",
                "Production": "Laser",
                "Finish": "Geanodiseerd",
                "RAL color": "",
                "Aantal": 1,
            },
        ]
    )

    cnt, _ = copy_per_production_and_orders(
        str(src),
        str(dest),
        df,
        [".pdf"],
        _make_db(),
        {"Laser": "ACME"},
        {},
        {},
        False,
        zip_parts=zip_parts,
        copy_finish_exports=True,
        zip_finish_exports=zip_finish,
    )

    assert cnt == 4

    finish_dir_1_name = (
        "Finish-"
        + _normalize_finish_folder("Poedercoating")
        + "-"
        + _normalize_finish_folder("RAL 9005")
    )
    finish_dir_2_name = (
        "Finish-"
        + _normalize_finish_folder("Anodized/Marine")
        + "-"
        + _normalize_finish_folder("RAL-9010")
    )
    finish_dir_1 = dest / finish_dir_1_name
    finish_dir_2 = dest / finish_dir_2_name
    finish_dir_3_name = "Finish-" + _normalize_finish_folder("Geanodiseerd")
    finish_dir_3 = dest / finish_dir_3_name
    missing_finish_dir_name = (
        "Finish-"
        + _normalize_finish_folder(" ")
        + "-"
        + _normalize_finish_folder("RAL 9016")
    )

    entries_finish_1 = sorted(p.name for p in finish_dir_1.iterdir())
    if zip_finish:
        zip_path_1 = finish_dir_1 / f"{finish_dir_1_name}.zip"
        assert zip_path_1.is_file()
        with zipfile.ZipFile(zip_path_1) as zf:
            assert "PN1.pdf" in zf.namelist()
    else:
        assert (finish_dir_1 / "PN1.pdf").is_file()
        assert "PN1.pdf" in entries_finish_1
    assert any(name.startswith("Bestelbon_") for name in entries_finish_1)
    if zip_finish:
        zip_path_2 = finish_dir_2 / f"{finish_dir_2_name}.zip"
        assert zip_path_2.is_file()
        with zipfile.ZipFile(zip_path_2) as zf:
            assert "PN2.pdf" in zf.namelist()
        zip_path_3 = finish_dir_3 / f"{finish_dir_3_name}.zip"
        assert zip_path_3.is_file()
        with zipfile.ZipFile(zip_path_3) as zf:
            assert "PN4.pdf" in zf.namelist()
    else:
        assert (finish_dir_2 / "PN2.pdf").is_file()
        assert (finish_dir_3 / "PN4.pdf").is_file()
    finish_dirs = sorted(
        p.name for p in dest.iterdir() if p.is_dir() and p.name.startswith("Finish-")
    )
    assert finish_dirs == sorted(
        [finish_dir_1_name, finish_dir_2_name, finish_dir_3_name]
    )
    assert not (dest / missing_finish_dir_name).exists()

    prod_dir = dest / "Laser"
    if zip_parts:
        archive = prod_dir / "Laser.zip"
        assert archive.is_file()
        with zipfile.ZipFile(archive) as zf:
            assert any(name.endswith("PN3.pdf") for name in zf.namelist())
    else:
        assert (prod_dir / "PN3.pdf").is_file()


@pytest.mark.parametrize(
    "flag, settings_value, expected",
    [
        (None, True, True),
        ("--finish-folders", False, True),
        ("--no-finish-folders", True, False),
    ],
)
def test_cli_finish_flag(monkeypatch, tmp_path, flag, settings_value, expected):
    parser = build_parser()
    args_list = [
        "copy-per-prod",
        "--source",
        str(tmp_path / "src"),
        "--dest",
        str(tmp_path / "dst"),
        "--bom",
        str(tmp_path / "bom.xlsx"),
        "--exts",
        "pdf",
    ]
    if flag:
        args_list.append(flag)
    args = parser.parse_args(args_list)

    (tmp_path / "src").mkdir()
    (tmp_path / "dst").mkdir()

    df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Description": "",
                "Production": "Laser",
                "Aantal": 1,
            }
        ]
    )
    monkeypatch.setattr(cli, "load_bom", lambda path: df)

    sdb = _make_db()
    monkeypatch.setattr(SuppliersDB, "load", classmethod(lambda cls, path: sdb))
    cdb = ClientsDB([])
    monkeypatch.setattr(ClientsDB, "load", classmethod(lambda cls, path: cdb))
    ddb = DeliveryAddressesDB([])
    monkeypatch.setattr(DeliveryAddressesDB, "load", classmethod(lambda cls, path: ddb))

    monkeypatch.setattr(
        AppSettings,
        "load",
        classmethod(
            lambda cls, path=...: AppSettings(
                copy_finish_exports=settings_value,
                zip_finish_exports=True,
            )
        ),
    )

    captured = {}

    def fake_copy(*_args, **kwargs):
        captured.update(kwargs)
        return 0, {}

    monkeypatch.setattr(cli, "copy_per_production_and_orders", fake_copy)

    cli_copy_per_prod(args)

    assert captured["copy_finish_exports"] is expected


@pytest.mark.parametrize(
    "flag, settings_value, expected",
    [
        (None, True, True),
        ("--zip-finish-folders", False, True),
        ("--no-zip-finish-folders", True, False),
    ],
)
def test_cli_zip_finish_flag(monkeypatch, tmp_path, flag, settings_value, expected):
    parser = build_parser()
    args_list = [
        "copy-per-prod",
        "--source",
        str(tmp_path / "src"),
        "--dest",
        str(tmp_path / "dst"),
        "--bom",
        str(tmp_path / "bom.xlsx"),
        "--exts",
        "pdf",
    ]
    if flag:
        args_list.append(flag)
    args = parser.parse_args(args_list)

    (tmp_path / "src").mkdir()
    (tmp_path / "dst").mkdir()

    df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Description": "",
                "Production": "Laser",
                "Aantal": 1,
            }
        ]
    )
    monkeypatch.setattr(cli, "load_bom", lambda path: df)

    sdb = _make_db()
    monkeypatch.setattr(SuppliersDB, "load", classmethod(lambda cls, path: sdb))
    cdb = ClientsDB([])
    monkeypatch.setattr(ClientsDB, "load", classmethod(lambda cls, path: cdb))
    ddb = DeliveryAddressesDB([])
    monkeypatch.setattr(DeliveryAddressesDB, "load", classmethod(lambda cls, path: ddb))

    monkeypatch.setattr(
        AppSettings,
        "load",
        classmethod(
            lambda cls, path=...: AppSettings(
                copy_finish_exports=True,
                zip_finish_exports=settings_value,
            )
        ),
    )

    captured = {}

    def fake_copy(*_args, **kwargs):
        captured.update(kwargs)
        return 0, {}

    monkeypatch.setattr(cli, "copy_per_production_and_orders", fake_copy)

    cli_copy_per_prod(args)

    assert captured["zip_finish_exports"] is expected


def test_finish_documents_and_defaults(tmp_path, monkeypatch):
    pytest.importorskip("reportlab")
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from PyPDF2 import PdfReader

    monkeypatch.setattr(orders, "SUPPLIERS_DB_FILE", str(tmp_path / "suppliers.json"))

    db = SuppliersDB([Supplier.from_any({"supplier": "ACME"})])
    src = tmp_path / "src"
    dest = tmp_path / "dest"
    src.mkdir()
    dest.mkdir()
    (src / "PN1.pdf").write_text("pdf", encoding="utf-8")

    df = pd.DataFrame(
        [
            {
                "PartNumber": "PN1",
                "Production": "Laser",
                "Finish": "Poedercoating",
                "RAL color": "RAL 9005",
                "Aantal": 1,
            }
        ]
    )

    finish_key = "Finish-" + _normalize_finish_folder("Poedercoating") + "-" + _normalize_finish_folder("RAL 9005")
    today = datetime.date.today().strftime("%Y-%m-%d")
    label = "Poedercoating – RAL 9005"
    filename_component = _normalize_finish_folder(label)

    doc_number_raw = "12/34"

    cnt, chosen = copy_per_production_and_orders(
        str(src),
        str(dest),
        df,
        [".pdf"],
        db,
        {"Laser": "ACME"},
        {},
        {},
        True,
        copy_finish_exports=True,
        finish_override_map={finish_key: "ACME"},
        finish_doc_num_map={finish_key: doc_number_raw},
    )

    assert cnt == 1
    assert chosen[make_finish_selection_key(finish_key)] == "ACME"
    assert db.get_default_finish(finish_key) == "ACME"

    finish_dir = dest / finish_key
    assert finish_dir.is_dir()
    excel_path = finish_dir / f"Bestelbon_BB-12_34_{filename_component}_{today}.xlsx"
    pdf_path = finish_dir / f"Bestelbon_BB-12_34_{filename_component}_{today}.pdf"
    assert excel_path.exists()
    assert pdf_path.exists()

    wb = load_workbook(excel_path)
    ws = wb.active
    values = {ws[f"A{i}"].value: ws[f"B{i}"].value for i in range(1, 10)}
    assert values.get("Nummer") == "BB-12/34"
    assert values.get("Afwerking") == label

    reader = PdfReader(str(pdf_path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Afwerking: Poedercoating – RAL 9005" in text
    assert "Nummer: BB-12/34" in text
