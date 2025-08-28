import pandas as pd
import openpyxl
import pytest

from models import Supplier, DeliveryAddress
from suppliers_db import SuppliersDB
from orders import copy_per_production_and_orders, REPORTLAB_OK


def test_delivery_address_in_documents(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    db = SuppliersDB()
    db.upsert(Supplier.from_any({"supplier": "ACME"}))
    src = tmp_path / "src"
    dst = tmp_path / "dst"
    src.mkdir(); dst.mkdir()
    (src / "PN1.pdf").write_text("dummy")
    bom_df = pd.DataFrame([
        {"PartNumber": "PN1", "Description": "", "Production": "Laser", "Aantal": 1}
    ])
    addr = DeliveryAddress.from_any({
        "name": "Magazijn",
        "address": "Straat 5, 1000 Brussel",
        "contact": "Jan",
        "phone": "012345",
        "email": "jan@example.com",
    })
    overrides = {"Laser": "ACME"}
    addr_map = {"Laser": addr}
    cnt, chosen = copy_per_production_and_orders(
        str(src),
        str(dst),
        bom_df,
        [".pdf"],
        db,
        overrides,
        False,
        addr_map,
    )
    assert cnt == 1
    prod_dir = dst / "Laser"
    xlsx = next(prod_dir.glob("*.xlsx"))
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active
    found = False
    for i in range(1, 20):
        if ws[f"A{i}"].value == "Leveringsadres":
            assert ws[f"B{i}"].value == addr.name
            assert ws[f"B{i+1}"].value == addr.address
            found = True
            break
    assert found, "Leveringsadres niet gevonden in Excel"
    if REPORTLAB_OK:
        pdf = next(prod_dir.glob("*.pdf"))
        content = pdf.read_bytes().decode("latin1")
        assert addr.address in content
        assert addr.name in content
