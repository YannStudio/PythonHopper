"""Order-related file operations and export orchestration.

This module contains the legacy order export implementation while the public
``orders`` package keeps backward-compatible imports.
"""

import os
import sys
import shutil
import datetime
import re
import unicodedata
import zipfile
import io
import tempfile
import hashlib
import math
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from html import escape
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from dataclasses import dataclass, field
from app_paths import resolve_runtime_path
try:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - optional dependency
    Alignment = None
    Font = None
    get_column_letter = None

try:
    from PyPDF2 import PdfMerger
except Exception:  # pragma: no cover - PyPDF2 might be absent
    PdfMerger = None

# ReportLab (PDF). Script works without it (PDF generation is skipped).
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        LongTable,
        Table,
        TableStyle,
        Spacer,
        Image as RLImage,
        KeepTogether,
    )
    from reportlab.pdfbase.pdfmetrics import stringWidth
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

    def stringWidth(text, font_name, font_size):
        """Approximate text width when ReportLab is unavailable."""

        return len(str(text)) * float(font_size) * 0.6

from en1090 import EN1090_NOTE_TEXT, should_require_en1090

from opticutter import (
    OpticutterAnalysis,
    OpticutterExportContext,
    OpticutterProductionExport,
    parse_length_to_mm,
    prepare_opticutter_export,
)
from spare_parts import (
    SPARE_PARTS_FULL_LIST_DOCUMENT_LABEL,
    SPARE_PARTS_FULL_LIST_KEY,
    is_spare_parts_production,
)

import step_previews
from helpers import (
    _to_str,
    _num_to_2dec,
    _material_nowrap,
    _build_file_index,
)
from models import Supplier, Client, DeliveryAddress, color_to_rgb, normalize_rgb_color
from pdf_workdossier_presets import PdfWorkDossierPreset, PdfWorkDossierSection
from progress import ProgressCallback, ProgressEvent
from .core import (
    DEFAULT_FOOTER_NOTE as _CORE_DEFAULT_FOOTER_NOTE,
    DEFAULT_ORDER_FOOTER_NOTE as _CORE_DEFAULT_ORDER_FOOTER_NOTE,
    DEFAULT_QUOTE_FOOTER_NOTE as _CORE_DEFAULT_QUOTE_FOOTER_NOTE,
    calculate_order_measure_totals,
    footer_note_for_doc_type,
    format_supplier_address,
    is_order_surface_column,
    is_order_weight_column,
)
from suppliers_db import SuppliersDB, SUPPLIERS_DB_FILE
from bom import load_bom  # noqa: F401 - imported for module dependency

MIAMI_PINK = "#FF77FF"
ORDER_RULE_COLOR = "#B9C1CA"
ORDER_TEXT_COLOR = "#1F2329"
ORDER_MUTED_TEXT_COLOR = "#5D6670"
ORDER_TABLE_OUTLINE_COLOR = "#B7BEC8"
ORDER_TABLE_GRID_COLOR = "#D5DAE1"
ORDER_TABLE_ALT_ROW_COLOR = "#FBFCFD"
ORDER_TOTAL_FILL_COLOR = "#FFF4FF"
ORDER_DELIVERY_FILL_COLOR = "#FFF8FD"
DEFAULT_ORDER_FOOTER_NOTE = _CORE_DEFAULT_ORDER_FOOTER_NOTE
DEFAULT_QUOTE_FOOTER_NOTE = _CORE_DEFAULT_QUOTE_FOOTER_NOTE
DEFAULT_FOOTER_NOTE = _CORE_DEFAULT_FOOTER_NOTE

STEP_EXTS = {".step", ".stp"}

NO_SUPPLIER_PLACEHOLDER = "(geen)"


def _mix_color_with_white(color: str, whiteness: float) -> str:
    rgb = color_to_rgb(color)
    if rgb is None:
        return color
    ratio = max(0.0, min(1.0, float(whiteness)))
    mixed = tuple(int(round(channel + (255 - channel) * ratio)) for channel in rgb)
    return "#{:02X}{:02X}{:02X}".format(*mixed)


def _accent_text_color(fill_color: str) -> str:
    rgb = color_to_rgb(fill_color)
    if rgb is None:
        return ORDER_TEXT_COLOR
    luminance = ((0.299 * rgb[0]) + (0.587 * rgb[1]) + (0.114 * rgb[2])) / 255.0
    return ORDER_TEXT_COLOR if luminance >= 0.68 else "#FFFFFF"


def _order_palette(company_info: Mapping[str, object] | None) -> Dict[str, str]:
    accent_color = normalize_rgb_color(
        company_info.get("accent_color") if company_info else None
    ) or MIAMI_PINK
    return {
        "accent": accent_color,
        "accent_text": _accent_text_color(accent_color),
        "total_fill": _mix_color_with_white(accent_color, 0.88),
    }


def _clean_order_cell_text(value: object) -> str:
    text = _to_str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return " ".join(text.split())


def _fit_text_to_width(
    text: str,
    width: float,
    font_name: str,
    font_size: float,
) -> Tuple[str, str]:
    if not text:
        return "", ""
    if stringWidth(text, font_name, font_size) <= width:
        return text, ""
    lo = 1
    hi = len(text)
    best = 1
    while lo <= hi:
        mid = (lo + hi) // 2
        probe = text[:mid]
        if stringWidth(probe, font_name, font_size) <= width:
            best = mid
            lo = mid + 1
        else:
            hi = mid - 1
    head = text[:best].rstrip()
    tail = text[best:].lstrip()
    if not head:
        head = text[:1]
        tail = text[1:].lstrip()
    return head, tail


def _truncate_text_to_width(
    text: str,
    width: float,
    font_name: str,
    font_size: float,
    suffix: str = "...",
    force_suffix: bool = False,
) -> str:
    text = text.rstrip()
    if not text:
        return ""
    if stringWidth(text, font_name, font_size) <= width and not force_suffix:
        return text
    while text and stringWidth(text + suffix, font_name, font_size) > width:
        text = text[:-1].rstrip()
    return (text + suffix) if text else suffix


def _wrap_words_to_lines(
    text: str,
    width: float,
    font_name: str,
    font_size: float,
    max_lines: int,
) -> List[str]:
    clean = _clean_order_cell_text(text)
    if not clean:
        return []
    pending = clean.split()
    lines: List[str] = []
    while pending and len(lines) < max_lines:
        current = pending.pop(0)
        if stringWidth(current, font_name, font_size) > width:
            current, remainder = _fit_text_to_width(
                current,
                width,
                font_name,
                font_size,
            )
            if remainder:
                pending.insert(0, remainder)
        while pending:
            probe = f"{current} {pending[0]}"
            if stringWidth(probe, font_name, font_size) > width:
                break
            current = probe
            pending.pop(0)
        lines.append(current)
    if pending and lines:
        lines[-1] = _truncate_text_to_width(
            lines[-1],
            width,
            font_name,
            font_size,
            force_suffix=True,
        )
    return lines


_BOM_STATUS_COLUMNS: Tuple[str, ...] = ("Bestanden gevonden", "Status", "Link")
_BOM_EXPORT_BASE_COLUMNS: Tuple[str, ...] = (
    "PartNumber",
    "Description",
    "QTY.",
    "Profile",
    "Length profile",
    "Production",
    "Materiaal",
    "Supplier",
    "Supplier code",
    "Manufacturer",
    "Manufacturer code",
    "Finish",
    "RAL color",
    "Oppervlakte",
    "Gewicht",
)


@dataclass(slots=True)
class CombinedPdfResult:
    """Metadata for combined PDF export operations."""

    count: int
    output_dir: str
    output_files: List[str] = field(default_factory=list)


@dataclass(slots=True)
class PdfWorkDossierPlanItem:
    """One PDF file in the final work dossier merge plan."""

    path: str
    section_name: str = ""
    production: str = ""
    role: str = "drawing"


PDF_WORKDOSSIER_SUPPLEMENTARY_SECTION = "Aanvullende bonnen zonder tekening"


@dataclass(slots=True)
class OpticutterProfileStats:
    """Aggregated length/weight data for a single Opticutter profile."""

    total_length_mm: float = 0.0
    total_weight_kg: float = 0.0

    @property
    def weight_per_mm(self) -> float | None:
        if self.total_length_mm <= 0 or self.total_weight_kg <= 0:
            return None
        return self.total_weight_kg / self.total_length_mm


@dataclass(slots=True)
class OpticutterOrderComputation:
    """Computed data for Opticutter raw material exports per production."""

    scenario_rows: List[Dict[str, object]]
    piece_rows: List[Dict[str, object]]
    order_rows: List[Dict[str, object]]
    raw_items: List[Dict[str, object]]
    has_valid_bars: bool
    total_bars: int
    total_weight_kg: float | None
    selection_count: int


@dataclass(slots=True)
class OrderDocumentSection:
    """A single logical section inside an order document."""

    context_label: str
    context_kind: str
    items: List[Dict[str, object]]
    selection_key: str = ""
    total_surface_m2: float | None = None
    total_weight_kg: float | None = None
    column_layout: Optional[List[Dict[str, object]]] = None


@dataclass(slots=True)
class OrderDocumentCandidate:
    """Collected data for one selectable order document before grouping."""

    selection_key: str
    context_label: str
    context_kind: str
    filename_context: str
    target_dir: str
    supplier: Supplier | None
    delivery: DeliveryAddress | None
    doc_type: str
    doc_num: str
    doc_num_display: str
    order_remark: str | None
    items: List[Dict[str, object]]
    total_surface_m2: float | None = None
    total_weight_kg: float | None = None
    column_layout: Optional[List[Dict[str, object]]] = None
    en1090_required: bool = False


@dataclass(slots=True)
class OrderDocumentJob:
    """A concrete PDF/XLSX document to render."""

    target_dir: str
    context_for_filename: str
    doc_type: str
    doc_num: str
    doc_num_display: str
    supplier: Supplier | None
    delivery: DeliveryAddress | None
    order_remark: str | None
    sections: List[OrderDocumentSection] = field(default_factory=list)
    en1090_required: bool = False


_PRICE_UNIT_KEY = "Eenheidsprijs"
_PRICE_TOTAL_KEY = "Totaalprijs"
_LINE_PRICE_KEY = "_line_price_key"


def build_order_pricing_item_key(
    item: Mapping[str, object],
    *,
    context_kind: str,
) -> str:
    """Return a stable key for storing/reloading line pricing."""

    context_kind_clean = (_to_str(context_kind) or "productie").strip().lower()
    is_raw = context_kind_clean.startswith("brutemateriaal")
    if is_raw:
        parts = [
            _to_str(item.get("Profiel")).strip(),
            _to_str(item.get("Materiaal")).strip(),
            _to_str(item.get("Lengte")).strip(),
        ]
        if any(parts):
            return "raw|" + "|".join(parts)

    part_number = _to_str(item.get("PartNumber")).strip()
    if part_number:
        return f"part|{part_number}"

    parts = [
        _to_str(item.get("Description")).strip(),
        _to_str(item.get("Materiaal")).strip(),
        _to_str(item.get("Aantal") or item.get("St.")).strip(),
    ]
    return "item|" + "|".join(parts)


def _clean_price_text(value: object) -> str:
    text = _to_str(value).strip()
    if not text:
        return ""
    text = (
        text.replace("\u00a0", "")
        .replace(" ", "")
        .replace("€", "")
        .replace("%", "")
    )
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    return text


def _price_decimal(value: object) -> Decimal | None:
    text = _clean_price_text(value)
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _format_price_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    rounded = value.quantize(Decimal("0.01"))
    return f"{rounded:.2f}"


def _clean_vat_rate_text(value: object) -> str:
    text = _clean_price_text(value).replace("%", "").strip()
    if not text:
        return ""
    rate = _price_decimal(text)
    if rate is None or rate < 0 or rate > 100:
        return ""
    rounded = rate.quantize(Decimal("0.01"))
    if rounded == rounded.to_integral_value():
        return str(int(rounded))
    return f"{rounded:.2f}".rstrip("0").rstrip(".")


def _price_summary_label_key(is_raw: bool) -> str:
    return "Profiel" if is_raw else "Description"


def _blank_price_summary_row(
    template_items: Sequence[Mapping[str, object]],
    *,
    is_raw: bool,
    label: str,
    total: Decimal,
) -> Dict[str, object]:
    keys = template_items[0].keys() if template_items else []
    row = {key: "" for key in keys}
    row[_price_summary_label_key(is_raw)] = label
    row[_PRICE_UNIT_KEY] = ""
    row[_PRICE_TOTAL_KEY] = _format_price_decimal(total)
    return row


def _pricing_subtotal(
    items: Sequence[Mapping[str, object]],
    *,
    is_raw: bool,
) -> tuple[Decimal | None, bool]:
    label_key = _price_summary_label_key(is_raw)
    line_sum = Decimal("0")
    has_line_total = False
    offered_total: Decimal | None = None
    for item in items:
        label = _to_str(item.get(label_key)).strip().lower()
        total = _price_decimal(item.get(_PRICE_TOTAL_KEY))
        if total is None:
            continue
        if label == "totaal aangeboden":
            offered_total = total
        elif label not in {
            "subtotaal excl. btw",
            "btw",
            "totaal incl. btw",
        } and not label.startswith("btw "):
            line_sum += total
            has_line_total = True
    if offered_total is not None:
        return offered_total, True
    if has_line_total:
        return line_sum, False
    return None, False


def _normalized_price_summary_label(value: object) -> str:
    return " ".join(_to_str(value).strip().lower().split())


def _is_price_summary_label(value: object) -> bool:
    label = _normalized_price_summary_label(value)
    return (
        label == "subtotaal excl. btw"
        or label == "btw"
        or label.startswith("btw ")
        or label == "totaal incl. btw"
    )


def _price_total_column_key(column_layout: Sequence[Mapping[str, object]]) -> str:
    for column in column_layout:
        key = _to_str(column.get("key")).strip()
        if key.lower() == _PRICE_TOTAL_KEY.lower():
            return key
    return _PRICE_TOTAL_KEY


def _price_summary_label_keys(
    column_layout: Sequence[Mapping[str, object]],
    *,
    is_raw: bool,
) -> List[str]:
    candidates: List[str] = [_price_summary_label_key(is_raw)]
    for column in column_layout:
        key = _to_str(column.get("key")).strip()
        if key and key not in candidates and key not in {_PRICE_UNIT_KEY, _PRICE_TOTAL_KEY}:
            candidates.append(key)
    return candidates


def _split_order_price_summary_rows(
    items: Sequence[Mapping[str, object]],
    column_layout: Sequence[Mapping[str, object]],
    *,
    is_raw: bool,
) -> tuple[List[Mapping[str, object]], object, List[tuple[str, object]]]:
    label_keys = _price_summary_label_keys(column_layout, is_raw=is_raw)
    total_key = _price_total_column_key(column_layout)
    data_items: List[Mapping[str, object]] = []
    subtotal: object = ""
    summary_rows: List[tuple[str, object]] = []

    for item in items:
        label_text = ""
        for key in label_keys:
            text = _to_str(item.get(key)).strip()
            if text:
                label_text = text
                break
        normalized = _normalized_price_summary_label(label_text)
        if normalized == "subtotaal excl. btw":
            subtotal = item.get(total_key, item.get(_PRICE_TOTAL_KEY, ""))
        elif normalized in {"btw", "totaal incl. btw"} or normalized.startswith("btw "):
            summary_rows.append(
                (label_text, item.get(total_key, item.get(_PRICE_TOTAL_KEY, "")))
            )
        else:
            data_items.append(item)

    return data_items, subtotal, summary_rows


def _append_price_summary_rows_to_df(
    df: pd.DataFrame,
    headers: Sequence[str],
    summary_rows: Sequence[tuple[str, object]],
) -> pd.DataFrame:
    if not summary_rows or not headers:
        return df
    price_header = next(
        (
            header
            for header in headers
            if _normalized_price_summary_label(header)
            in {"totaal (\u20ac)", "totaalprijs", "totalprice"}
        ),
        headers[-1],
    )
    blank_row = {header: "" for header in headers}
    rows = [blank_row]
    for label, total in summary_rows:
        row = {header: "" for header in headers}
        row[headers[0]] = label
        row[price_header] = total
        rows.append(row)
    return pd.concat([df, pd.DataFrame(rows)], ignore_index=True)


def _clean_line_pricing_items(
    pricing: Mapping[str, object] | None,
) -> Dict[str, Dict[str, str]]:
    if not isinstance(pricing, Mapping):
        return {}
    raw_items = pricing.get("items")
    if not isinstance(raw_items, Mapping):
        return {}

    cleaned: Dict[str, Dict[str, str]] = {}
    for raw_key, raw_value in raw_items.items():
        key = _to_str(raw_key).strip()
        if not key or not isinstance(raw_value, Mapping):
            continue
        entry = {
            "unit_price": _clean_price_text(raw_value.get("unit_price")),
            "total_price": _clean_price_text(raw_value.get("total_price")),
            "quote_ref": _to_str(raw_value.get("quote_ref")).strip(),
            "note": _to_str(raw_value.get("note")).strip(),
        }
        if any(entry.values()):
            cleaned[key] = entry
    return cleaned


def _pricing_has_values(pricing: Mapping[str, object] | None) -> bool:
    if not pricing:
        return False
    if _clean_price_text(pricing.get("unit_price")) or _clean_price_text(
        pricing.get("total_price")
    ):
        return True
    return bool(_clean_line_pricing_items(pricing))


def _line_pricing_for_item(
    item: Mapping[str, object],
    line_pricing: Mapping[str, Mapping[str, str]],
    *,
    context_kind: str,
) -> Mapping[str, str] | None:
    if not line_pricing:
        return None
    key = _to_str(item.get(_LINE_PRICE_KEY)).strip()
    if not key:
        key = build_order_pricing_item_key(item, context_kind=context_kind)
    candidates = [key]

    context_kind_clean = (_to_str(context_kind) or "productie").strip().lower()
    if context_kind_clean.startswith("brutemateriaal"):
        raw_plain = "|".join(
            [
                _to_str(item.get("Profiel")).strip(),
                _to_str(item.get("Materiaal")).strip(),
                _to_str(item.get("Lengte")).strip(),
            ]
        )
        candidates.append(raw_plain)
    else:
        part_number = _to_str(item.get("PartNumber")).strip()
        if part_number:
            candidates.append(part_number)

    for candidate in candidates:
        if not candidate:
            continue
        value = line_pricing.get(candidate)
        if value is not None:
            return value

    lowered = {stored_key.lower(): value for stored_key, value in line_pricing.items()}
    for candidate in candidates:
        if not candidate:
            continue
        value = lowered.get(candidate.lower())
        if value is not None:
            return value
    return None


def _spare_part_column_layout(*, include_internal: bool = True) -> List[Dict[str, object]]:
    """Return the order table layout for spare-parts documents."""

    columns = [
        {"key": "PartNumber", "label": "Artikel nr.", "width": 16, "justify": "left", "wrap": True, "weight": 1.35},
        {"key": "Description", "label": "Omschrijving", "width": 30, "justify": "left", "wrap": True, "weight": 2.6},
        {"key": "Aantal", "label": "St.", "width": 8, "justify": "right", "numeric": True, "integer": True, "weight": 0.45},
        {"key": "Supplier", "label": "Supplier", "width": 18, "justify": "left", "wrap": True, "weight": 1.15},
        {"key": "Supplier code", "label": "Supplier code", "width": 18, "justify": "left", "wrap": True, "weight": 1.2},
        {"key": "Manufacturer", "label": "Manufacturer", "width": 18, "justify": "left", "wrap": True, "weight": 1.15},
        {"key": "Manufacturer code", "label": "Manufacturer code", "width": 18, "justify": "left", "wrap": True, "weight": 1.2},
    ]
    if include_internal:
        columns.extend(
            [
                {"key": "Bestelgroep", "label": "Bestelgroep", "width": 16, "justify": "left", "wrap": True, "weight": 1.0},
                {"key": "Status", "label": "Status", "width": 16, "justify": "left", "wrap": True, "weight": 0.95},
            ]
        )
    return columns


def _default_priced_column_layout(
    context_kind: str,
    *,
    include_length: bool = False,
    include_spare_internal: bool = True,
) -> List[Dict[str, object]]:
    context_kind_clean = (_to_str(context_kind) or "productie").strip().lower()
    is_raw = context_kind_clean.startswith("brutemateriaal")
    is_spare = context_kind_clean.startswith("spare")
    if is_raw:
        columns = [
            {"key": "Profiel", "label": "Profiel", "width": 20, "justify": "left", "wrap": True, "weight": 2.0},
            {"key": "Materiaal", "label": "Materiaal", "width": 18, "justify": "left", "wrap": True, "weight": 1.6},
            {"key": "Lengte", "label": "Lengte", "width": 10, "justify": "right", "numeric": True, "weight": 0.9},
            {"key": "St.", "label": "St.", "width": 8, "justify": "right", "numeric": True, "integer": True, "weight": 0.7},
            {"key": "kg", "label": "kg", "width": 10, "justify": "right", "numeric": True, "total_weight": True, "weight": 0.8},
        ]
    elif is_spare:
        columns = _spare_part_column_layout(include_internal=include_spare_internal)
    else:
        columns = [
            {"key": "PartNumber", "label": "PartNumber", "width": 22, "justify": "left", "wrap": True, "weight": 1.8},
            {"key": "Description", "label": "Omschrijving", "width": 32, "justify": "left", "wrap": True, "weight": 2.4},
            {"key": "Materiaal", "label": "Materiaal", "width": 16, "justify": "left", "wrap": False, "weight": 1.2},
        ]
        if include_length:
            columns.append({"key": "Lengte", "label": "Lengte", "width": 10, "justify": "right", "numeric": True, "integer": True, "weight": 0.8})
        columns.extend(
            [
                {"key": "Aantal", "label": "St.", "width": 8, "justify": "right", "numeric": True, "integer": True, "weight": 0.7},
                {"key": "Oppervlakte", "label": "m\u00b2/st", "width": 10, "justify": "right", "numeric": True, "total_surface": True, "weight": 0.8},
                {"key": "Gewicht", "label": "kg/st", "width": 10, "justify": "right", "numeric": True, "total_weight": True, "weight": 0.8},
            ]
        )
    columns.extend(
        [
            {"key": _PRICE_UNIT_KEY, "label": "Prijs/st. (\u20ac)", "width": 12, "justify": "right", "numeric": True, "weight": 0.9},
            {"key": _PRICE_TOTAL_KEY, "label": "Totaal (\u20ac)", "width": 14, "justify": "right", "numeric": True, "weight": 1.0},
        ]
    )
    return columns

def _apply_order_pricing(
    items: List[Dict[str, object]],
    pricing: Mapping[str, object] | None,
    *,
    context_kind: str,
    vat_rate: object = "",
) -> tuple[List[Dict[str, object]], Optional[List[Dict[str, object]]]]:
    if not _pricing_has_values(pricing):
        return items, None

    unit_price_text = _clean_price_text(pricing.get("unit_price"))
    total_price_text = _clean_price_text(pricing.get("total_price"))
    line_pricing = _clean_line_pricing_items(pricing)
    priced_items = [dict(item) for item in items]
    is_raw = (_to_str(context_kind).strip().lower()).startswith("brutemateriaal")
    qty_key = "St." if is_raw else "Aantal"

    for item in priced_items:
        line_info = _line_pricing_for_item(
            item,
            line_pricing,
            context_kind=context_kind,
        )
        line_unit_text = (
            _clean_price_text(line_info.get("unit_price")) if line_info else ""
        )
        line_total_text = (
            _clean_price_text(line_info.get("total_price")) if line_info else ""
        )
        effective_unit_text = line_unit_text or unit_price_text
        effective_unit = _price_decimal(effective_unit_text)
        if effective_unit_text:
            item[_PRICE_UNIT_KEY] = effective_unit_text
            qty = _price_decimal(item.get(qty_key))
            if line_total_text:
                item[_PRICE_TOTAL_KEY] = line_total_text
            elif effective_unit is not None and qty is not None:
                item[_PRICE_TOTAL_KEY] = _format_price_decimal(effective_unit * qty)
            else:
                item.setdefault(_PRICE_TOTAL_KEY, "")
        elif line_total_text:
            item[_PRICE_TOTAL_KEY] = line_total_text

    if total_price_text:
        total_row = {key: "" for key in (priced_items[0].keys() if priced_items else [])}
        if is_raw:
            total_row["Profiel"] = "Totaal aangeboden"
        else:
            total_row["Description"] = "Totaal aangeboden"
        total_row[_PRICE_TOTAL_KEY] = total_price_text
        priced_items.append(total_row)

    vat_rate_text = _clean_vat_rate_text(vat_rate)
    if vat_rate_text:
        subtotal, has_offered_total = _pricing_subtotal(priced_items, is_raw=is_raw)
        vat_decimal = _price_decimal(vat_rate_text)
        if subtotal is not None and vat_decimal is not None:
            if not has_offered_total:
                priced_items.append(
                    _blank_price_summary_row(
                        priced_items,
                        is_raw=is_raw,
                        label="Subtotaal excl. BTW",
                        total=subtotal,
                    )
                )
            vat_amount = subtotal * vat_decimal / Decimal("100")
            priced_items.append(
                _blank_price_summary_row(
                    priced_items,
                    is_raw=is_raw,
                    label=f"BTW {vat_rate_text}%",
                    total=vat_amount,
                )
            )
            priced_items.append(
                _blank_price_summary_row(
                    priced_items,
                    is_raw=is_raw,
                    label="Totaal incl. BTW",
                    total=subtotal + vat_amount,
                )
            )

    for item in priced_items:
        item.setdefault(_PRICE_UNIT_KEY, "")
        item.setdefault(_PRICE_TOTAL_KEY, "")
    include_length = (not is_raw) and any(
        _clean_order_cell_text(item.get("Lengte", "")) for item in priced_items
    )
    return priced_items, _default_priced_column_layout(
        context_kind,
        include_length=include_length,
    )


_INVALID_PATH_CHARS = set('<>:"/\\|?*')
_WINDOWS_MAX_PATH = 240


def _sanitize_component(value: object) -> str:
    """Return a filesystem-friendly representation of ``value``."""

    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    text = " ".join(text.split())
    cleaned: List[str] = []
    for ch in text:
        if ch in _INVALID_PATH_CHARS or ord(ch) < 32:
            cleaned.append("_")
        elif ch == os.sep or (os.altsep and ch == os.altsep):
            cleaned.append("-")
        else:
            cleaned.append(ch)
    return "".join(cleaned).strip(" .-_")


_NATURAL_SORT_RE = re.compile(r"(\d+)")


def _natural_pdf_name_key(path: str) -> Tuple[Tuple[int, object], ...]:
    """Return a natural, case-insensitive sort key for PDF filenames."""

    name = os.path.basename(_to_str(path)).casefold()
    key_parts: List[Tuple[int, object]] = []
    for part in _NATURAL_SORT_RE.split(name):
        if not part:
            continue
        if part.isdigit():
            try:
                key_parts.append((0, int(part)))
            except Exception:
                key_parts.append((1, part))
        else:
            key_parts.append((1, part))
    return tuple(key_parts)


def _pdf_match_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", _to_str(value))
    text = text.encode("ASCII", "ignore").decode("ASCII")
    text = re.sub(r"[^0-9a-zA-Z]+", " ", text).casefold()
    return " ".join(text.split())


def _pdf_section_match_score(production: str, identifier: str) -> int:
    production_key = _pdf_match_text(production)
    identifier_key = _pdf_match_text(identifier)
    if not production_key or not identifier_key:
        return 0
    if production_key == identifier_key:
        return 10000 + len(identifier_key)
    if identifier_key in production_key:
        return 1000 + len(identifier_key)
    if production_key in identifier_key:
        return 500 + len(production_key)
    return 0


def _assign_pdf_sections(
    productions: Sequence[str],
    sections: Sequence[PdfWorkDossierSection],
) -> Dict[str, int]:
    assigned: Dict[str, int] = {}
    active_sections = [
        (index, section)
        for index, section in enumerate(sections)
        if getattr(section, "enabled", True)
    ]
    for production in productions:
        best_index: int | None = None
        best_score = 0
        for index, section in active_sections:
            for identifier in getattr(section, "identifiers", []) or []:
                score = _pdf_section_match_score(production, identifier)
                if score <= 0:
                    continue
                if score > best_score or (
                    score == best_score
                    and (best_index is None or index < best_index)
                ):
                    best_score = score
                    best_index = index
        if best_index is not None:
            assigned[production] = best_index
    return assigned


def _is_order_interleaf_pdf(filename: str, *, include_offers: bool = False) -> bool:
    name = os.path.basename(_to_str(filename)).strip().casefold()
    if not name.endswith(".pdf"):
        return False
    if name.startswith("bestelbon") or name.startswith("standaard"):
        return True
    return include_offers and name.startswith("offerteaanvraag")


def _find_order_interleaf_pdfs(
    root_dir: str | None,
    production: str,
    *,
    include_offers: bool = False,
) -> List[str]:
    """Find generated order PDFs for a production folder."""

    root = _to_str(root_dir).strip()
    if not root:
        return []
    production_text = _to_str(production).strip()
    if not production_text:
        return []

    candidates: List[str] = []
    seen: set[str] = set()

    direct_dir = os.path.join(root, production_text)
    search_roots = [direct_dir] if os.path.isdir(direct_dir) else []
    if not search_roots and os.path.isdir(root):
        production_key = production_text.casefold()
        for dirpath, dirnames, _filenames in os.walk(root):
            basename = os.path.basename(dirpath).casefold()
            if basename == production_key:
                search_roots.append(dirpath)
                dirnames[:] = []

    for search_root in search_roots:
        try:
            filenames = os.listdir(search_root)
        except OSError:
            continue
        for filename in filenames:
            if not _is_order_interleaf_pdf(filename, include_offers=include_offers):
                continue
            path = os.path.join(search_root, filename)
            if not os.path.isfile(path):
                continue
            key = os.path.abspath(path).casefold()
            if key in seen:
                continue
            seen.add(key)
            candidates.append(path)

    candidates.sort(key=_natural_pdf_name_key)
    return candidates


def _unique_pdf_paths(paths: Sequence[str]) -> List[str]:
    result: List[str] = []
    seen: set[str] = set()
    for path in paths:
        clean = _to_str(path).strip()
        if not clean:
            continue
        key = os.path.abspath(clean).casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(clean)
    return result


def _slugify_name(value: object, fallback: str) -> str:
    """Slugify ``value`` similar to export bundle directories."""

    normalized = unicodedata.normalize("NFKD", "" if value is None else str(value))
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.lower()
    ascii_text = ascii_text.replace(" ", "-")
    ascii_text = re.sub(r"[^a-z0-9-]", "", ascii_text)
    ascii_text = re.sub(r"-+", "-", ascii_text).strip("-")
    if len(ascii_text) > 40:
        ascii_text = ascii_text[:40].rstrip("-")
    if ascii_text:
        return ascii_text
    fallback_norm = unicodedata.normalize("NFKD", fallback)
    ascii_fallback = fallback_norm.encode("ascii", "ignore").decode("ascii") or fallback
    ascii_fallback = re.sub(r"[^a-zA-Z0-9-]", "", ascii_fallback)
    ascii_fallback = ascii_fallback.lower()[:40].rstrip("-")
    return ascii_fallback or "export"


def _fit_filename_within_path(
    directory: str,
    filename: str,
    *,
    max_path: int = _WINDOWS_MAX_PATH,
) -> str:
    """Return ``filename`` possibly shortened so ``directory/filename`` fits ``max_path``.

    On Windows, opening files with paths longer than 260 characters raises ``FileNotFoundError``.
    ``max_path`` defaults to a slightly smaller value (240) to provide some safety margin for
    runtime conversions that may add characters (e.g. via extended paths). When the composed path
    exceeds the limit, the base filename is truncated and suffixed with an 8-character hash so the
    resulting name remains unique and deterministic while staying within the limit. ``max_path`` can
    be overridden (primarily for tests). When the directory path already exceeds the limit the
    function raises :class:`OSError` as there is no filename that could satisfy the constraint.
    """

    if max_path is None:
        return filename

    directory_abs = os.path.abspath(directory)
    full_abs = os.path.join(directory_abs, filename)
    if len(full_abs) <= max_path:
        return filename

    stem, ext = os.path.splitext(filename)
    if not stem:
        stem = "_"

    available = max_path - len(directory_abs) - len(os.sep) - len(ext)
    if available <= 0:
        raise OSError(
            "Basispad is te lang voor het genereren van exportbestanden. Verkort het exportpad."
        )

    digest = hashlib.sha1(full_abs.encode("utf-8")).hexdigest()[:8]
    if available <= len(digest):
        safe_stem = digest[:available]
    else:
        keep = max(1, available - len(digest) - 1)
        safe_stem = f"{stem[:keep].rstrip(' _-.')}_{digest}" if keep < len(stem) else f"{stem}_{digest}"

    safe_filename = f"{safe_stem}{ext}"
    safe_abs = os.path.join(directory_abs, safe_filename)
    if len(safe_abs) > max_path and available > len(digest):
        safe_stem = digest[: available]
        safe_filename = f"{safe_stem}{ext}"

    return safe_filename


def _create_combined_output_dir(
    base_dir: str,
    project_number: Optional[str],
    project_name: Optional[str],
    *,
    timestamp: Optional[datetime.datetime] = None,
) -> str:
    """Create a combined PDF export directory inside ``base_dir``."""

    os.makedirs(base_dir, exist_ok=True)
    ts = timestamp or datetime.datetime.now()
    ts_token = ts.strftime("%Y-%m-%dT%H%M%S")
    pn_clean = _sanitize_component(project_number) or "project"
    slug = _slugify_name(project_name, pn_clean)
    name_parts = ["Combined pdf", ts_token, pn_clean]
    if slug and slug != pn_clean.lower():
        name_parts.append(slug)
    folder_name = "_".join(part for part in name_parts if part)
    base_path = os.path.join(base_dir, folder_name)
    candidate = base_path
    index = 1
    while os.path.exists(candidate):
        candidate = f"{base_path}_{index}"
        index += 1
    os.makedirs(candidate, exist_ok=True)
    return candidate


def _export_bom_workbook(bom_df: pd.DataFrame, dest: str, filename: str) -> str:
    """Write the processed BOM dataframe to an Excel workbook."""

    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"
    target_path = os.path.join(dest, filename)
    export_df = bom_df.reset_index(drop=True).copy()
    # Drop status-related columns that are only useful inside the app.
    to_drop = [col for col in _BOM_STATUS_COLUMNS if col in export_df.columns]
    if to_drop:
        export_df = export_df.drop(columns=to_drop)

    # Normalise quantity column naming to ``QTY.`` and drop aliases.
    qty_aliases: Tuple[str, ...] = ("QTY.", "Qty.", "Qty", "Quantity", "Aantal")
    qty_columns = [col for col in qty_aliases if col in export_df.columns]
    if "QTY." not in export_df.columns:
        if qty_columns:
            export_df = export_df.rename(columns={qty_columns[0]: "QTY."})
        else:
            export_df["QTY."] = ""
    for alias in qty_columns:
        if alias == "QTY." or alias not in export_df.columns:
            continue
        source = export_df[alias]
        destination = export_df["QTY."]
        dest_str = destination.astype(str)
        missing_mask = dest_str.str.strip().isin(("", "nan"))
        export_df.loc[missing_mask, "QTY."] = source[missing_mask]
        export_df = export_df.drop(columns=alias)

    # Ensure all primary BOM columns are present and appear first.
    for column in _BOM_EXPORT_BASE_COLUMNS:
        if column not in export_df.columns:
            export_df[column] = ""
    ordered_columns = [c for c in _BOM_EXPORT_BASE_COLUMNS if c in export_df.columns]
    remaining_columns = [c for c in export_df.columns if c not in ordered_columns]
    export_df = export_df[ordered_columns + remaining_columns]

    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="BOM")
        if Alignment is not None and get_column_letter is not None:
            ws = writer.sheets["BOM"]
            alignment = Alignment(wrap_text=False, vertical="top")
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment
            for col_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[column_letter]:
                    value = cell.value
                    cell_length = len(str(value)) if value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                ws.column_dimensions[column_letter].width = min(max(12, max_length + 2), 80)

    return target_path


def make_bom_export_filename(
    bom_source_path: Optional[str],
    date_iso: str,
    transform: Callable[[str], str],
) -> str:
    """Return a normalized filename for exporting the processed BOM workbook."""

    source_stem = ""
    if bom_source_path:
        source_stem = Path(bom_source_path).stem
        match = re.search(r"(.*?\bBOM\b)", source_stem, flags=re.IGNORECASE)
        if match:
            source_stem = match.group(1)
        source_stem = source_stem.rstrip(" -_.")
    stem = source_stem or "BOM-FileHopper-Export"
    stem_with_date = f"{stem}-{date_iso}"
    filename = f"{stem_with_date}.xlsx"
    return transform(filename)


def find_related_bom_exports(
    bom_source_path: Optional[str],
    file_index: Mapping[str, Sequence[str]],
) -> List[str]:
    """Return export files whose stem appears in the BOM filename."""

    if not bom_source_path:
        return []
    stem = Path(bom_source_path).stem.lower()
    if not stem:
        return []

    def _component_contains(haystack: str, needle: str) -> bool:
        idx = haystack.find(needle)
        while idx != -1:
            if idx > 0 and haystack[idx - 1].isalnum():
                idx = haystack.find(needle, idx + 1)
                continue
            end = idx + len(needle)
            if end < len(haystack) and haystack[end].isalnum():
                idx = haystack.find(needle, idx + 1)
                continue
            return True
        return False

    def _candidate_allowed(value: str) -> bool:
        return bool(value) and (len(value) >= 4 or any(ch.isdigit() for ch in value))

    bom_candidates = [stem]
    bom_match = re.search(r"(.*?)(?:[-_.\s]*\bBOM\b)", stem, flags=re.IGNORECASE)
    if bom_match:
        base = bom_match.group(1).rstrip(" -_.")
        if base:
            bom_candidates.append(base)
    bom_candidates = [
        candidate
        for candidate in dict.fromkeys(bom_candidates)
        if _candidate_allowed(candidate)
    ]

    matches: List[str] = []
    seen: set[str] = set()
    for key in sorted(file_index.keys(), key=len, reverse=True):
        if not key:
            continue
        key_lower = key.lower()
        if not _candidate_allowed(key_lower):
            continue
        if not any(
            _component_contains(candidate, key_lower)
            or _component_contains(key_lower, candidate)
            for candidate in bom_candidates
        ):
            continue
        for src_file in file_index.get(key, []):
            if src_file not in seen:
                matches.append(src_file)
                seen.add(src_file)
    return matches


def _normalize_crop_box(
    crop: object, width: int, height: int
) -> Optional[tuple[int, int, int, int]]:
    """Validate and clamp crop data against an image size."""

    if not crop or width <= 0 or height <= 0:
        return None

    left = top = 0
    right, bottom = width, height

    try:
        if isinstance(crop, dict):
            left = int(float(crop.get("left", 0)))
            top = int(float(crop.get("top", 0)))
            right = int(float(crop.get("right", width)))
            bottom = int(float(crop.get("bottom", height)))
        elif isinstance(crop, (list, tuple)) and len(crop) == 4:
            left, top, right, bottom = [int(float(v)) for v in crop]
        else:
            return None
    except Exception:
        return None

    left = max(0, min(width, left))
    top = max(0, min(height, top))
    right = max(left + 1, min(width, right))
    bottom = max(top + 1, min(height, bottom))

    if right <= left or bottom <= top:
        return None
    return left, top, right, bottom

def _parse_qty(val: object) -> int:
    """Parse quantity values to int within [1, 999]."""
    s = _to_str(val).strip()
    if not s:
        return 1
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.]+", "", s)
    try:
        q = int(float(s))
    except Exception:
        q = 1
    return max(1, min(999, q))


def _coerce_integer_like(value: object) -> object:
    """Return ``value`` as an int when it represents a whole number."""

    if value in ("", None):
        return ""
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value)) if math.isfinite(value) else value
    text = _to_str(value).strip()
    if not text:
        return ""
    text = text.replace(",", ".")
    try:
        number = float(text)
    except Exception:
        return value
    if not math.isfinite(number):
        return value
    return int(round(number))


def _format_integer_like(value: object) -> str:
    """Format integer-like values without decimal places."""

    coerced = _coerce_integer_like(value)
    if coerced in ("", None):
        return ""
    if isinstance(coerced, int):
        return str(coerced)
    return _to_str(coerced)


def _prefix_for_doc_type(doc_type: str) -> str:
    """Return standard document number prefix for a ``doc_type``.

    ``"Bestelbon"`` uses ``"BB-"`` while ``"Offerteaanvraag"`` uses ``"OFF-"``.
    Unknown types return an empty prefix.
    """
    t = (doc_type or "").strip().lower()
    if t.startswith("standaard"):
        return "BOM-"
    if t.startswith("bestel"):
        return "BB-"
    if t.startswith("offerte"):
        return "OFF-"
    return ""


def _normalize_doc_number(value: object, doc_type: object) -> str:
    """Return a cleaned document number for a given ``doc_type``.

    The GUI provides placeholder prefixes such as ``"BB-"``. When the user
    pastes a value that already contains the prefix the placeholder should be
    replaced instead of duplicated (``"BB-BB123"`` → ``"BB-123"``).
    """

    doc_num = _to_str(value).strip()
    if not doc_num:
        return ""

    prefix = _prefix_for_doc_type(_to_str(doc_type))
    if not prefix:
        return doc_num

    prefix_upper = prefix.upper()
    doc_upper = doc_num.upper()
    prefix_compact = re.sub(r"[^A-Z0-9]", "", prefix_upper)

    if doc_upper.startswith(prefix_upper):
        remainder = doc_num[len(prefix) :]
        remainder_stripped = remainder.lstrip(" -_")
        remainder_upper = remainder_stripped.upper()
        if remainder_upper.startswith(prefix_upper):
            remainder = remainder_stripped[len(prefix) :]
            doc_num = prefix + remainder.lstrip(" -_")
        elif prefix_compact and remainder_upper.startswith(prefix_compact):
            remainder = remainder_stripped[len(prefix_compact) :]
            doc_num = prefix + remainder.lstrip(" -_")
    elif prefix_compact and doc_upper.startswith(prefix_compact):
        remainder = doc_num[len(prefix_compact) :]
        doc_num = prefix + remainder.lstrip(" -_")

    normalized_compact = re.sub(r"[^A-Z0-9]", "", doc_num.upper())
    if prefix_compact and normalized_compact == prefix_compact:
        return ""

    return doc_num


DOCUMENT_FILENAME_PROFILE_STANDARD = "standard"
DOCUMENT_FILENAME_PROFILE_SHORT = "short"
DOCUMENT_FILENAME_PROFILE_COMPACT = "compact"
DOCUMENT_FILENAME_PROFILE_CUSTOM = "custom"
DOCUMENT_FILENAME_PROFILES = {
    DOCUMENT_FILENAME_PROFILE_STANDARD,
    DOCUMENT_FILENAME_PROFILE_SHORT,
    DOCUMENT_FILENAME_PROFILE_COMPACT,
    DOCUMENT_FILENAME_PROFILE_CUSTOM,
}
DOCUMENT_FILENAME_SEPARATOR_MAP = {
    "underscore": "_",
    "dash": "-",
    "none": "",
}


def normalize_document_filename_profile(value: object) -> str:
    """Return a supported document filename profile key."""

    text = _to_str(value).strip().lower()
    if text in DOCUMENT_FILENAME_PROFILES:
        return text
    return DOCUMENT_FILENAME_PROFILE_STANDARD


def normalize_document_filename_separator(value: object) -> str:
    """Return a supported document filename separator key."""

    text = _to_str(value).strip().lower()
    if text in DOCUMENT_FILENAME_SEPARATOR_MAP:
        return text
    if text == "_":
        return "underscore"
    if text == "-":
        return "dash"
    if text in {"", "geen"}:
        return "none"
    return "underscore"


def _format_doc_number_for_filename(
    doc_number: object,
    doc_type: object,
    *,
    compact: bool = False,
) -> str:
    """Return the filename-safe document number component."""

    normalized = _normalize_doc_number(doc_number, doc_type)
    if not normalized:
        return ""
    if compact:
        normalized = re.sub(r"[\s\-_]+", "", normalized)
    return _sanitize_component(normalized)


def format_document_number_for_display(
    doc_number: object,
    doc_type: object,
    *,
    compact: bool = False,
) -> str:
    """Return the document number as it should appear in PDF/XLSX headers."""

    normalized = _normalize_doc_number(doc_number, doc_type)
    if not normalized:
        return ""
    if compact:
        normalized = re.sub(r"[\s\-_]+", "", normalized)
    return normalized


def _join_filename_parts(parts: Sequence[str], separator: str) -> str:
    cleaned = [part for part in (_sanitize_component(part) for part in parts) if part]
    if not cleaned:
        return ""
    if separator:
        return separator.join(cleaned)
    return "".join(cleaned)


def build_document_export_basename(
    doc_type: object,
    doc_number: object = "",
    context_label: object = "",
    export_date: object = None,
    *,
    profile: object = DOCUMENT_FILENAME_PROFILE_STANDARD,
    show_doc_type: bool = True,
    show_doc_number: bool = True,
    show_context: bool = True,
    show_date: bool = True,
    compact_doc_number: bool = False,
    separator: object = "underscore",
    extra_context_label: object = "",
) -> str:
    """Return the export basename for order-related PDF/XLSX files."""

    doc_type_text = _to_str(doc_type).strip() or "Bestelbon"
    context_text = _to_str(context_label).strip()
    extra_context_text = _to_str(extra_context_label).strip()
    date_text = _to_str(export_date).strip() or datetime.date.today().strftime("%Y-%m-%d")
    profile_key = normalize_document_filename_profile(profile)
    separator_key = normalize_document_filename_separator(separator)
    separator_text = DOCUMENT_FILENAME_SEPARATOR_MAP[separator_key]

    doc_num_default = _format_doc_number_for_filename(doc_number, doc_type_text, compact=False)
    doc_num_compact = _format_doc_number_for_filename(doc_number, doc_type_text, compact=True)

    def _standard_name() -> str:
        return _join_filename_parts(
            [
                doc_type_text,
                doc_num_default,
                context_text,
                extra_context_text,
                date_text,
            ],
            "_",
        )

    if profile_key == DOCUMENT_FILENAME_PROFILE_SHORT:
        if doc_num_default:
            return doc_num_default
        profile_key = DOCUMENT_FILENAME_PROFILE_STANDARD
    elif profile_key == DOCUMENT_FILENAME_PROFILE_COMPACT:
        if doc_num_compact:
            return doc_num_compact
        profile_key = DOCUMENT_FILENAME_PROFILE_STANDARD

    if profile_key == DOCUMENT_FILENAME_PROFILE_STANDARD:
        basename = _standard_name()
        return basename or "document"

    doc_num_custom = doc_num_compact if compact_doc_number else doc_num_default
    parts: List[str] = []
    if show_doc_type:
        parts.append(doc_type_text)
    if show_doc_number and doc_num_custom:
        parts.append(doc_num_custom)
    if show_context:
        if context_text:
            parts.append(context_text)
        if extra_context_text:
            parts.append(extra_context_text)
    if show_date and date_text:
        parts.append(date_text)

    basename = _join_filename_parts(parts, separator_text)
    if basename:
        return basename
    fallback = _standard_name()
    return fallback or "document"


def _should_place_remark_in_delivery_block(
    *,
    order_remark_has_content: bool,
    doc_type_text_slug: str,
    is_standaard_doc: bool,
    delivery: DeliveryAddress | None,
) -> bool:
    """Return :data:`True` when remarks belong in the delivery block."""

    doc_type_is_export = "export" in doc_type_text_slug

    return (
        order_remark_has_content
        and doc_type_is_export
        and not is_standaard_doc
    )


FINISH_KEY_PREFIX = "finish::"
PRODUCTION_KEY_PREFIX = "production::"
OPTICUTTER_KEY_PREFIX = "opticutter::"
SPARE_PART_KEY_PREFIX = "sparepart::"
OPTICUTTER_DEFAULT_SUFFIX = "::Opticutter"
SPARE_PART_DEFAULT_PREFIX = "Spare Parts::"


def _normalize_finish_folder(value: object) -> str:
    """Return a filesystem-friendly folder component for finish/RAL names."""

    text = _to_str(value).strip()
    if not text:
        text = "_Onbekend"
    text = re.sub(r"[\\/:]+", "-", text)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^0-9A-Za-z._ \-]+", "_", text)
    text = text.strip(" .-_")
    if not text:
        text = "_Onbekend"
    return text


def _selection_key(kind: str, identifier: str) -> str:
    """Return a disambiguated key for supplier/doc selections."""

    identifier = _to_str(identifier)
    if kind == "finish":
        return f"{FINISH_KEY_PREFIX}{identifier}"
    if kind == "opticutter":
        return f"{OPTICUTTER_KEY_PREFIX}{identifier}"
    if kind == "sparepart":
        return f"{SPARE_PART_KEY_PREFIX}{identifier}"
    return f"{PRODUCTION_KEY_PREFIX}{identifier}"


def make_production_selection_key(name: str) -> str:
    """Return a stable selection key for a production name."""

    return _selection_key("production", name)


def make_finish_selection_key(finish_key: str) -> str:
    """Return a stable selection key for a finish combination."""

    return _selection_key("finish", finish_key)


def make_opticutter_selection_key(name: str) -> str:
    """Return a stable selection key for Opticutter raw material orders."""

    return _selection_key("opticutter", name)


def make_spare_part_selection_key(group_key: str) -> str:
    """Return a stable selection key for a spare-parts order group."""

    return _selection_key("sparepart", group_key)


def make_opticutter_default_key(name: str) -> str:
    """Return the SuppliersDB default key for Opticutter raw material orders."""

    base = _to_str(name)
    return f"{base}{OPTICUTTER_DEFAULT_SUFFIX}" if base else OPTICUTTER_DEFAULT_SUFFIX


def make_spare_part_default_key(group_key: str) -> str:
    """Return the SuppliersDB default key for a spare-parts order group."""

    base = _to_str(group_key)
    return f"{SPARE_PART_DEFAULT_PREFIX}{base}" if base else SPARE_PART_DEFAULT_PREFIX.rstrip(":")


def parse_selection_key(key: str) -> Tuple[str, str]:
    """Return the selection kind and identifier."""

    if key.startswith(FINISH_KEY_PREFIX):
        return "finish", key[len(FINISH_KEY_PREFIX) :]
    if key.startswith(PRODUCTION_KEY_PREFIX):
        return "production", key[len(PRODUCTION_KEY_PREFIX) :]
    if key.startswith(OPTICUTTER_KEY_PREFIX):
        return "opticutter", key[len(OPTICUTTER_KEY_PREFIX) :]
    if key.startswith(SPARE_PART_KEY_PREFIX):
        return "sparepart", key[len(SPARE_PART_KEY_PREFIX) :]
    # Fallback for legacy keys without explicit prefix.
    return "production", key


def _clean_document_group_map(
    group_map: Mapping[str, str] | None,
) -> Dict[str, str]:
    cleaned: Dict[str, str] = {}
    for raw_follower, raw_master in (group_map or {}).items():
        follower = _to_str(raw_follower).strip()
        master = _to_str(raw_master).strip()
        if not follower or not master or follower == master:
            continue
        follower_kind, _ = parse_selection_key(follower)
        master_kind, _ = parse_selection_key(master)
        if follower_kind != master_kind or follower_kind not in {
            "production",
            "finish",
            "sparepart",
        }:
            continue
        cleaned[follower] = master
    return cleaned


def _resolve_document_group_root(
    selection_key: str,
    group_map: Mapping[str, str],
) -> str:
    current = selection_key
    seen = {selection_key}
    while True:
        parent = _to_str(group_map.get(current)).strip()
        if not parent or parent == current:
            return current
        if parent in seen:
            return selection_key
        seen.add(parent)
        current = parent


def _build_grouped_document_jobs(
    candidates: Sequence[OrderDocumentCandidate],
    group_map: Mapping[str, str] | None,
) -> List[OrderDocumentJob]:
    cleaned_group_map = _clean_document_group_map(group_map)
    candidate_by_key = {candidate.selection_key: candidate for candidate in candidates}
    grouped_members: Dict[str, List[OrderDocumentCandidate]] = {}

    for candidate in candidates:
        root_key = _resolve_document_group_root(candidate.selection_key, cleaned_group_map)
        root_candidate = candidate_by_key.get(root_key)
        follower_kind, _ = parse_selection_key(candidate.selection_key)
        root_kind, _ = parse_selection_key(root_key)
        if root_candidate is None or follower_kind != root_kind:
            root_key = candidate.selection_key
            root_candidate = candidate
        grouped_members.setdefault(root_key, []).append(candidate)

    jobs: List[OrderDocumentJob] = []
    for root_key, members in grouped_members.items():
        master = candidate_by_key.get(root_key) or members[0]
        ordered_members = [master] + [
            member for member in members if member.selection_key != master.selection_key
        ]
        master_kind, _ = parse_selection_key(master.selection_key)
        en1090_states = {bool(member.en1090_required) for member in ordered_members}
        if master_kind == "production" and len(en1090_states) > 1:
            raise ValueError(
                "Gekoppelde producties moeten dezelfde EN 1090-instelling hebben."
            )
        jobs.append(
            OrderDocumentJob(
                target_dir=master.target_dir,
                context_for_filename=master.filename_context,
                doc_type=master.doc_type,
                doc_num=master.doc_num,
                doc_num_display=master.doc_num_display,
                supplier=master.supplier,
                delivery=master.delivery,
                order_remark=master.order_remark,
                sections=[
                    OrderDocumentSection(
                        context_label=member.context_label,
                        context_kind=member.context_kind,
                        items=list(member.items),
                        selection_key=member.selection_key,
                        total_surface_m2=member.total_surface_m2,
                        total_weight_kg=member.total_weight_kg,
                        column_layout=(
                            [dict(col) for col in member.column_layout]
                            if member.column_layout
                            else None
                        ),
                    )
                    for member in ordered_members
                ],
                en1090_required=bool(en1090_states and True in en1090_states),
            )
        )
    return jobs


def _parse_weight_kg(value: object) -> float | None:
    """Parse a textual kilogram value to float."""

    text = _to_str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    text = text.replace(",", ".")
    cleaned = []
    for ch in text:
        if ch.isdigit():
            cleaned.append(ch)
        elif ch in "+-":
            if not cleaned:
                cleaned.append(ch)
        elif ch == ".":
            cleaned.append(".")
    if not cleaned:
        return None
    candidate = "".join(cleaned)
    if candidate.count(".") > 1:
        first = candidate.find(".")
        candidate = candidate[: first + 1] + candidate[first + 1 :].replace(".", "")
    if candidate in {"", "+", "-", ".", "+.", "-."}:
        return None
    try:
        return float(candidate)
    except Exception:
        return None


def _collect_opticutter_profile_stats(
    bom_df: pd.DataFrame,
) -> Dict[tuple[str, str, str], OpticutterProfileStats]:
    """Aggregate length and weight per (profile, material, production)."""

    stats: Dict[tuple[str, str, str], OpticutterProfileStats] = defaultdict(
        OpticutterProfileStats
    )
    for _, row in bom_df.iterrows():
        profile_name = _to_str(row.get("Profile")).strip()
        if not profile_name:
            continue
        material_name = _to_str(row.get("Materiaal") or row.get("Material")).strip()
        production_name = _to_str(row.get("Production")).strip() or "_Onbekend"
        weight_each = _parse_weight_kg(row.get("Gewicht"))
        if weight_each is None or weight_each <= 0:
            continue
        qty_raw = row.get("Aantal", 0)
        try:
            qty = int(float(qty_raw))
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        length_mm: int | None = None
        length_value = row.get("Length profile mm")
        if length_value is not None and not pd.isna(length_value):
            try:
                length_mm = int(round(float(length_value)))
            except Exception:
                length_mm = None
        if length_mm is None or length_mm <= 0:
            length_mm = parse_length_to_mm(row.get("Length profile"))
        if length_mm is None or length_mm <= 0:
            continue
        key = (profile_name, material_name, production_name)
        entry = stats[key]
        entry.total_length_mm += float(length_mm) * qty
        entry.total_weight_kg += float(weight_each) * qty
    return stats


def _format_weight_kg(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}"


@dataclass(frozen=True, slots=True)
class _StandardOrderColumn:
    key: str
    pdf_label: str
    excel_label: str
    width_fraction: float
    align: str = "RIGHT"
    small: bool = True
    wrap: bool = False


_STANDARD_ORDER_COLUMNS: Tuple[_StandardOrderColumn, ...] = (
    _StandardOrderColumn(
        "PartNumber",
        "PartNumber",
        "PartNumber",
        0.22,
        align="LEFT",
        small=False,
        wrap=True,
    ),
    _StandardOrderColumn(
        "Description",
        "Omschrijving",
        "Description",
        0.40,
        align="LEFT",
        small=False,
        wrap=True,
    ),
    _StandardOrderColumn("Materiaal", "Materiaal", "Materiaal", 0.14),
    _StandardOrderColumn("Lengte", "Lengte", "Lengte", 0.10),
    _StandardOrderColumn("Aantal", "St.", "Aantal", 0.06),
    _StandardOrderColumn("Oppervlakte", "m\u00b2/st", "m\u00b2/st", 0.09),
    _StandardOrderColumn("Gewicht", "kg/st", "kg/st", 0.09),
)
_STANDARD_ORDER_TOTAL_KEYS = {"Oppervlakte", "Gewicht"}


def _standard_order_column_item_has_value(
    column: _StandardOrderColumn,
    item: Mapping[str, object],
) -> bool:
    value = item.get(column.key, "")
    if column.key in _STANDARD_ORDER_TOTAL_KEYS:
        value = _num_to_2dec(value)
    return bool(_clean_order_cell_text(value))


def _standard_order_visible_columns(
    items: Sequence[Mapping[str, object]],
    *,
    total_surface_m2: float | None,
    total_weight_kg: float | None,
) -> List[_StandardOrderColumn]:
    visible: List[_StandardOrderColumn] = []
    for column in _STANDARD_ORDER_COLUMNS:
        has_value = any(
            _standard_order_column_item_has_value(column, item) for item in items
        )
        if column.key == "Oppervlakte" and total_surface_m2 is not None:
            has_value = True
        elif column.key == "Gewicht" and total_weight_kg is not None:
            has_value = True
        if has_value:
            visible.append(column)

    has_total_row = total_surface_m2 is not None or total_weight_kg is not None
    if has_total_row and (
        not visible
        or visible[0].key in _STANDARD_ORDER_TOTAL_KEYS
        or visible[0].align.upper() != "LEFT"
    ):
        visible.insert(0, _STANDARD_ORDER_COLUMNS[0])
    if not visible:
        visible.append(_STANDARD_ORDER_COLUMNS[0])
    return visible


def _standard_order_column_widths(
    columns: Sequence[_StandardOrderColumn],
    items: Sequence[Mapping[str, object]],
    *,
    total_width: float,
    small_font_size: float,
    header_font_size: float,
) -> List[float]:
    if not columns:
        return []

    fractions = {column.key: column.width_fraction for column in _STANDARD_ORDER_COLUMNS}
    visible_keys = {column.key for column in columns}
    if {"PartNumber", "Description"}.issubset(visible_keys):
        non_empty_desc_count = sum(
            1 for item in items if _clean_order_cell_text(item.get("Description", ""))
        )
        empty_desc_ratio = (
            1.0 - (non_empty_desc_count / len(items)) if items else 0.0
        )
        extra_pn_frac = 0.12 * max(0.0, min(1.0, empty_desc_ratio))
        fractions["PartNumber"] += extra_pn_frac
        fractions["Description"] -= extra_pn_frac

    weights = [fractions[column.key] for column in columns]
    minimums: List[float] = []
    for column in columns:
        if column.key == "Description":
            minimums.append(max(_order_table_header_min_width(column.pdf_label, header_font_size), 42 * mm))
            continue
        values = [item.get(column.key, "") for item in items]
        minimums.append(
            _order_table_column_min_width(
                column.pdf_label,
                values,
                header_font_size=header_font_size,
                body_font_size=small_font_size if column.small else 9,
                wrap=column.wrap,
                max_width=total_width * (0.34 if column.key == "PartNumber" else 0.24),
            )
        )

    return _weighted_widths_with_minimums(total_width, weights, minimums)


def _build_standard_order_excel_dataframe(
    items: Sequence[Mapping[str, object]],
    *,
    total_surface_m2: float | None,
    total_weight_kg: float | None,
) -> pd.DataFrame:
    columns = _standard_order_visible_columns(
        items,
        total_surface_m2=total_surface_m2,
        total_weight_kg=total_weight_kg,
    )
    df = pd.DataFrame(items, columns=[column.key for column in columns])
    df = df.rename(
        columns={
            column.key: column.excel_label
            for column in columns
            if column.excel_label != column.key
        }
    )

    if total_surface_m2 is not None or total_weight_kg is not None:
        total_row = {column.excel_label: "" for column in columns}
        total_row[columns[0].excel_label] = "Totaal"
        for column in columns:
            if column.key == "Oppervlakte" and total_surface_m2 is not None:
                total_row[column.excel_label] = _format_weight_kg(total_surface_m2)
            elif column.key == "Gewicht" and total_weight_kg is not None:
                total_row[column.excel_label] = _format_weight_kg(total_weight_kg)
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    return df


def _standard_order_left_columns(columns: Sequence[_StandardOrderColumn]) -> set[str]:
    return {
        column.excel_label
        for column in columns
        if column.align.upper() != "RIGHT"
    }


def _standard_order_wrap_columns(columns: Sequence[_StandardOrderColumn]) -> set[str]:
    return {column.excel_label for column in columns if column.wrap}


def _compute_opticutter_order_exports(
    opticutter_prod: OpticutterProductionExport,
    stats_map: Mapping[tuple[str, str, str], OpticutterProfileStats],
) -> OpticutterOrderComputation:
    scenario_rows: List[Dict[str, object]] = []
    piece_rows: List[Dict[str, object]] = []
    order_rows: List[Dict[str, object]] = []
    raw_items: List[Dict[str, object]] = []
    has_valid_bars = False
    total_bars = 0
    total_weight_known = False
    total_weight_kg = 0.0
    selection_count = 0

    for selection in opticutter_prod.selections:
        selection_count += 1
        profile = selection.profile
        result = selection.result
        stock_length = selection.stock_length_mm

        remark_lines: List[str] = []
        if selection.choice == "input":
            remark_lines.append("Per stuk zagen (inputlengte).")
        elif result is None:
            remark_lines.append("Geen scenario berekend.")
        elif result.dropped_pieces:
            remark_lines.append(
                f"Niet mogelijk: {result.dropped_pieces} stuk(ken) zijn te lang."
            )
        if selection.blockers:
            remark_lines.append("Blokkerende stukken:")
            remark_lines.extend(f"- {text}" for text in selection.blockers)
        elif result is not None and not result.dropped_pieces:
            remark_lines.append(f"Totale restlengte: {result.waste_mm:.0f} mm.")
        scenario_remark = "\n".join(remark_lines).strip()

        bars_value = (
            result.bars if result is not None and not result.dropped_pieces else None
        )
        waste_pct_value = (
            round(result.waste_pct, 1)
            if result is not None and not result.dropped_pieces
            else None
        )
        waste_mm_value = (
            round(result.waste_mm, 0)
            if result is not None and not result.dropped_pieces
            else None
        )
        cuts_value = (
            result.cuts if result is not None and not result.dropped_pieces else None
        )

        scenario_rows.append(
            {
                "Profiel": profile.profile,
                "Materiaal": profile.material,
                "Productie": profile.production,
                "Keuze": selection.choice_label,
                "Staaflengte (mm)": stock_length,
                "Aantal staven": bars_value,
                "Afval %": waste_pct_value,
                "Afval (mm)": waste_mm_value,
                "Zaagsneden": cuts_value,
                "Opmerking": scenario_remark,
            }
        )

        for piece in profile.pieces:
            piece_rows.append(
                {
                    "Profiel": profile.profile,
                    "Materiaal": profile.material,
                    "Productie": profile.production,
                    "Onderdeel": piece.label,
                    "Lengte (mm)": piece.length_mm,
                }
            )

        if selection.choice == "input":
            order_remark_text = "Handmatig zagen per stuk."
        elif result is None or result.dropped_pieces:
            blockers_text = "; ".join(selection.blockers)
            if blockers_text:
                order_remark_text = f"Handmatig controleren – {blockers_text}"
            else:
                order_remark_text = "Handmatig controleren – scenario niet mogelijk."
        else:
            order_remark_text = (
                f"Afval {result.waste_pct:.1f}% ({result.waste_mm:.0f} mm)."
            )

        total_length_m = None
        if result is not None and not result.dropped_pieces and stock_length is not None:
            total_length_m = round(result.bars * stock_length / 1000, 3)

        weight_total = None
        stats_key = (profile.profile, profile.material, profile.production)
        stats_entry = stats_map.get(stats_key)
        if (
            stats_entry is not None
            and stats_entry.weight_per_mm is not None
            and stock_length is not None
            and bars_value
        ):
            weight_total = (
                float(stats_entry.weight_per_mm)
                * float(stock_length)
                * float(bars_value)
                / 1.0
            )
            total_weight_known = True
            total_weight_kg += weight_total

        order_rows.append(
            {
                "Profiel": profile.profile,
                "Materiaal": profile.material,
                "Productie": profile.production,
                "Keuze": selection.choice_label,
                "Staaflengte (mm)": stock_length,
                "Aantal staven": bars_value,
                "Totale lengte (m)": total_length_m,
                "Totaal gewicht (kg)": round(weight_total, 3)
                if weight_total is not None
                else None,
                "Opmerking": order_remark_text,
            }
        )

        if bars_value and bars_value > 0:
            has_valid_bars = True
            total_bars += int(bars_value)
            raw_items.append(
                {
                    "Profiel": profile.profile or "Brutemateriaal",
                    "Materiaal": profile.material,
                    "Lengte": stock_length or "",
                    "St.": int(bars_value),
                    "kg": _format_weight_kg(weight_total),
                }
            )

    summary_weight = total_weight_kg if total_weight_known else None
    return OpticutterOrderComputation(
        scenario_rows=scenario_rows,
        piece_rows=piece_rows,
        order_rows=order_rows,
        raw_items=raw_items,
        has_valid_bars=has_valid_bars,
        total_bars=total_bars,
        total_weight_kg=summary_weight,
        selection_count=selection_count,
    )


def compute_opticutter_order_details(
    bom_df: pd.DataFrame,
    context: OpticutterExportContext | None,
) -> Dict[str, OpticutterOrderComputation]:
    """Return computed Opticutter export data per production."""

    if context is None:
        return {}
    stats_map = _collect_opticutter_profile_stats(bom_df)
    details: Dict[str, OpticutterOrderComputation] = {}
    for prod_key, export in context.productions.items():
        details[prod_key] = _compute_opticutter_order_exports(export, stats_map)
    return details


def describe_finish_combo(
    finish_value: object,
    ral_value: object,
) -> Dict[str, str]:
    """Return normalized metadata for a finish/RAL combination."""

    finish_text = _to_str(finish_value).strip()
    ral_text = _to_str(ral_value).strip()
    finish_norm = _normalize_finish_folder(finish_text)
    if not finish_norm:
        finish_norm = "_Onbekend"
    finish_display = finish_text if finish_text else "Onbekend"
    ral_norm = _normalize_finish_folder(ral_text) if ral_text else ""
    ral_display = ral_text
    folder_name = f"Finish-{finish_norm}"
    if ral_norm:
        folder_name = f"{folder_name}-{ral_norm}"
    label = finish_display
    if ral_display:
        label = f"{finish_display} – {ral_display}"
    filename_component = _normalize_finish_folder(label)
    if not filename_component:
        suffix = f"-{ral_norm}" if ral_norm else ""
        filename_component = f"{finish_norm}{suffix}"
    return {
        "finish_display": finish_display,
        "finish_norm": finish_norm,
        "ral_display": ral_display,
        "ral_norm": ral_norm,
        "folder_name": folder_name,
        "label": label,
        "filename_component": filename_component,
        "key": folder_name,
    }


def _format_order_section_title(context_kind: object, context_label: object) -> str:
    kind_text = (_to_str(context_kind) or "productie").strip() or "productie"
    label_text = _to_str(context_label).strip()
    kind_title = kind_text[0].upper() + kind_text[1:] if kind_text else "Productie"
    if label_text:
        return f"{kind_title}: {label_text}"
    return kind_title


def _normalize_order_sections(
    production: object,
    items: List[Dict[str, object]],
    label_kind: object,
    total_surface_m2: float | None,
    total_weight_kg: float | None,
    column_layout: Optional[List[Dict[str, object]]],
    sections: Optional[List[OrderDocumentSection]],
) -> List[OrderDocumentSection]:
    if sections:
        normalized: List[OrderDocumentSection] = []
        for section in sections:
            if isinstance(section, OrderDocumentSection):
                normalized.append(
                    OrderDocumentSection(
                        context_label=_to_str(section.context_label),
                        context_kind=_to_str(section.context_kind) or "productie",
                        items=list(section.items or []),
                        selection_key=_to_str(section.selection_key),
                        total_surface_m2=section.total_surface_m2,
                        total_weight_kg=section.total_weight_kg,
                        column_layout=(
                            [dict(col) for col in section.column_layout]
                            if section.column_layout
                            else None
                        ),
                    )
                )
                continue
            normalized.append(
                OrderDocumentSection(
                    context_label=_to_str(section.get("context_label")),
                    context_kind=_to_str(section.get("context_kind")) or "productie",
                    items=list(section.get("items") or []),
                    selection_key=_to_str(section.get("selection_key")),
                    total_surface_m2=section.get("total_surface_m2"),
                    total_weight_kg=section.get("total_weight_kg"),
                    column_layout=(
                        [dict(col) for col in section.get("column_layout")]
                        if section.get("column_layout")
                        else None
                    ),
                )
            )
        if normalized:
            return normalized

    return [
        OrderDocumentSection(
            context_label=_to_str(production),
            context_kind=_to_str(label_kind) or "productie",
            items=list(items or []),
            selection_key="",
            total_surface_m2=total_surface_m2,
            total_weight_kg=total_weight_kg,
            column_layout=[dict(col) for col in column_layout] if column_layout else None,
        )
    ]


def _order_group_summary_text(sections: Sequence[OrderDocumentSection]) -> str:
    labels = [_to_str(section.context_label).strip() for section in sections]
    labels = [label for label in labels if label]
    if len(labels) <= 1:
        return ""
    return ", ".join(labels)


def _build_order_excel_section_data(
    section: OrderDocumentSection,
) -> tuple[pd.DataFrame, set[str], set[str]]:
    context_kind_clean = (
        (_to_str(section.context_kind) or "productie").strip() or "productie"
    )
    is_raw_material_order = context_kind_clean.lower().startswith("brutemateriaal")
    column_layout = (
        [dict(col) for col in section.column_layout] if section.column_layout else []
    )
    custom_layout = bool(column_layout)
    display_items, price_subtotal, price_summary_rows = _split_order_price_summary_rows(
        section.items,
        column_layout,
        is_raw=is_raw_material_order,
    )

    if custom_layout:
        headers: List[str] = []
        for column in column_layout:
            header = _order_column_export_label(column)
            if not header:
                header = column.get("key", "")
            column["label"] = header
            headers.append(header)
        rows: List[Dict[str, object]] = []
        for item in display_items:
            row: Dict[str, object] = {}
            for column, header in zip(column_layout, headers):
                key = column.get("key")
                value = item.get(key, "") if key else ""
                if column.get("integer"):
                    value = _coerce_integer_like(value)
                row[header] = value
            rows.append(row)
        df = pd.DataFrame(rows, columns=headers)
        surface_header: str | None = None
        weight_header: str | None = None
        total_price_header: str | None = None
        for column in column_layout:
            if surface_header is None and is_order_surface_column(column):
                surface_header = column["label"]
            if weight_header is None and is_order_weight_column(column):
                weight_header = column["label"]
            if total_price_header is None and _to_str(column.get("key")).strip().lower() == _PRICE_TOTAL_KEY.lower():
                total_price_header = column["label"]
        if (
            (surface_header and section.total_surface_m2 is not None)
            or (weight_header and section.total_weight_kg is not None)
            or (total_price_header and _to_str(price_subtotal).strip())
        ):
            total_row = {header: "" for header in headers}
            if headers:
                total_row[headers[0]] = "Totaal"
            if surface_header and section.total_surface_m2 is not None:
                total_row[surface_header] = _format_weight_kg(section.total_surface_m2)
            if weight_header and section.total_weight_kg is not None:
                total_row[weight_header] = _format_weight_kg(section.total_weight_kg)
            if total_price_header and _to_str(price_subtotal).strip():
                total_row[total_price_header] = price_subtotal
            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        df = _append_price_summary_rows_to_df(df, headers, price_summary_rows)
        left_cols = {
            _to_str(column.get("label") or column.get("key") or "").strip()
            or column.get("key", "")
            for column in column_layout
            if _to_str(column.get("justify") or "left").strip().lower() != "right"
        }
        wrap_cols = {
            _to_str(column.get("label") or column.get("key") or "").strip()
            or column.get("key", "")
            for column in column_layout
            if bool(column.get("wrap"))
        }
        return df, left_cols, wrap_cols

    if is_raw_material_order:
        df_columns = ["Profiel", "Materiaal", "Lengte", "St.", "kg"]
        df = pd.DataFrame(display_items, columns=df_columns)
        if section.total_weight_kg is not None:
            total_row = {
                "Profiel": "Totaal",
                "Materiaal": "",
                "Lengte": "",
                "St.": "",
                "kg": _format_weight_kg(section.total_weight_kg),
            }
            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        return df, {"Profiel", "Materiaal"}, {"Profiel", "Materiaal"}

    visible_columns = _standard_order_visible_columns(
        display_items,
        total_surface_m2=section.total_surface_m2,
        total_weight_kg=section.total_weight_kg,
    )
    df = _build_standard_order_excel_dataframe(
        display_items,
        total_surface_m2=section.total_surface_m2,
        total_weight_kg=section.total_weight_kg,
    )
    return (
        df,
        _standard_order_left_columns(visible_columns),
        _standard_order_wrap_columns(visible_columns),
    )


def _order_column_export_label(column: Mapping[str, object]) -> str:
    """Return a compact header label for exported order tables."""

    key = _to_str(column.get("key")).strip().lower()
    label = _to_str(column.get("label") or column.get("key") or "").strip()
    label_lower = label.lower()
    label_compact = (
        label_lower.replace("€", "")
        .replace("(euro)", "")
        .replace("\u20ac", "")
        .replace("(", "")
        .replace(")", "")
        .replace(" ", "")
    )

    if key == "oppervlakte" or label_lower in {"oppervlakte", "oppervlakte/st"}:
        return "m\u00b2/st"
    if key == "gewicht" or label_lower in {"gewicht", "gewicht (kg)", "gewicht/st"}:
        return "kg/st"
    if key == _PRICE_UNIT_KEY.lower() or label_compact in {
        "eenheidsprijs",
        "unitprice",
    }:
        return "Prijs/st. (\u20ac)"
    if key == _PRICE_TOTAL_KEY.lower() or label_compact in {
        "totaalprijs",
        "totalprice",
    }:
        return "Totaal (\u20ac)"
    return label or _to_str(column.get("key")).strip()


def _pdf_order_column_label(column: Mapping[str, object]) -> str:
    """Return a compact header label for PDF tables."""

    return _order_column_export_label(column)


def _order_title_style(styles, palette: Mapping[str, str], title: object, usable_w: float):
    title_text = _to_str(title)
    font_name = "Helvetica-Bold"
    font_size = 20.0
    try:
        title_width = stringWidth(title_text, font_name, font_size)
    except Exception:
        title_width = len(title_text) * font_size * 0.55

    if title_width > usable_w:
        font_size = 18.0
    if title_width > usable_w * 1.55:
        font_size = 16.0
    if title_width > usable_w * 2.2:
        font_size = 14.5

    return ParagraphStyle(
        "OrderTitle",
        parent=styles["Heading1"],
        textColor=colors.HexColor(palette["accent"]),
        fontName=font_name,
        fontSize=font_size,
        leading=font_size + 2.0,
        spaceAfter=1,
    )


def _order_table_header_cell(
    label: object,
    *,
    align: str = "LEFT",
    text_color: str,
    font_size: float,
):
    alignment = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align.upper(), 0)
    style = ParagraphStyle(
        "OrderTableHeader",
        fontName="Helvetica-Bold",
        fontSize=font_size,
        leading=font_size + 1.5,
        alignment=alignment,
        textColor=colors.HexColor(text_color),
        splitLongWords=0,
    )
    return Paragraph(escape(_to_str(label)), style)


def _order_table_header_min_width(label: object, font_size: float) -> float:
    """Return the minimum width needed to keep header words intact."""

    text = re.sub(r"\s+", " ", _to_str(label).replace("\n", " ")).strip()
    tokens = [token for token in text.split(" ") if token]
    if not tokens:
        return 0.0
    try:
        token_width = max(
            stringWidth(token, "Helvetica-Bold", font_size) for token in tokens
        )
    except Exception:
        token_width = max(len(token) for token in tokens) * font_size * 0.55
    return token_width + 14.0


def _order_table_value_min_width(
    value: object,
    *,
    font_size: float,
    wrap: bool,
) -> float:
    text = _clean_order_cell_text(value).replace("\u00a0", " ")
    if not text:
        return 0.0
    tokens = [token for token in re.split(r"\s+", text) if token] if wrap else [text]
    if not tokens:
        return 0.0
    try:
        width = max(stringWidth(token, "Helvetica", font_size) for token in tokens)
    except Exception:
        width = max(len(token) for token in tokens) * font_size * 0.55
    return width + 14.0


def _order_table_column_min_width(
    label: object,
    values: Iterable[object],
    *,
    header_font_size: float,
    body_font_size: float,
    wrap: bool,
    max_width: float | None = None,
) -> float:
    minimum = _order_table_header_min_width(label, header_font_size)
    for value in values:
        minimum = max(
            minimum,
            _order_table_value_min_width(
                value,
                font_size=body_font_size,
                wrap=wrap,
            ),
        )
    if max_width is not None and max_width > 0:
        minimum = min(minimum, max_width)
    return minimum


def _weighted_widths_with_minimums(
    total_width: float,
    weights: Sequence[float],
    minimums: Sequence[float],
) -> List[float]:
    """Distribute width by weight while honoring column minimums when possible."""

    count = len(weights)
    if count == 0:
        return []
    safe_weights = [weight if weight > 0 else 1.0 for weight in weights]
    safe_minimums = [max(0.0, minimum) for minimum in minimums]
    while len(safe_minimums) < count:
        safe_minimums.append(0.0)
    safe_minimums = safe_minimums[:count]
    minimum_total = sum(safe_minimums)
    if minimum_total >= total_width and minimum_total > 0:
        return [total_width * (minimum / minimum_total) for minimum in safe_minimums]

    fixed: set[int] = set()
    widths = [0.0] * count
    while True:
        fixed_width = sum(safe_minimums[index] for index in fixed)
        flexible = [index for index in range(count) if index not in fixed]
        if not flexible:
            break
        flexible_width = max(0.0, total_width - fixed_width)
        flexible_weight = sum(safe_weights[index] for index in flexible) or len(flexible)
        for index in fixed:
            widths[index] = safe_minimums[index]
        for index in flexible:
            widths[index] = flexible_width * (safe_weights[index] / flexible_weight)
        new_fixed = [
            index
            for index in flexible
            if widths[index] < safe_minimums[index]
        ]
        if not new_fixed:
            break
        fixed.update(new_fixed)
    return widths


def _order_metadata_table(
    rows: Sequence[Tuple[str, object]],
    *,
    table_width: float,
    base_style,
    value_max_width: float | None = None,
    right_gap: float | None = None,
):
    """Return a compact label/value table for order PDF metadata."""

    visible_rows = [
        (_to_str(label).strip(), _to_str(value).strip())
        for label, value in rows
        if _to_str(label).strip() or _to_str(value).strip()
    ]
    if not visible_rows:
        return Paragraph("", base_style)

    font_size = float(getattr(base_style, "fontSize", 10) or 10)
    leading = float(getattr(base_style, "leading", font_size + 2.2) or font_size + 2.2)
    text_color = getattr(base_style, "textColor", colors.HexColor(ORDER_TEXT_COLOR))
    try:
        widest_label = max(
            stringWidth(f"{label}: ", "Helvetica-Bold", font_size)
            for label, _value in visible_rows
            if label
        )
    except Exception:
        widest_label = max((len(label) for label, _value in visible_rows), default=10) * font_size * 0.55

    gap = right_gap if right_gap is not None else 11 * mm
    value_cap = value_max_width if value_max_width is not None else 60 * mm
    available_width = max(80.0, table_width - gap)
    metadata_width = min(available_width, widest_label + value_cap)
    metadata_width = max(min(available_width, widest_label + 56.0), metadata_width)

    def value_html(value: str) -> str:
        lines = value.splitlines() or [value]
        return "<br/>".join(escape(line.strip()) for line in lines)

    data = []
    for index, (label, value) in enumerate(visible_rows):
        try:
            indent = stringWidth(f"{label}: ", "Helvetica-Bold", font_size) + 2.0
        except Exception:
            indent = (len(label) + 2) * font_size * 0.55
        row_style = ParagraphStyle(
            f"OrderMetaRow{index}",
            parent=base_style,
            fontSize=font_size,
            leading=leading,
            textColor=text_color,
            leftIndent=indent,
            firstLineIndent=-indent,
            wordWrap="CJK",
        )
        data.append(
            [
                Paragraph(
                    f"<b>{escape(label)}:</b> {value_html(value)}" if label else value_html(value),
                    row_style,
                )
            ]
        )
    table = Table(data, colWidths=[metadata_width], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0.8),
            ]
        )
    )
    return table


def _order_company_lines(company_info: Mapping[str, object] | None) -> List[str]:
    """Return formatted company lines for order headers."""

    if not company_info:
        return []
    lines = [
        f"<b>{company_info.get('name','')}</b>",
        f"{company_info.get('address','')}",
        f"BTW: {company_info.get('vat','')}",
        f"E-mail: {company_info.get('email','')}",
    ]
    website = _to_str(company_info.get("website")).strip()
    if website:
        lines.append(f"Website: {website}")
    return lines


def _append_order_price_summary_story(
    story: List[object],
    summary_rows: Sequence[tuple[str, object]],
    *,
    usable_w: float,
    palette: Mapping[str, str],
) -> None:
    if not summary_rows:
        return

    label_style = ParagraphStyle(
        "OrderPriceSummaryLabel",
        fontName="Helvetica",
        fontSize=8.7,
        leading=10.8,
        alignment=0,
    )
    value_style = ParagraphStyle(
        "OrderPriceSummaryValue",
        fontName="Helvetica",
        fontSize=8.7,
        leading=10.8,
        alignment=2,
    )
    data = [
        [
            Paragraph(escape(_to_str(label)), label_style),
            Paragraph(escape(_num_to_2dec(total)), value_style),
        ]
        for label, total in summary_rows
    ]
    table_width = min(usable_w * 0.42, 82 * mm)
    value_width = min(32 * mm, table_width * 0.42)
    label_width = table_width - value_width
    summary_tbl = Table(
        data,
        colWidths=[label_width, value_width],
        hAlign="RIGHT",
    )
    style_cmds = [
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(ORDER_TEXT_COLOR)),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
    ]
    if data:
        last_row = len(data) - 1
        style_cmds.extend(
            [
                ("FONTNAME", (0, last_row), (-1, last_row), "Helvetica-Bold"),
                (
                    "LINEABOVE",
                    (0, last_row),
                    (-1, last_row),
                    0.45,
                    colors.HexColor(ORDER_TABLE_OUTLINE_COLOR),
                ),
                (
                    "BACKGROUND",
                    (0, last_row),
                    (-1, last_row),
                    colors.HexColor(palette["total_fill"]),
                ),
            ]
        )
    summary_tbl.setStyle(TableStyle(style_cmds))
    story.append(Spacer(0, 8))
    story.append(summary_tbl)


def _build_order_pdf_section_story(
    section: OrderDocumentSection,
    *,
    story: List[object],
    usable_w: float,
    palette: Mapping[str, str],
    section_title_style: ParagraphStyle,
    show_title: bool,
    start_item_number: int = 1,
) -> None:
    context_kind_clean = (
        (_to_str(section.context_kind) or "productie").strip() or "productie"
    )
    is_raw_material_order = context_kind_clean.lower().startswith("brutemateriaal")
    column_layout = (
        [dict(col) for col in section.column_layout] if section.column_layout else []
    )
    custom_layout = bool(column_layout)
    display_items, price_subtotal, price_summary_rows = _split_order_price_summary_rows(
        section.items,
        column_layout,
        is_raw=is_raw_material_order,
    )
    table_header_font_size = 9.5
    table_body_font_size = 8.7
    table_small_font_size = 8.2
    item_column_title = "Nr."
    item_col_width = min(11 * mm, usable_w * 0.07)
    item_col_width = max(item_col_width, min(9 * mm, usable_w * 0.045))
    content_usable_w = max(usable_w - item_col_width, usable_w * 0.7)
    item_col_width = usable_w - content_usable_w
    try:
        next_item_number = max(1, int(start_item_number))
    except Exception:
        next_item_number = 1
    standard_columns: List[_StandardOrderColumn] = []
    if not custom_layout and not is_raw_material_order:
        standard_columns = _standard_order_visible_columns(
            display_items,
            total_surface_m2=section.total_surface_m2,
            total_weight_kg=section.total_weight_kg,
        )

    if show_title:
        story.append(
            Paragraph(
                _format_order_section_title(
                    context_kind_clean,
                    section.context_label,
                ),
                section_title_style,
            )
        )
        story.append(Spacer(0, 6))

    if custom_layout:
        head = [
            _order_table_header_cell(
                item_column_title,
                align="CENTER",
                text_color=palette["accent_text"],
                font_size=table_header_font_size,
            )
        ]
        for column in column_layout:
            header = _pdf_order_column_label(column)
            if not header:
                header = column.get("key", "")
            column["label"] = header
            align = _to_str(column.get("justify") or "left").strip().upper() or "LEFT"
            if align not in {"LEFT", "RIGHT", "CENTER"}:
                align = "LEFT"
            head.append(
                _order_table_header_cell(
                    header,
                    align=align,
                    text_color=palette["accent_text"],
                    font_size=table_header_font_size,
                )
            )
    elif is_raw_material_order:
        head = [item_column_title, "Profiel", "Materiaal", "Lengte", "St.", "kg"]
    else:
        head = [item_column_title] + [column.pdf_label for column in standard_columns]

    def wrap_cell_html(val: str, small=False, align=None):
        style = ParagraphStyle(
            "cellsmall" if small else "cell",
            fontName="Helvetica",
            fontSize=table_small_font_size if small else table_body_font_size,
            leading=10.0 if small else 10.7,
            wordWrap="CJK",
        )
        if align:
            style.alignment = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(
                align.upper(), 0
            )
        return Paragraph(str(val if (val is not None) else ""), style)

    standard_col_widths: List[float] | None = None
    if not custom_layout and not is_raw_material_order:
        standard_col_widths = _standard_order_column_widths(
            standard_columns,
            display_items,
            total_width=content_usable_w,
            small_font_size=table_small_font_size,
            header_font_size=table_header_font_size,
        )

    def description_cell_html(val: object, width: float) -> str:
        lines = _wrap_words_to_lines(
            _clean_order_cell_text(val),
            max(24.0, width),
            "Helvetica",
            table_body_font_size,
            max_lines=2,
        )
        return "<br/>".join(escape(line) for line in lines)

    data = [head]
    total_row_index: int | None = None
    if custom_layout:
        surface_idx: int | None = None
        weight_idx: int | None = None
        total_price_idx: int | None = None
        for idx, column in enumerate(column_layout):
            if surface_idx is None and is_order_surface_column(column):
                surface_idx = idx
            if weight_idx is None and is_order_weight_column(column):
                weight_idx = idx
            if total_price_idx is None and _to_str(column.get("key")).strip().lower() == _PRICE_TOTAL_KEY.lower():
                total_price_idx = idx
        for row_offset, item in enumerate(display_items):
            row_cells: List[Paragraph] = [
                wrap_cell_html(
                    str(next_item_number + row_offset),
                    small=True,
                    align="CENTER",
                )
            ]
            for idx, column in enumerate(column_layout):
                key = column.get("key")
                value = item.get(key, "") if key else ""
                if column.get("numeric"):
                    if column.get("integer"):
                        value = _format_integer_like(value)
                    else:
                        value = _num_to_2dec(value)
                    small = True
                else:
                    value = _to_str(value)
                    small = False
                align = (
                    _to_str(column.get("justify") or "left").strip().upper()
                    or "LEFT"
                )
                if align not in {"LEFT", "RIGHT", "CENTER"}:
                    align = "LEFT"
                row_cells.append(wrap_cell_html(value, small=small, align=align))
            data.append(row_cells)

        if (
            (surface_idx is not None and section.total_surface_m2 is not None)
            or (weight_idx is not None and section.total_weight_kg is not None)
            or (total_price_idx is not None and _to_str(price_subtotal).strip())
        ):
            total_row: List[Paragraph] = [
                wrap_cell_html("", small=True, align="CENTER")
            ]
            for idx, column in enumerate(column_layout):
                align = (
                    _to_str(column.get("justify") or "left").strip().upper()
                    or "LEFT"
                )
                if align not in {"LEFT", "RIGHT", "CENTER"}:
                    align = "LEFT"
                if idx == surface_idx and section.total_surface_m2 is not None:
                    total_row.append(
                        wrap_cell_html(
                            _num_to_2dec(section.total_surface_m2),
                            small=True,
                            align=align,
                        )
                    )
                elif idx == weight_idx and section.total_weight_kg is not None:
                    total_row.append(
                        wrap_cell_html(
                            _num_to_2dec(section.total_weight_kg),
                            small=True,
                            align=align,
                        )
                    )
                elif idx == total_price_idx and _to_str(price_subtotal).strip():
                    total_row.append(
                        wrap_cell_html(
                            _num_to_2dec(price_subtotal),
                            small=True,
                            align=align,
                        )
                    )
                elif idx == 0:
                    total_row.append(
                        wrap_cell_html("Totaal", small=False, align="LEFT")
                    )
                else:
                    total_row.append(
                        wrap_cell_html(
                            "",
                            small=bool(column.get("numeric")),
                            align=align,
                        )
                    )
            data.append(total_row)
            total_row_index = len(data) - 1
    elif is_raw_material_order:
        for row_offset, item in enumerate(display_items):
            prof = _to_str(item.get("Profiel", ""))
            mat = _to_str(item.get("Materiaal", ""))
            length_val = item.get("Lengte", "")
            length = _to_str("" if length_val in (None, "") else length_val)
            qty_val = item.get("St.", "")
            qty = _to_str("" if qty_val in (None, "") else qty_val)
            weight_val = item.get("kg", "")
            weight = _num_to_2dec(weight_val)
            data.append(
                [
                    wrap_cell_html(
                        str(next_item_number + row_offset),
                        small=True,
                        align="CENTER",
                    ),
                    wrap_cell_html(prof, small=False, align="LEFT"),
                    wrap_cell_html(mat, small=False, align="LEFT"),
                    wrap_cell_html(length, small=True, align="RIGHT"),
                    wrap_cell_html(qty, small=True, align="RIGHT"),
                    wrap_cell_html(weight, small=True, align="RIGHT"),
                ]
            )
        if section.total_weight_kg is not None:
            total_row = [
                wrap_cell_html("", small=True, align="CENTER"),
                wrap_cell_html("Totaal", small=False, align="LEFT"),
                wrap_cell_html("", small=False, align="LEFT"),
                wrap_cell_html("", small=True, align="RIGHT"),
                wrap_cell_html("", small=True, align="RIGHT"),
                wrap_cell_html(
                    _num_to_2dec(section.total_weight_kg),
                    small=True,
                    align="RIGHT",
                ),
            ]
            data.append(total_row)
            total_row_index = len(data) - 1
    else:
        standard_widths_by_key = {
            column.key: width
            for column, width in zip(standard_columns, standard_col_widths or [])
        }
        for row_offset, item in enumerate(display_items):
            row_cells: List[Paragraph] = [
                wrap_cell_html(
                    str(next_item_number + row_offset),
                    small=True,
                    align="CENTER",
                )
            ]
            for column in standard_columns:
                if column.key == "PartNumber":
                    value = escape(_clean_order_cell_text(item.get("PartNumber", "")))
                elif column.key == "Description":
                    desc_width = max(
                        24.0,
                        standard_widths_by_key.get(
                            "Description",
                            content_usable_w * 0.40,
                        )
                        - 10,
                    )
                    value = description_cell_html(item.get("Description", ""), desc_width)
                elif column.key == "Materiaal":
                    value = _material_nowrap(
                        _clean_order_cell_text(item.get("Materiaal", ""))
                    )
                elif column.key == "Aantal":
                    value = item.get("Aantal", "")
                elif column.key == "Oppervlakte":
                    value = _num_to_2dec(item.get("Oppervlakte", ""))
                elif column.key == "Gewicht":
                    value = _num_to_2dec(item.get("Gewicht", ""))
                else:
                    value = item.get(column.key, "")
                row_cells.append(
                    wrap_cell_html(
                        value,
                        small=column.small,
                        align=column.align,
                    )
                )
            data.append(row_cells)
        if section.total_surface_m2 is not None or section.total_weight_kg is not None:
            total_row = [
                wrap_cell_html("", small=True, align="CENTER")
            ]
            for idx, column in enumerate(standard_columns):
                if idx == 0:
                    value = "Totaal"
                    small = False
                    align = "LEFT"
                elif column.key == "Oppervlakte" and section.total_surface_m2 is not None:
                    value = _num_to_2dec(section.total_surface_m2)
                    small = True
                    align = column.align
                elif column.key == "Gewicht" and section.total_weight_kg is not None:
                    value = _num_to_2dec(section.total_weight_kg)
                    small = True
                    align = column.align
                else:
                    value = ""
                    small = column.small
                    align = column.align
                total_row.append(
                    wrap_cell_html(
                        value,
                        small=small,
                        align=align,
                    )
                )
            data.append(total_row)
            total_row_index = len(data) - 1

    if custom_layout and column_layout:
        weights: List[float] = []
        for column in column_layout:
            try:
                weight_val = float(column.get("weight", 0))
            except Exception:
                weight_val = 0.0
            weights.append(weight_val if weight_val > 0 else 1.0)
        minimums = [
            _order_table_column_min_width(
                column.get("label"),
                [
                    item.get(column.get("key"), "") if column.get("key") else ""
                    for item in display_items
                ],
                header_font_size=table_header_font_size,
                body_font_size=table_small_font_size
                if column.get("numeric")
                else table_body_font_size,
                wrap=bool(column.get("wrap")),
                max_width=content_usable_w * 0.34,
            )
            for column in column_layout
        ]
        col_widths = [item_col_width] + [
            width for width in _weighted_widths_with_minimums(
                content_usable_w,
                weights,
                minimums,
            )
        ]
    elif is_raw_material_order:
        col_fracs = [0.32, 0.24, 0.16, 0.12, 0.16]
        col_widths = [item_col_width] + [
            content_usable_w * frac for frac in col_fracs
        ]
    else:
        col_widths = [item_col_width] + (
            standard_col_widths
            or [content_usable_w / max(1, len(standard_columns))]
            * max(1, len(standard_columns))
        )

    tbl = LongTable(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["accent"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(palette["accent_text"])),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor(ORDER_TEXT_COLOR)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), table_header_font_size),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(ORDER_TABLE_OUTLINE_COLOR)),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor(ORDER_TABLE_GRID_COLOR)),
        (
            "ROWBACKGROUNDS",
            (0, 1),
            (-1, -1),
            [colors.white, colors.HexColor(ORDER_TABLE_ALT_ROW_COLOR)],
        ),
        ("LEFTPADDING", (0, 0), (-1, 0), 6),
        ("RIGHTPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("LEFTPADDING", (0, 1), (-1, -1), 5),
        ("RIGHTPADDING", (0, 1), (-1, -1), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]
    if custom_layout and column_layout:
        style_cmds.append(("ALIGN", (0, 0), (0, -1), "CENTER"))
        for idx, column in enumerate(column_layout, start=1):
            align = _to_str(column.get("justify") or "left").strip().upper() or "LEFT"
            if align not in {"LEFT", "RIGHT", "CENTER"}:
                align = "LEFT"
            style_cmds.append(("ALIGN", (idx, 0), (idx, -1), align))
    elif is_raw_material_order:
        style_cmds.extend(
            [
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (5, -1), "RIGHT"),
            ]
        )
    else:
        style_cmds.append(("ALIGN", (0, 0), (0, -1), "CENTER"))
        for idx, column in enumerate(standard_columns, start=1):
            style_cmds.append(("ALIGN", (idx, 0), (idx, -1), column.align))
    if total_row_index is not None:
        style_cmds.extend(
            [
                ("FONTNAME", (0, total_row_index), (-1, total_row_index), "Helvetica-Bold"),
                (
                    "BACKGROUND",
                    (0, total_row_index),
                    (-1, total_row_index),
                    colors.HexColor(palette["total_fill"]),
                ),
                (
                    "LINEABOVE",
                    (0, total_row_index),
                    (-1, total_row_index),
                    0.45,
                    colors.HexColor(ORDER_TABLE_OUTLINE_COLOR),
                ),
            ]
        )
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    _append_order_price_summary_story(
        story,
        price_summary_rows,
        usable_w=usable_w,
        palette=palette,
    )


def generate_pdf_order_platypus(
    path: str,
    company_info: Dict[str, object],
    supplier: Supplier | None,
    production: str,
    items: List[Dict[str, object]],
    doc_type: str = "Bestelbon",
    doc_number: str | None = None,
    footer_note: Optional[str] = None,
    quote_footer_note: Optional[str] = None,
    delivery: DeliveryAddress | None = None,
    project_number: str | None = None,
    project_name: str | None = None,
    label_kind: str = "productie",
    order_remark: str | None = None,
    total_surface_m2: float | None = None,
    total_weight_kg: float | None = None,
    en1090_required: bool = False,
    en1090_note: Optional[str] = None,
    include_bruto_note: bool = False,
    column_layout: Optional[List[Dict[str, object]]] = None,
    sections: Optional[List[OrderDocumentSection]] = None,
) -> None:
    """Generate a PDF order using ReportLab if available.

    ``doc_type`` determines the document title, e.g. ``"Bestelbon"`` or
    ``"Offerteaanvraag"``.
    """
    if not REPORTLAB_OK:
        return

    normalized_sections = _normalize_order_sections(
        production,
        items,
        label_kind,
        total_surface_m2,
        total_weight_kg,
        column_layout,
        sections,
    )
    multiple_sections = len(normalized_sections) > 1

    margin = 18 * mm
    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )
    width, _ = A4
    usable_w = width - 2 * margin
    left_col_width = usable_w * 0.58
    right_col_width = usable_w - left_col_width
    palette = _order_palette(company_info)
    styles = getSampleStyleSheet()
    text_style = styles["Normal"]
    text_style.fontSize = 10
    text_style.leading = 12.2
    text_style.textColor = colors.HexColor(ORDER_TEXT_COLOR)
    meta_style = ParagraphStyle("meta", parent=text_style, leading=12.4)
    delivery_style = ParagraphStyle(
        "delivery",
        parent=text_style,
        fontSize=9.2,
        leading=11.2,
        textColor=colors.HexColor(ORDER_TEXT_COLOR),
    )
    small_style = ParagraphStyle(
        "small",
        parent=text_style,
        fontSize=8.4,
        leading=10.3,
        textColor=colors.HexColor(ORDER_MUTED_TEXT_COLOR),
    )
    section_title_style = ParagraphStyle(
        "sectiontitle",
        parent=text_style,
        fontName="Helvetica-Bold",
        fontSize=11.2,
        leading=13.2,
        textColor=colors.HexColor(ORDER_TEXT_COLOR),
    )

    doc_type_text = (_to_str(doc_type).strip() or "Bestelbon")
    doc_type_text_lower = doc_type_text.lower()
    doc_type_text_slug = re.sub(r"[^0-9a-z]+", "", doc_type_text_lower)
    is_standaard_doc = doc_type_text_lower.startswith("standaard")
    primary_section = normalized_sections[0]
    if not multiple_sections:
        if total_surface_m2 is None:
            total_surface_m2 = primary_section.total_surface_m2
        if total_weight_kg is None:
            total_weight_kg = primary_section.total_weight_kg
        if column_layout is None and primary_section.column_layout:
            column_layout = [dict(col) for col in primary_section.column_layout]
        if not items and primary_section.items:
            items = list(primary_section.items)
    production_text = _to_str(production).strip()
    label_kind_clean = (_to_str(label_kind) or "productie").strip() or "productie"
    # Per-section layout flags
    column_layout = [dict(col) for col in column_layout] if column_layout else []
    custom_layout = bool(column_layout)
    is_raw_material_order = label_kind_clean.lower().startswith("brutemateriaal")
    if not multiple_sections:
        if not production_text:
            production_text = _to_str(primary_section.context_label).strip()
        label_kind_clean = (
            (_to_str(primary_section.context_kind) or label_kind_clean).strip()
            or label_kind_clean
        )
    order_remark_text = _to_str(order_remark) if order_remark is not None else ""
    order_remark_has_content = bool(order_remark_text.strip())
    place_remark_in_delivery_block = _should_place_remark_in_delivery_block(
        order_remark_has_content=order_remark_has_content,
        doc_type_text_slug=doc_type_text_slug,
        is_standaard_doc=is_standaard_doc,
        delivery=delivery,
    )

    doc_meta_rows: List[Tuple[str, object]] = []
    if doc_number:
        doc_meta_rows.append(("Nummer", doc_number))
    today = datetime.date.today().strftime("%Y-%m-%d")
    doc_meta_rows.append(("Datum", today))
    label_title = label_kind_clean[0].upper() + label_kind_clean[1:]
    group_summary = _order_group_summary_text(normalized_sections)
    if multiple_sections:
        if group_summary:
            doc_meta_rows.append(("Gecombineerde bon voor", group_summary))
    elif production_text:
        doc_meta_rows.append((label_title, production_text))
    if project_number:
        doc_meta_rows.append(("Projectnummer", project_number))
    if project_name:
        doc_meta_rows.append(("Projectnaam", project_name))
    if order_remark_has_content and not place_remark_in_delivery_block:
        doc_meta_rows.append(("Opmerking", order_remark_text))

    company_lines = _order_company_lines(company_info)

    logo_flowable = None
    logo_path_info = company_info.get("logo_path") if company_info else None
    if logo_path_info:
        logo_path = resolve_runtime_path(str(logo_path_info))
        if logo_path and logo_path.exists():
            try:
                from PIL import Image as PILImage  # type: ignore
            except Exception:  # pragma: no cover - Pillow missing during runtime
                PILImage = None  # type: ignore
            if PILImage is not None:
                try:
                    with PILImage.open(logo_path) as src_logo:  # type: ignore[union-attr]
                        logo_img = src_logo.convert("RGBA")
                        crop_box = _normalize_crop_box(
                            company_info.get("logo_crop"),
                            logo_img.width,
                            logo_img.height,
                        )
                        if crop_box:
                            logo_img = logo_img.crop(crop_box)
                        if logo_img.width > 0 and logo_img.height > 0:
                            buffer = io.BytesIO()
                            logo_img.save(buffer, format="PNG")
                            buffer.seek(0)
                            aspect = (
                                logo_img.width / logo_img.height
                                if logo_img.height
                                else 1.0
                            )
                            max_width = 38 * mm
                            max_height = 18 * mm
                            width_pt = max_width
                            height_pt = width_pt / aspect if aspect else max_height
                            if height_pt > max_height:
                                height_pt = max_height
                                width_pt = height_pt * aspect
                            logo_flowable = RLImage(
                                buffer, width=width_pt, height=height_pt
                            )
                            logo_flowable.hAlign = "LEFT"
                except Exception:
                    logo_flowable = None

    supp_lines: List[str] = []
    if supplier is not None and not is_standaard_doc:
        full_addr = format_supplier_address(supplier)
        supplier_label = (
            "Offerte aangevraagd bij:"
            if doc_type_text_slug.startswith("offerte")
            else "Besteld bij:"
        )

        supp_lines = [f"<b>{supplier_label}</b> {supplier.supplier}"]
        if full_addr:
            supp_lines.append(full_addr)
        supp_lines.append(f"BTW: {supplier.btw or ''}")
        if supplier.contact_sales:
            supp_lines.append(f"Contact sales: {supplier.contact_sales}")
        if supplier.sales_email:
            supp_lines.append(f"E-mail: {supplier.sales_email}")
        if supplier.phone:
            supp_lines.append(f"Tel: {supplier.phone}")

    client_block = Paragraph("<br/>".join(company_lines), text_style)
    doc_block = _order_metadata_table(
        doc_meta_rows,
        table_width=left_col_width,
        base_style=meta_style,
        value_max_width=58 * mm,
    )

    supplier_block_parts: List[object] = []
    if supp_lines:
        supplier_block_parts.append(Paragraph("<br/>".join(supp_lines), text_style))

    delivery_block: object | None = None
    include_delivery_block = not is_standaard_doc and (
        delivery is not None or place_remark_in_delivery_block
    )
    if include_delivery_block:
        delivery_text_parts: List[str] = []
        if delivery:
            if _to_str(delivery.name).strip():
                delivery_text_parts.append(_to_str(delivery.name).strip())
            address_text = ", ".join(
                line.strip()
                for line in _to_str(delivery.address).splitlines()
                if line.strip()
            )
            if address_text:
                delivery_text_parts.append(address_text)
            if _to_str(delivery.remarks).strip():
                delivery_text_parts.append(_to_str(delivery.remarks).strip())
        delivery_rows: List[Tuple[str, object]] = []
        if delivery_text_parts:
            delivery_rows.append(("Leveradres", " | ".join(delivery_text_parts)))
        if place_remark_in_delivery_block:
            remark_lines = order_remark_text.splitlines()
            if not remark_lines:
                remark_lines = [order_remark_text]
            delivery_rows.append(
                ("Opmerking", "\n".join(line for line in remark_lines if line.strip()))
            )
        if delivery_rows:
            delivery_block = _order_metadata_table(
                delivery_rows,
                table_width=min(usable_w, 106 * mm),
                base_style=delivery_style,
                value_max_width=70 * mm,
                right_gap=0,
            )

    left_block_parts: List[object] = [doc_block]
    if supplier_block_parts:
        left_block_parts.append(Spacer(0, 8))
        left_block_parts.extend(supplier_block_parts)
    left_block: object = left_block_parts

    story = []
    title = (
        doc_type_text
        if multiple_sections or not production_text
        else f"{doc_type_text} {label_kind_clean}: {production_text}"
    )
    title_style = _order_title_style(styles, palette, title, usable_w)
    story.append(Paragraph(title, title_style))
    title_rule = Table([[""]], colWidths=[usable_w], rowHeights=[2])
    title_rule.setStyle(
        TableStyle(
            [
                ("LINEBELOW", (0, 0), (-1, -1), 0.55, colors.HexColor(ORDER_RULE_COLOR)),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(title_rule)
    story.append(Spacer(0, 12))

    right_block_parts: List[object] = []
    if logo_flowable is not None:
        logo_flowable.hAlign = "LEFT"
        right_block_parts.append(logo_flowable)
        right_block_parts.append(Spacer(0, 6))
    right_block_parts.append(client_block)
    right_block: object = right_block_parts

    header_tbl = LongTable(
        [[left_block, right_block]],
        colWidths=[left_col_width, right_col_width],
    )
    header_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_tbl)
    if delivery_block:
        story.append(Spacer(0, 6))
        story.append(delivery_block)
        story.append(Spacer(0, 12))
    else:
        story.append(Spacer(0, 10))

    if multiple_sections:
        next_item_number = 1
        for index, section in enumerate(normalized_sections):
            _build_order_pdf_section_story(
                section,
                story=story,
                usable_w=usable_w,
                palette=palette,
                section_title_style=section_title_style,
                show_title=True,
                start_item_number=next_item_number,
            )
            section_display_items, _section_subtotal, _section_summary_rows = (
                _split_order_price_summary_rows(
                    section.items,
                    section.column_layout or [],
                    is_raw=_to_str(section.context_kind).strip().lower().startswith("brutemateriaal"),
                )
            )
            next_item_number += len(section_display_items)
            if index < len(normalized_sections) - 1:
                story.append(Spacer(0, 10))

        if en1090_required:
            note_text = EN1090_NOTE_TEXT if en1090_note is None else _to_str(en1090_note)
            if note_text:
                story.append(Spacer(0, 12))
                en1090_note_html = note_text.replace("\n", "<br/>")
                if note_text == EN1090_NOTE_TEXT:
                    en1090_note_html = f"<b>{en1090_note_html}</b>"
                story.append(Paragraph(en1090_note_html, small_style))

        note = footer_note_for_doc_type(doc_type_text, footer_note, quote_footer_note)
        if note:
            story.append(Spacer(0, 8))
            story.append(Paragraph(note, small_style))

        doc.build(story)
        return

    display_items, price_subtotal, price_summary_rows = _split_order_price_summary_rows(
        items,
        column_layout,
        is_raw=is_raw_material_order,
    )
    standard_columns: List[_StandardOrderColumn] = []
    if not custom_layout and not is_raw_material_order:
        standard_columns = _standard_order_visible_columns(
            display_items,
            total_surface_m2=total_surface_m2,
            total_weight_kg=total_weight_kg,
        )

    # Headers and data
    if custom_layout:
        head = []
        for column in column_layout:
            header = _pdf_order_column_label(column)
            if not header:
                header = column.get("key", "")
            column["label"] = header
            align = _to_str(column.get("justify") or "left").strip().upper() or "LEFT"
            if align not in {"LEFT", "RIGHT", "CENTER"}:
                align = "LEFT"
            head.append(
                _order_table_header_cell(
                    header,
                    align=align,
                    text_color=palette["accent_text"],
                    font_size=9.5,
                )
            )
    elif is_raw_material_order:
        head = ["Profiel", "Materiaal", "Lengte", "St.", "kg"]
    else:
        head = [column.pdf_label for column in standard_columns]

    def wrap_cell_html(val: str, small=False, align=None):
        style = ParagraphStyle(
            "cellsmall" if small else "cell",
            fontName="Helvetica",
            fontSize=8.5 if small else 9,
            leading=10.5 if small else 11,
            wordWrap="CJK",
        )
        if align:
            style.alignment = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align.upper(), 0)
        return Paragraph(str(val if (val is not None) else ""), style)

    standard_col_widths: List[float] | None = None
    if not custom_layout and not is_raw_material_order:
        standard_col_widths = _standard_order_column_widths(
            standard_columns,
            display_items,
            total_width=usable_w,
            small_font_size=8.5,
            header_font_size=10,
        )

    def description_cell_html(val: object, width: float) -> str:
        lines = _wrap_words_to_lines(
            _clean_order_cell_text(val),
            max(24.0, width),
            "Helvetica",
            9,
            max_lines=2,
        )
        return "<br/>".join(escape(line) for line in lines)

    data = [head]
    total_row_index: int | None = None
    if custom_layout:
        surface_idx: int | None = None
        weight_idx: int | None = None
        total_price_idx: int | None = None
        for idx, column in enumerate(column_layout):
            if surface_idx is None and is_order_surface_column(column):
                surface_idx = idx
            if weight_idx is None and is_order_weight_column(column):
                weight_idx = idx
            if total_price_idx is None and _to_str(column.get("key")).strip().lower() == _PRICE_TOTAL_KEY.lower():
                total_price_idx = idx
        for it in display_items:
            row_cells: List[Paragraph] = []
            for idx, column in enumerate(column_layout):
                key = column.get("key")
                value = it.get(key, "") if key else ""
                if column.get("numeric"):
                    if column.get("integer"):
                        value = _format_integer_like(value)
                    else:
                        value = _num_to_2dec(value)
                    small = True
                else:
                    value = _to_str(value)
                    small = False
                align = _to_str(column.get("justify") or "left").strip().upper() or "LEFT"
                if align not in {"LEFT", "RIGHT", "CENTER"}:
                    align = "LEFT"
                row_cells.append(wrap_cell_html(value, small=small, align=align))
            data.append(row_cells)

        if (
            (surface_idx is not None and total_surface_m2 is not None)
            or (weight_idx is not None and total_weight_kg is not None)
            or (total_price_idx is not None and _to_str(price_subtotal).strip())
        ):
            total_row: List[Paragraph] = []
            for idx, column in enumerate(column_layout):
                align = _to_str(column.get("justify") or "left").strip().upper() or "LEFT"
                if align not in {"LEFT", "RIGHT", "CENTER"}:
                    align = "LEFT"
                if idx == surface_idx and total_surface_m2 is not None:
                    surface_text = _num_to_2dec(total_surface_m2)
                    total_row.append(wrap_cell_html(surface_text, small=True, align=align))
                elif idx == weight_idx and total_weight_kg is not None:
                    weight_text = _num_to_2dec(total_weight_kg)
                    total_row.append(wrap_cell_html(weight_text, small=True, align=align))
                elif idx == total_price_idx and _to_str(price_subtotal).strip():
                    subtotal_text = _num_to_2dec(price_subtotal)
                    total_row.append(wrap_cell_html(subtotal_text, small=True, align=align))
                elif idx == 0:
                    total_row.append(wrap_cell_html("Totaal", small=False, align="LEFT"))
                else:
                    total_row.append(
                        wrap_cell_html("", small=bool(column.get("numeric")), align=align)
                    )
            data.append(total_row)
            total_row_index = len(data) - 1
    elif is_raw_material_order:
        for it in display_items:
            prof = _to_str(it.get("Profiel", ""))
            mat = _to_str(it.get("Materiaal", ""))
            length_val = it.get("Lengte", "")
            length = _to_str("" if length_val in (None, "") else length_val)
            qty_val = it.get("St.", "")
            qty = _to_str("" if qty_val in (None, "") else qty_val)
            weight_val = it.get("kg", "")
            weight = _num_to_2dec(weight_val)
            data.append(
                [
                    wrap_cell_html(prof, small=False, align="LEFT"),
                    wrap_cell_html(mat, small=False, align="LEFT"),
                    wrap_cell_html(length, small=True, align="RIGHT"),
                    wrap_cell_html(qty, small=True, align="RIGHT"),
                    wrap_cell_html(weight, small=True, align="RIGHT"),
                ]
            )
        if total_weight_kg is not None:
            total_row_index = len(data)
            total_row = [
                wrap_cell_html("Totaal", small=False, align="LEFT"),
                wrap_cell_html("", small=False, align="LEFT"),
                wrap_cell_html("", small=True, align="RIGHT"),
                wrap_cell_html("", small=True, align="RIGHT"),
                wrap_cell_html(_num_to_2dec(total_weight_kg), small=True, align="RIGHT"),
            ]
            data.append(total_row)
            total_row_index = len(data) - 1
    else:
        standard_widths_by_key = {
            column.key: width
            for column, width in zip(standard_columns, standard_col_widths or [])
        }
        for it in display_items:
            row_cells: List[Paragraph] = []
            for column in standard_columns:
                if column.key == "PartNumber":
                    value = escape(_clean_order_cell_text(it.get("PartNumber", "")))
                elif column.key == "Description":
                    desc_width = max(
                        24.0,
                        standard_widths_by_key.get("Description", usable_w * 0.40)
                        - 10,
                    )
                    value = description_cell_html(it.get("Description", ""), desc_width)
                elif column.key == "Materiaal":
                    value = _material_nowrap(_clean_order_cell_text(it.get("Materiaal", "")))
                elif column.key == "Aantal":
                    value = it.get("Aantal", "")
                elif column.key == "Oppervlakte":
                    value = _num_to_2dec(it.get("Oppervlakte", ""))
                elif column.key == "Gewicht":
                    value = _num_to_2dec(it.get("Gewicht", ""))
                else:
                    value = it.get(column.key, "")
                row_cells.append(
                    wrap_cell_html(
                        value,
                        small=column.small,
                        align=column.align,
                    )
                )
            data.append(row_cells)
        if total_surface_m2 is not None or total_weight_kg is not None:
            total_row = []
            for idx, column in enumerate(standard_columns):
                if idx == 0:
                    value = "Totaal"
                    small = False
                    align = "LEFT"
                elif column.key == "Oppervlakte" and total_surface_m2 is not None:
                    value = _num_to_2dec(total_surface_m2)
                    small = True
                    align = column.align
                elif column.key == "Gewicht" and total_weight_kg is not None:
                    value = _num_to_2dec(total_weight_kg)
                    small = True
                    align = column.align
                else:
                    value = ""
                    small = column.small
                    align = column.align
                total_row.append(wrap_cell_html(value, small=small, align=align))
            data.append(total_row)
            total_row_index = len(data) - 1

    if custom_layout and column_layout:
        weights: List[float] = []
        for column in column_layout:
            try:
                weight_val = float(column.get("weight", 0))
            except Exception:
                weight_val = 0.0
            weights.append(weight_val if weight_val > 0 else 1.0)
        minimums = [
            _order_table_column_min_width(
                column.get("label"),
                [
                    item.get(column.get("key"), "") if column.get("key") else ""
                    for item in display_items
                ],
                header_font_size=9.5,
                body_font_size=8.5 if column.get("numeric") else 9,
                wrap=bool(column.get("wrap")),
                max_width=usable_w * 0.34,
            )
            for column in column_layout
        ]
        col_widths = _weighted_widths_with_minimums(usable_w, weights, minimums)
    elif is_raw_material_order:
        col_fracs = [0.32, 0.24, 0.16, 0.12, 0.16]
        col_widths = [usable_w * frac for frac in col_fracs]
    else:
        col_widths = standard_col_widths or [
            usable_w / max(1, len(standard_columns))
        ] * max(1, len(standard_columns))

    tbl = LongTable(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["accent"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(palette["accent_text"])),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor(ORDER_TEXT_COLOR)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(ORDER_TABLE_OUTLINE_COLOR)),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor(ORDER_TABLE_GRID_COLOR)),
        (
            "ROWBACKGROUNDS",
            (0, 1),
            (-1, -1),
            [colors.white, colors.HexColor(ORDER_TABLE_ALT_ROW_COLOR)],
        ),
        ("LEFTPADDING", (0, 0), (-1, 0), 6),
        ("RIGHTPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("LEFTPADDING", (0, 1), (-1, -1), 5),
        ("RIGHTPADDING", (0, 1), (-1, -1), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]
    if custom_layout and column_layout:
        for idx, column in enumerate(column_layout):
            align = _to_str(column.get("justify") or "left").strip().upper() or "LEFT"
            if align not in {"LEFT", "RIGHT", "CENTER"}:
                align = "LEFT"
            style_cmds.append(("ALIGN", (idx, 0), (idx, -1), align))
    elif is_raw_material_order:
        style_cmds.append(("ALIGN", (2, 0), (4, -1), "RIGHT"))
    else:
        for idx, column in enumerate(standard_columns):
            style_cmds.append(("ALIGN", (idx, 0), (idx, -1), column.align))
    if total_row_index is not None:
        style_cmds.extend(
            [
                ("FONTNAME", (0, total_row_index), (-1, total_row_index), "Helvetica-Bold"),
                (
                    "BACKGROUND",
                    (0, total_row_index),
                    (-1, total_row_index),
                    colors.HexColor(palette["total_fill"]),
                ),
                (
                    "LINEABOVE",
                    (0, total_row_index),
                    (-1, total_row_index),
                    0.45,
                    colors.HexColor(ORDER_TABLE_OUTLINE_COLOR),
                ),
            ]
        )
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    _append_order_price_summary_story(
        story,
        price_summary_rows,
        usable_w=usable_w,
        palette=palette,
    )

    if en1090_required:
        note_text = EN1090_NOTE_TEXT if en1090_note is None else _to_str(en1090_note)
        if note_text:
            story.append(Spacer(0, 12))
            en1090_note_html = note_text.replace("\n", "<br/>")
            if note_text == EN1090_NOTE_TEXT:
                en1090_note_html = f"<b>{en1090_note_html}</b>"
            story.append(Paragraph(en1090_note_html, small_style))

    # Optional brief explanatory note for bruto materiaal (controlled by GUI checkbox).
    # Show only when checkbox enabled AND a supplier is present AND the
    # production or supplier product_type indicates tube laser work.
    if include_bruto_note and supplier is not None:
        prod_lower = _to_str(production).strip().lower() if production else ""
        supp_prod_type = (
            _to_str(getattr(supplier, "product_type", "")).strip().lower()
        )

        tube_keywords = ("tube", "tube laser", "tube laser cutting", "tube-laser")

        def _has_tube_kw(text: str) -> bool:
            return any(k in text for k in tube_keywords if k)

        if _has_tube_kw(prod_lower) or _has_tube_kw(supp_prod_type):
            story.append(Spacer(0, 12))
            bruto_note = (
                "Deze bruto materiaalbon is aanvullende productie-informatie voor snijwerk. "
                "De bon geeft aan hoeveel bruto profielmateriaal nodig is om de snedes uit te voeren. "
                "Alleen ter ondersteuning van productie; vervangt geen order of factuur."
            )
            story.append(Paragraph(bruto_note, small_style))

    note = footer_note_for_doc_type(doc_type_text, footer_note, quote_footer_note)
    if note:
        story.append(Spacer(0, 8))
        story.append(Paragraph(note, small_style))

    doc.build(story)


def generate_packlist_pdf(
    path: str,
    production: str,
    previews: List[dict],
    doc_date: str | None = None,
    columns: int = 2,
) -> bool:
    """Generate a packing list PDF containing thumbnails."""

    if not REPORTLAB_OK or not previews:
        return False
    columns = max(1, int(columns))
    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        title=f"Paklijst {production}",
    )
    styles = getSampleStyleSheet()
    title_style = styles["Heading1"]
    subtitle_style = styles["Normal"]
    subtitle_style.leading = 12
    label_style = ParagraphStyle(
        "PacklistLabel",
        parent=styles["Normal"],
        alignment=1,
        leading=11,
    )
    story: List[object] = []
    story.append(Paragraph(f"Paklijst – {production}", title_style))
    if doc_date:
        story.append(Paragraph(f"Datum: {doc_date}", subtitle_style))
    story.append(Spacer(0, 8 * mm))

    data: List[List[object]] = []
    row: List[object] = []
    usable_width = doc.width
    col_width = usable_width / columns
    image_width = col_width
    image_height = col_width
    for entry in previews:
        thumb = entry.get("thumbnail")
        label = entry.get("label") or os.path.basename(entry.get("source", ""))
        try:
            img = RLImage(thumb, width=image_width, height=image_height)
        except Exception:
            continue
        cell = KeepTogether([img, Spacer(0, 4), Paragraph(label, label_style)])
        row.append(cell)
        if len(row) == columns:
            data.append(row)
            row = []
    if row:
        while len(row) < columns:
            row.append(Spacer(0, 0))
        data.append(row)

    if not data:
        return False

    table = Table(data, colWidths=[col_width] * columns, hAlign="CENTER")
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return True


def write_order_excel(
    path: str,
    items: List[Dict[str, object]],
    company_info: Dict[str, str] | None = None,
    supplier: Supplier | None = None,
    delivery: DeliveryAddress | None = None,
    doc_type: str = "Bestelbon",
    doc_number: str | None = None,
    project_number: str | None = None,
    project_name: str | None = None,
    context_label: str | None = None,
    context_kind: str = "productie",
    order_remark: str | None = None,
    total_surface_m2: float | None = None,
    total_weight_kg: float | None = None,
    en1090_required: bool = False,
    en1090_note: Optional[str] = None,
    column_layout: Optional[List[Dict[str, object]]] = None,
    sections: Optional[List[OrderDocumentSection]] = None,
) -> None:
    """Write order information to an Excel file with header info."""
    context_kind_clean = (_to_str(context_kind) or "productie").strip() or "productie"
    is_raw_material_order = context_kind_clean.lower().startswith("brutemateriaal")
    excel_sections = (
        _normalize_order_sections(
            context_label or "",
            items,
            context_kind_clean,
            total_surface_m2,
            total_weight_kg,
            column_layout,
            sections,
        )
        if sections
        else []
    )
    column_layout = [dict(col) for col in column_layout] if column_layout else []
    custom_layout = bool(column_layout)
    display_items, price_subtotal, price_summary_rows = _split_order_price_summary_rows(
        items,
        column_layout,
        is_raw=is_raw_material_order,
    )
    df_columns: List[str] = []
    if excel_sections:
        df = pd.DataFrame()
    elif custom_layout:
        headers: List[str] = []
        for column in column_layout:
            header = _to_str(column.get("label") or column.get("key") or "").strip()
            if not header:
                header = column.get("key", "")
            column["label"] = header
            headers.append(header)
        rows: List[Dict[str, object]] = []
        for item in display_items:
            row: Dict[str, object] = {}
            for column, header in zip(column_layout, headers):
                key = column.get("key")
                value = item.get(key, "") if key else ""
                if column.get("integer"):
                    value = _coerce_integer_like(value)
                row[header] = value
            rows.append(row)
        df = pd.DataFrame(rows, columns=headers)
        surface_header: str | None = None
        weight_header: str | None = None
        total_price_header: str | None = None
        for column in column_layout:
            if surface_header is None and is_order_surface_column(column):
                surface_header = column["label"]
            if weight_header is None and is_order_weight_column(column):
                weight_header = column["label"]
            if total_price_header is None and _to_str(column.get("key")).strip().lower() == _PRICE_TOTAL_KEY.lower():
                total_price_header = column["label"]
        if (
            (surface_header and total_surface_m2 is not None)
            or (weight_header and total_weight_kg is not None)
            or (total_price_header and _to_str(price_subtotal).strip())
        ):
            total_row = {header: "" for header in headers}
            if headers:
                total_row[headers[0]] = "Totaal"
            if surface_header and total_surface_m2 is not None:
                total_row[surface_header] = _format_weight_kg(total_surface_m2)
            if weight_header and total_weight_kg is not None:
                total_row[weight_header] = _format_weight_kg(total_weight_kg)
            if total_price_header and _to_str(price_subtotal).strip():
                total_row[total_price_header] = price_subtotal
            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        df = _append_price_summary_rows_to_df(df, headers, price_summary_rows)
    else:
        if is_raw_material_order:
            df_columns = ["Profiel", "Materiaal", "Lengte", "St.", "kg"]
            df = pd.DataFrame(display_items, columns=df_columns)
            if total_weight_kg is not None:
                total_row = {
                    "Profiel": "Totaal",
                    "Materiaal": "",
                    "Lengte": "",
                    "St.": "",
                    "kg": _format_weight_kg(total_weight_kg),
                }
                df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        else:
            df = _build_standard_order_excel_dataframe(
                display_items,
                total_surface_m2=total_surface_m2,
                total_weight_kg=total_weight_kg,
            )
            df_columns = list(df.columns)

    note_text = EN1090_NOTE_TEXT if en1090_note is None else _to_str(en1090_note)

    append_note_to_df = (not excel_sections) and en1090_required and note_text and (
        Alignment is None or not hasattr(pd, "ExcelWriter")
    )
    if append_note_to_df:
        blank_row = {col: "" for col in df_columns}
        note_row = {col: "" for col in df_columns}
        if df_columns:
            note_row[df_columns[0]] = note_text
        df = pd.concat(
            [df, pd.DataFrame([blank_row, note_row])], ignore_index=True
        )

    doc_type_text = (_to_str(doc_type).strip() or "Bestelbon")
    doc_type_text_lower = doc_type_text.lower()
    doc_type_text_slug = re.sub(r"[^0-9a-z]+", "", doc_type_text_lower)
    is_standaard_doc = doc_type_text_lower.startswith("standaard")
    order_remark_text = _to_str(order_remark) if order_remark is not None else ""
    order_remark_has_content = bool(order_remark_text.strip())
    place_remark_in_delivery_block = _should_place_remark_in_delivery_block(
        order_remark_has_content=order_remark_has_content,
        doc_type_text_slug=doc_type_text_slug,
        is_standaard_doc=is_standaard_doc,
        delivery=delivery,
    )

    header_lines: List[Tuple[str, str]] = []
    today = datetime.date.today().strftime("%Y-%m-%d")
    if doc_number:
        header_lines.append(("Nummer", str(doc_number)))
    header_lines.append(("Datum", today))
    if excel_sections and len(excel_sections) > 1:
        group_summary = _order_group_summary_text(excel_sections)
        if group_summary:
            header_lines.append(("Gecombineerde bon voor", group_summary))
    if context_label:
        header_lines.append((context_kind_clean.capitalize(), context_label))
    if project_number:
        header_lines.append(("Projectnummer", project_number))
    if project_name:
        header_lines.append(("Projectnaam", project_name))
    if order_remark_has_content and not place_remark_in_delivery_block:
        header_lines.append(("Opmerking", order_remark_text))
    header_lines.append(("", ""))
    if company_info:
        header_lines.extend(
            [
                ("Bedrijf", company_info.get("name", "")),
                ("Adres", company_info.get("address", "")),
                ("BTW", company_info.get("vat", "")),
                ("E-mail", company_info.get("email", "")),
            ]
        )
        website = _to_str(company_info.get("website")).strip()
        if website:
            header_lines.append(("Website", website))
        header_lines.append(("", ""))
    supplier_name = _to_str(supplier.supplier).strip() if supplier else ""
    include_supplier_block = supplier is not None and (
        not is_standaard_doc or bool(supplier_name)
    )
    if include_supplier_block:
        full_addr = format_supplier_address(supplier)
        header_lines.extend(
            [
                ("Leverancier", supplier.supplier),
                ("Adres", full_addr),
                ("BTW", supplier.btw or ""),
                ("E-mail", supplier.sales_email or ""),
                ("Tel", supplier.phone or ""),
                ("", ""),
            ]
        )
    include_delivery_block = False
    if delivery is not None:
        delivery_has_content = any(
            _to_str(value).strip()
            for value in (delivery.name, delivery.address, delivery.remarks)
        )
        include_delivery_block = (
            not is_standaard_doc
            or delivery_has_content
            or place_remark_in_delivery_block
        )
    elif place_remark_in_delivery_block and not is_standaard_doc:
        include_delivery_block = True
    if include_delivery_block:
        if delivery:
            header_lines.extend(
                [
                    ("Leveradres", ""),
                    ("", delivery.name),
                    ("Adres", delivery.address or ""),
                    ("Opmerking", delivery.remarks or ""),
                ]
            )
        if place_remark_in_delivery_block and order_remark_has_content:
            header_lines.append(("Opmerking", order_remark_text))
        header_lines.append(("", ""))

    startrow = len(header_lines)
    if Alignment is not None and hasattr(pd, "ExcelWriter"):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            if excel_sections:
                pd.DataFrame().to_excel(
                    writer,
                    index=False,
                    header=False,
                    sheet_name="Bestelbon",
                )
            else:
                df.to_excel(writer, index=False, startrow=startrow, sheet_name="Bestelbon")
            ws = writer.sheets[list(writer.sheets.keys())[0]]
            for r, (label, value) in enumerate(header_lines, start=1):
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=value)

            if excel_sections:
                row_cursor = startrow + 1
                for section in excel_sections:
                    title = _format_order_section_title(
                        section.context_kind,
                        section.context_label,
                    )
                    title_cell = ws.cell(row=row_cursor, column=1, value=title)
                    if Font is not None:
                        title_cell.font = Font(bold=True)
                    row_cursor += 1

                    section_df, left_cols, wrap_cols = _build_order_excel_section_data(section)
                    section_df.to_excel(
                        writer,
                        index=False,
                        startrow=row_cursor - 1,
                        sheet_name="Bestelbon",
                    )
                    for col_idx, col_name in enumerate(section_df.columns, start=1):
                        align = Alignment(
                            horizontal="left" if col_name in left_cols else "right",
                            wrap_text=col_name in wrap_cols,
                        )
                        if col_name in {"PartNumber", "Profiel"} and get_column_letter is not None:
                            ws.column_dimensions[get_column_letter(col_idx)].width = 25
                        for row in range(row_cursor + 1, row_cursor + len(section_df) + 2):
                            ws.cell(row=row, column=col_idx).alignment = align
                    row_cursor += len(section_df) + 3

                if en1090_required and note_text:
                    note_row = row_cursor + 1
                    cell = ws.cell(row=note_row, column=1, value=note_text)
                    if Font is not None:
                        cell.font = Font(bold=True)
                    if Alignment is not None:
                        cell.alignment = Alignment(horizontal="left", wrap_text=True)
                return

            if custom_layout:
                left_cols = {
                    _to_str(column.get("label") or column.get("key") or "").strip()
                    or column.get("key", "")
                    for column in column_layout
                    if _to_str(column.get("justify") or "left").strip().lower() != "right"
                }
                wrap_cols = {
                    _to_str(column.get("label") or column.get("key") or "").strip()
                    or column.get("key", "")
                    for column in column_layout
                    if bool(column.get("wrap"))
                }
            else:
                if is_raw_material_order:
                    left_cols = {"Profiel", "Materiaal"}
                    wrap_cols = {"Profiel", "Materiaal"}
                else:
                    visible_columns = _standard_order_visible_columns(
                        display_items,
                        total_surface_m2=total_surface_m2,
                        total_weight_kg=total_weight_kg,
                    )
                    left_cols = _standard_order_left_columns(visible_columns)
                    wrap_cols = _standard_order_wrap_columns(visible_columns)
            for col_idx, col_name in enumerate(df.columns, start=1):
                align = Alignment(
                    horizontal="left" if col_name in left_cols else "right",
                    wrap_text=col_name in wrap_cols,
                )
                if (
                    not custom_layout
                    and col_name in {"PartNumber", "Profiel"}
                    and get_column_letter is not None
                ):
                    column_letter = get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = 25
                for row in range(startrow + 1, startrow + len(df) + 2):
                    ws.cell(row=row, column=col_idx).alignment = align

            if en1090_required and note_text and not append_note_to_df:
                note_row = ws.max_row + 2
                cell = ws.cell(row=note_row, column=1, value=note_text)
                if Font is not None:
                    cell.font = Font(bold=True)
                if Alignment is not None:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)





def pick_supplier_for_production(
    prod: str,
    db: SuppliersDB,
    override_map: Dict[str, str],
    suppliers_sorted: List[Supplier] | None = None,
) -> Supplier:
    """Select a supplier for a given production.

    ``suppliers_sorted`` allows callers to provide a pre-sorted supplier list in
    order to avoid repeated :meth:`SuppliersDB.suppliers_sorted` lookups.
    """
    name = override_map.get(prod)
    sups = suppliers_sorted if suppliers_sorted is not None else db.suppliers_sorted()
    if name is not None:
        if not name.strip():
            return Supplier(supplier="")
        for s in sups:
            if s.supplier.lower() == name.lower():
                return s
        return Supplier(supplier=name)
    if prod.strip().lower() in {"dummy part", "nan", "spare part"}:
        return Supplier(supplier="")
    default = db.get_default(prod)
    if default:
        for s in sups:
            if s.supplier.lower() == default.lower():
                return s
    # Geen eerder geselecteerde leverancier onthouden: vul de placeholder in.
    return Supplier(supplier=NO_SUPPLIER_PLACEHOLDER)


def pick_supplier_for_opticutter(
    prod: str,
    db: SuppliersDB,
    override_map: Dict[str, str],
    suppliers_sorted: List[Supplier] | None = None,
) -> Supplier:
    """Select a supplier for Opticutter raw material orders."""

    key = make_opticutter_default_key(prod)
    name = override_map.get(prod)
    sups = suppliers_sorted if suppliers_sorted is not None else db.suppliers_sorted()
    if name is not None:
        if not name.strip():
            return Supplier(supplier="")
        for s in sups:
            if s.supplier.lower() == name.lower():
                return s
        return Supplier(supplier=name)
    default = db.get_default(key)
    if default:
        for s in sups:
            if s.supplier.lower() == default.lower():
                return s
    return Supplier(supplier=NO_SUPPLIER_PLACEHOLDER)


def pick_supplier_for_finish(
    finish_key: str,
    db: SuppliersDB,
    override_map: Dict[str, str],
    suppliers_sorted: List[Supplier] | None = None,
) -> Supplier:
    """Select a supplier for a finish combination."""

    name = override_map.get(finish_key)
    sups = suppliers_sorted if suppliers_sorted is not None else db.suppliers_sorted()
    if name is not None:
        if not name.strip():
            return Supplier(supplier="")
        for s in sups:
            if s.supplier.lower() == name.lower():
                return s
        return Supplier(supplier=name)
    default = db.get_default_finish(finish_key)
    if default:
        for s in sups:
            if s.supplier.lower() == default.lower():
                return s
    return Supplier(supplier=NO_SUPPLIER_PLACEHOLDER)




def pick_supplier_for_spare_part_group(
    group_key: str,
    default_supplier: str,
    db: SuppliersDB,
    override_map: Dict[str, str],
    suppliers_sorted: List[Supplier] | None = None,
) -> Supplier:
    """Select a supplier for a spare-parts order group."""

    name = override_map.get(group_key)
    sups = suppliers_sorted if suppliers_sorted is not None else db.suppliers_sorted()
    if name is not None:
        if not name.strip():
            return Supplier(supplier="")
        for s in sups:
            if s.supplier.lower() == name.lower():
                return s
        return Supplier(supplier=name)

    default = db.get_default(make_spare_part_default_key(group_key))
    if default:
        for s in sups:
            if s.supplier.lower() == default.lower():
                return s
        return Supplier(supplier=default)

    default_supplier = _to_str(default_supplier).strip()
    if default_supplier:
        for s in sups:
            if s.supplier.lower() == default_supplier.lower():
                return s
        return Supplier(supplier=default_supplier)

    return Supplier(supplier="")


def copy_per_production_and_orders(
    source: str,
    dest: str,
    bom_df: pd.DataFrame,
    selected_exts: List[str],
    db: SuppliersDB,
    override_map: Dict[str, str],
    doc_type_map: Dict[str, str] | None,
    doc_num_map: Dict[str, str] | None,
    remember_defaults: bool,
    client: Client | None = None,
    delivery_map: Dict[str, DeliveryAddress] | None = None,
    footer_note: Optional[str] = None,
    quote_footer_note: Optional[str] = None,
    zip_parts: bool = False,
    date_prefix_exports: bool = False,
    date_suffix_exports: bool = False,
    project_number: str | None = None,
    project_name: str | None = None,
    export_name_prefix_text: str = "",
    export_name_prefix_enabled: bool | None = None,
    export_name_suffix_text: str = "",
    export_name_suffix_enabled: bool | None = None,
    document_filename_profile: str = DOCUMENT_FILENAME_PROFILE_STANDARD,
    document_filename_show_doc_type: bool = True,
    document_filename_show_doc_number: bool = True,
    document_filename_show_context: bool = True,
    document_filename_show_date: bool = True,
    document_filename_compact_doc_number: bool = False,
    document_filename_separator: str = "underscore",
    document_display_compact_doc_number: bool = False,
    copy_finish_exports: bool = False,
    zip_finish_exports: bool = True,
    export_bom: bool = True,
    export_related_files: bool = True,
    finish_override_map: Dict[str, str] | None = None,
    finish_doc_type_map: Dict[str, str] | None = None,
    finish_doc_num_map: Dict[str, str] | None = None,
    finish_delivery_map: Dict[str, DeliveryAddress | None] | None = None,
    remarks_map: Dict[str, str] | None = None,
    finish_remarks_map: Dict[str, str] | None = None,
    document_group_map: Mapping[str, str] | None = None,
    bom_source_path: str | None = None,
    path_limit_warnings: List[str] | None = None,
    opticutter_analysis: OpticutterAnalysis | None = None,
    opticutter_choices: Mapping[tuple[str, str, str], str] | None = None,
    opticutter_override_map: Dict[str, str] | None = None,
    opticutter_doc_type_map: Dict[str, str] | None = None,
    opticutter_doc_num_map: Dict[str, str] | None = None,
    opticutter_delivery_map: Dict[str, DeliveryAddress | None] | None = None,
    opticutter_remarks_map: Dict[str, str] | None = None,
    pricing_map: Mapping[str, Mapping[str, object]] | None = None,
    finish_pricing_map: Mapping[str, Mapping[str, object]] | None = None,
    opticutter_pricing_map: Mapping[str, Mapping[str, object]] | None = None,
    spare_part_pricing_map: Mapping[str, Mapping[str, object]] | None = None,
    vat_rate_map: Mapping[str, object] | None = None,
    finish_vat_rate_map: Mapping[str, object] | None = None,
    opticutter_vat_rate_map: Mapping[str, object] | None = None,
    spare_part_vat_rate_map: Mapping[str, object] | None = None,
    production_export_filter: Mapping[str, bool] | None = None,
    finish_export_filter: Mapping[str, bool] | None = None,
    opticutter_export_filter: Mapping[str, bool] | None = None,
    spare_part_groups: Sequence[Mapping[str, object]] | None = None,
    spare_part_override_map: Dict[str, str] | None = None,
    spare_part_doc_type_map: Dict[str, str] | None = None,
    spare_part_doc_num_map: Dict[str, str] | None = None,
    spare_part_delivery_map: Dict[str, DeliveryAddress | None] | None = None,
    spare_part_remarks_map: Dict[str, str] | None = None,
    spare_part_export_filter: Mapping[str, bool] | None = None,
    en1090_enabled: bool = True,
    en1090_overrides: Mapping[str, bool] | None = None,
    en1090_note: Optional[str] = None,
    document_status_messages: List[str] | None = None,
    generated_documents: List[Dict[str, object]] | None = None,
    progress_callback: ProgressCallback | None = None,
) -> Tuple[int, Dict[str, str]]:
    """Copy files per production and create accompanying order documents.

    ``doc_type_map`` may specify per production whether a *Bestelbon* or an
    *Offerteaanvraag* should be generated. Missing entries default to
    ``"Bestelbon"``.

    ``doc_num_map`` provides document numbers per production which are used in
    filenames and document headers. The filename and displayed document number
    can be formatted independently through the document filename/display flags.

    ``delivery_map`` can provide a :class:`DeliveryAddress` per production.

    ``remarks_map`` and ``finish_remarks_map`` allow passing additional notes for
    productions and finishes respectively. When provided, the remarks are added
    to the generated Excel- en PDF-bestanden.

    ``opticutter_analysis`` en ``opticutter_choices`` laten toe om per
    productie extra Opticutter-overzichten en bestelbonnen voor brutemateriaal
    aan te maken. Wanneer beide waarden aanwezig zijn wordt in iedere
    productiemap een Opticutter-werkboek met scenario-informatie en een
    besteloverzicht voor volle lengten geschreven. Extra kaarten voor
    Opticutter-bestellingen kunnen voorzien worden via
    ``opticutter_override_map``, ``opticutter_doc_type_map``,
    ``opticutter_doc_num_map``, ``opticutter_delivery_map`` en
    ``opticutter_remarks_map``.

    Gebruik ``production_export_filter``, ``finish_export_filter`` of
    ``opticutter_export_filter`` om respectievelijk producties,
    afwerkingen of Opticutter-selecties over te slaan zonder ze uit de BOM
    te verwijderen. Wanneer een filtermap ``False`` bevat voor een bepaalde
    sleutel wordt er niets gekopieerd of aangemaakt voor die selectie.

    If ``zip_parts`` is ``True``, all export files for a production are
    collected into a single ``<production>.zip`` archive instead of individual
    ``PartNumber`` files. Only the generated order Excel/PDF remain unzipped in
    the production folder.

    When ``date_prefix_exports`` is ``True`` the copied export files will start
    with ``YYYYMMDD-``. When ``date_suffix_exports`` is ``True`` they will end
    with ``-YYYYMMDD`` before the extension. Both transformations are applied
    consistently to copied files and ZIP archive members.

    When custom export prefix/suffix tokens are provided and enabled they are
    added before and/or after the filename. Both transformations are applied
    consistently to copied files and ZIP archive members. The enable flags
    default to active when the corresponding text is non-empty unless
    explicitly set to ``False``.

    When ``copy_finish_exports`` is ``True`` each referenced export file is
    additionally copied to folders named ``Finish-<finish>`` with an optional
    ``-<ral>`` suffix when the BOM "RAL color" property is filled in. The
    folder components are normalized versions of the BOM "Finish" and
    ``RAL color`` values. When ``zip_finish_exports`` is ``True`` (default)
    the finish folders receive a single ZIP archive containing the export
    files instead of loose copies.

    When ``export_bom`` is ``True`` (default) the processed BOM, including any
    changes made inside Filehopper, is written as an Excel workbook in the root
    of ``dest``. The filename contains the export date in ISO-formaat.

    When ``bom_source_path`` refers to the loaded BOM file, Filehopper tries to
    copy export files from ``source`` whose filename stem appears inside the BOM
    name (for example ``123-BOM-PartsOnly`` → ``123.pdf``). Matching files are
    placed in the root export folder and use the same optional
    name-transformations (date/prefix/suffix). When no related files are found
    nothing is copied. Set ``export_related_files`` to ``False`` to skip copying
    these auxiliary files.

    Finish-specific overrides, document types/numbers and deliveries can be
    provided via the ``finish_*`` mappings. Keys correspond to the normalized
    ``Finish-...`` folder names produced by :func:`describe_finish_combo`.

    The returned ``chosen`` mapping uses selection keys produced by
    :func:`make_production_selection_key` for productions and
    :func:`make_finish_selection_key` for finish selections.

    When ``generated_documents`` is provided, Filehopper appends relative paths
    and document metadata for generated PDF/XLSX files so callers can store a
    traceable export log.
    """
    def _emit_progress(
        phase: str,
        message: str,
        *,
        percent: int | None = None,
        done: int = 0,
        total: int = 0,
    ) -> None:
        if progress_callback is None:
            return
        progress_callback(
            ProgressEvent(
                phase=phase,
                message=message,
                done=done,
                total=total,
                percent=percent,
            )
        )

    def _phase_percent(start: int, end: int, index: int, total: int) -> int:
        if total <= 0:
            return start
        fraction = max(0.0, min(1.0, index / total))
        return int(round(start + (end - start) * fraction))

    _emit_progress("prepare", "Export voorbereiden...", percent=0)
    en1090_active = bool(en1090_enabled)
    en1090_note_text = (
        EN1090_NOTE_TEXT if en1090_note is None else _to_str(en1090_note)
    )
    os.makedirs(dest, exist_ok=True)
    _emit_progress("scan", "Bestanden zoeken...", percent=4)
    file_index = _build_file_index(source, selected_exts)
    selected_exts_set = {ext.lower() for ext in selected_exts}
    count_copied = 0
    chosen: Dict[str, str] = {}
    doc_type_map = doc_type_map or {}
    doc_num_map = doc_num_map or {}
    finish_override_map = finish_override_map or {}
    finish_doc_type_map = finish_doc_type_map or {}
    finish_doc_num_map = finish_doc_num_map or {}
    finish_delivery_map = finish_delivery_map or {}
    opticutter_override_map = opticutter_override_map or {}
    opticutter_doc_type_map = opticutter_doc_type_map or {}
    opticutter_doc_num_map = opticutter_doc_num_map or {}
    opticutter_delivery_map = opticutter_delivery_map or {}
    opticutter_remarks_map = {
        key: _to_str(value).strip()
        for key, value in (opticutter_remarks_map or {}).items()
        if _to_str(value).strip()
    }
    spare_part_groups = [
        dict(group)
        for group in (spare_part_groups or [])
        if isinstance(group, Mapping) and _to_str(group.get("key")).strip()
    ]
    spare_part_override_map = spare_part_override_map or {}
    spare_part_doc_type_map = spare_part_doc_type_map or {}
    spare_part_doc_num_map = spare_part_doc_num_map or {}
    spare_part_delivery_map = spare_part_delivery_map or {}
    spare_part_remarks_map = {
        key: _to_str(value).strip()
        for key, value in (spare_part_remarks_map or {}).items()
        if _to_str(value).strip()
    }
    pricing_map = pricing_map or {}
    finish_pricing_map = finish_pricing_map or {}
    opticutter_pricing_map = opticutter_pricing_map or {}
    spare_part_pricing_map = spare_part_pricing_map or {}
    vat_rate_map = vat_rate_map or {}
    finish_vat_rate_map = finish_vat_rate_map or {}
    opticutter_vat_rate_map = opticutter_vat_rate_map or {}
    spare_part_vat_rate_map = spare_part_vat_rate_map or {}
    remarks_clean: Dict[str, str] = {}
    for key, value in (remarks_map or {}).items():
        text = _to_str(value).strip()
        if text:
            remarks_clean[key] = text
    remarks_map = remarks_clean
    document_group_map = _clean_document_group_map(document_group_map)

    finish_remarks_clean: Dict[str, str] = {}
    for key, value in (finish_remarks_map or {}).items():
        text = _to_str(value).strip()
        if text:
            finish_remarks_clean[key] = text
    finish_remarks_map = finish_remarks_clean

    def _clean_export_filter(values: Mapping[str, bool] | None) -> Dict[str, bool]:
        cleaned: Dict[str, bool] = {}
        for key, flag in (values or {}).items():
            identifier = _to_str(key).strip()
            if not identifier:
                continue
            cleaned[identifier] = bool(flag)
        return cleaned

    production_export_filter = _clean_export_filter(production_export_filter)
    finish_export_filter = _clean_export_filter(finish_export_filter)
    opticutter_export_filter = _clean_export_filter(opticutter_export_filter)
    spare_part_export_filter = _clean_export_filter(spare_part_export_filter)

    def _normalized_production_name(value: object) -> str:
        return _to_str(value).strip() or "_Onbekend"

    def _production_enabled(value: object) -> bool:
        production_name = _normalized_production_name(value)
        if not production_export_filter:
            return True
        return production_export_filter.get(production_name, True)

    opticutter_context: OpticutterExportContext | None = None
    if opticutter_analysis is not None and opticutter_analysis.profiles:
        try:
            opticutter_context = prepare_opticutter_export(
                opticutter_analysis, opticutter_choices or {}
            )
        except Exception:
            opticutter_context = None

    opticutter_details_map: Dict[str, OpticutterOrderComputation] = {}
    opticutter_stats_map: Dict[
        tuple[str, str, str], OpticutterProfileStats
    ] = {}
    if opticutter_context is not None:
        try:
            opticutter_stats_map = _collect_opticutter_profile_stats(bom_df)
            for prod_key, export in opticutter_context.productions.items():
                opticutter_details_map[prod_key] = _compute_opticutter_order_exports(
                    export, opticutter_stats_map
                )
        except Exception:
            opticutter_details_map = {}
            opticutter_stats_map = {}

    prod_to_rows: Dict[str, List[dict]] = defaultdict(list)
    step_entries: Dict[str, List[tuple[str, str]]] = defaultdict(list)
    step_seen: Dict[str, set[str]] = defaultdict(set)
    finish_groups: Dict[str, Dict[str, object]] = {}
    for _, row in bom_df.iterrows():
        prod = _normalized_production_name(row.get("Production"))
        if not _production_enabled(prod):
            continue
        prod_to_rows[prod].append(row)
        pn = _to_str(row.get("PartNumber")).strip()
        finish_text = _to_str(row.get("Finish")).strip()
        if finish_text:
            finish_meta = describe_finish_combo(
                row.get("Finish"), row.get("RAL color")
            )
            finish_key = finish_meta["key"]
            group = finish_groups.get(finish_key)
            if group is None:
                group = {
                    **finish_meta,
                    "rows": [],
                    "part_numbers": set(),
                }
                finish_groups[finish_key] = group
            group["rows"].append(row)
            if pn:
                group["part_numbers"].add(pn)

    today_date = datetime.date.today()
    today = today_date.strftime("%Y-%m-%d")
    date_token = today_date.strftime("%Y%m%d")
    delivery_map = delivery_map or {}
    dest_abs = os.path.abspath(dest)
    export_name_prefix_text = (export_name_prefix_text or "").strip()
    export_name_suffix_text = (export_name_suffix_text or "").strip()
    document_filename_profile = normalize_document_filename_profile(
        document_filename_profile
    )
    document_filename_separator = normalize_document_filename_separator(
        document_filename_separator
    )
    document_display_compact_doc_number = bool(document_display_compact_doc_number)
    prefix_has_text = bool(export_name_prefix_text)
    suffix_has_text = bool(export_name_suffix_text)
    if export_name_prefix_enabled is None:
        token_prefix_active = prefix_has_text
    else:
        token_prefix_active = bool(export_name_prefix_enabled) and prefix_has_text
    if export_name_suffix_enabled is None:
        token_suffix_active = suffix_has_text
    else:
        token_suffix_active = bool(export_name_suffix_enabled) and suffix_has_text

    def _transform_export_name(filename: str) -> str:
        """Apply date/custom tokens to ``filename`` when requested."""

        if not (
            date_prefix_exports
            or date_suffix_exports
            or token_prefix_active
            or token_suffix_active
        ):
            return filename
        stem, ext = os.path.splitext(filename)
        prefix_parts: List[str] = []
        if date_prefix_exports:
            prefix_parts.append(date_token)
        if token_prefix_active:
            prefix_parts.append(export_name_prefix_text)
        suffix_parts: List[str] = []
        if date_suffix_exports:
            suffix_parts.append(date_token)
        if token_suffix_active:
            suffix_parts.append(export_name_suffix_text)
        new_stem = "-".join(prefix_parts + [stem] + suffix_parts)
        return f"{new_stem}{ext}"

    suppliers_sorted = db.suppliers_sorted()

    def _record_path_warning(
        directory: str,
        requested: str,
        final: str,
        *,
        context: str,
    ) -> None:
        if path_limit_warnings is None or requested == final:
            return
        directory_abs = os.path.abspath(directory)
        original_abs = os.path.join(directory_abs, requested)
        detail = (
            f"{context}: '{requested}' → '{final}' "
            f"(padlengte {len(original_abs)} tekens, limiet {_WINDOWS_MAX_PATH})"
        )
        if detail not in path_limit_warnings:
            path_limit_warnings.append(detail)

    def _append_document_status(message: object) -> None:
        if document_status_messages is None:
            return
        clean = _to_str(message).strip()
        if clean and clean not in document_status_messages:
            document_status_messages.append(clean)

    def _report_export_path(path: str) -> str:
        try:
            relative = os.path.relpath(os.path.abspath(path), dest_abs)
        except Exception:
            return os.path.abspath(path)
        if relative.startswith(".."):
            return os.path.abspath(path)
        return relative

    def _append_generated_document(
        path: str,
        *,
        kind: str,
        file_format: str,
        selection_key: str = "",
        selection_keys: Sequence[str] | None = None,
        context_kind: str = "",
        context_label: str = "",
        doc_type: str = "",
        doc_number: str = "",
        supplier: str = "",
    ) -> None:
        if generated_documents is None:
            return
        clean_path = _to_str(path).strip()
        if not clean_path:
            return
        record: Dict[str, object] = {
            "path": _report_export_path(clean_path),
            "kind": _to_str(kind).strip(),
            "format": _to_str(file_format).strip(),
        }
        optional_values = {
            "selection_key": selection_key,
            "context_kind": context_kind,
            "context_label": context_label,
            "doc_type": doc_type,
            "doc_number": doc_number,
            "supplier": supplier,
        }
        for key, value in optional_values.items():
            text = _to_str(value).strip()
            if text:
                record[key] = text
        clean_selection_keys = [
            _to_str(key).strip()
            for key in (selection_keys or [])
            if _to_str(key).strip()
        ]
        if clean_selection_keys:
            record["selection_keys"] = clean_selection_keys
        if record not in generated_documents:
            generated_documents.append(record)

    footer_note_text = (
        None
        if footer_note is None
        else _to_str(footer_note).replace("\r\n", "\n")
    )
    quote_footer_note_text = (
        None
        if quote_footer_note is None
        else _to_str(quote_footer_note).replace("\r\n", "\n")
    )
    company = {
        "name": client.name if client else "",
        "address": client.address if client else "",
        "vat": client.vat if client else "",
        "email": client.email if client else "",
        "website": client.website if client else "",
        "accent_color": client.accent_color if client else "",
        "logo_path": client.logo_path if client else "",
        "logo_crop": client.logo_crop if client else None,
    }
    order_candidates: List[OrderDocumentCandidate] = []
    production_items = list(prod_to_rows.items())
    production_total = len(production_items)
    finish_copy_items = list(finish_groups.items()) if copy_finish_exports and finish_groups else []
    finish_copy_total = len(finish_copy_items)
    finish_doc_items = sorted(
        finish_groups.items(),
        key=lambda item: _to_str(item[1].get("label", "")).lower(),
    )
    finish_doc_total = len(finish_doc_items)

    if production_total:
        _emit_progress(
            "productions",
            f"Producties verwerken... 0 van {production_total}",
            percent=8,
            done=0,
            total=production_total,
        )

    for prod_index, (prod, rows) in enumerate(production_items, start=1):
        _emit_progress(
            "productions",
            f"Productie verwerken... {prod}",
            percent=_phase_percent(8, 45, prod_index - 1, production_total),
            done=prod_index - 1,
            total=production_total,
        )
        if production_export_filter and not production_export_filter.get(prod, True):
            continue
        prod_folder = os.path.join(dest, prod)
        os.makedirs(prod_folder, exist_ok=True)

        opticutter_prod = None
        opticutter_comp: OpticutterOrderComputation | None = None
        if opticutter_context is not None:
            opticutter_prod = opticutter_context.productions.get(prod)
            opticutter_comp = opticutter_details_map.get(prod)

        raw_doc_type = doc_type_map.get(prod, "Bestelbon")
        doc_type = _to_str(raw_doc_type).strip() or "Bestelbon"
        doc_num = _normalize_doc_number(doc_num_map.get(prod, ""), doc_type)
        prefix = _prefix_for_doc_type(doc_type)
        if doc_num and prefix and doc_num.upper() == prefix.upper():
            doc_num = ""
        doc_num_display = format_document_number_for_display(
            doc_num,
            doc_type,
            compact=document_display_compact_doc_number,
        )
        doc_num_token = _sanitize_component(doc_num) if doc_num else ""
        num_part = f"_{doc_num_token}" if doc_num_token else ""
        doc_type_lower = doc_type.lower()
        is_standaard_doc = doc_type_lower.startswith("standaard")

        zf = None
        if zip_parts:
            zip_name = _fit_filename_within_path(
                prod_folder, f"{prod}{num_part}.zip"
            )
            zip_path = os.path.join(prod_folder, zip_name)
            try:
                zf = zipfile.ZipFile(
                    zip_path,
                    "w",
                    compression=zipfile.ZIP_DEFLATED,
                    compresslevel=6,
                )
            except TypeError:
                # ``compresslevel`` not supported (older Python). Retry without it.
                zf = zipfile.ZipFile(
                    zip_path,
                    "w",
                    compression=zipfile.ZIP_DEFLATED,
                )
            except (RuntimeError, NotImplementedError):
                print(
                    "[WAARSCHUWING] ZIP_DEFLATED niet beschikbaar, val terug op ZIP_STORED",
                    file=sys.stderr,
                )
                zf = zipfile.ZipFile(
                    zip_path,
                    "w",
                    compression=zipfile.ZIP_STORED,
                )

        processed_pairs: set[tuple[str, str]] = set()
        for row in rows:
            pn = str(row["PartNumber"])
            files = file_index.get(pn, [])
            for src_file in files:
                transformed = _transform_export_name(os.path.basename(src_file))
                combo = (src_file, transformed)
                if combo in processed_pairs:
                    continue
                processed_pairs.add(combo)
                ext = os.path.splitext(src_file)[1].lower()
                if selected_exts_set and ext not in selected_exts_set:
                    continue
                if ext in STEP_EXTS:
                    seen_paths = step_seen[prod]
                    if src_file not in seen_paths:
                        seen_paths.add(src_file)
                        label = f"{pn} — {os.path.basename(src_file)}"
                        step_entries[prod].append((label, src_file))
                if zip_parts:
                    if zf is not None:
                        zf.write(src_file, arcname=transformed)
                else:
                    dst = os.path.join(prod_folder, transformed)
                    shutil.copy2(src_file, dst)
                count_copied += 1

        if zf is not None:
            zf.close()

        if spare_part_groups and is_spare_parts_production(prod):
            _emit_progress(
                "productions",
                f"Productie verwerkt... {prod}",
                percent=_phase_percent(8, 45, prod_index, production_total),
                done=prod_index,
                total=production_total,
            )
            continue

        supplier = pick_supplier_for_production(
            prod, db, override_map, suppliers_sorted=suppliers_sorted
        )
        chosen[make_production_selection_key(prod)] = supplier.supplier
        if remember_defaults and supplier.supplier not in ("", "Onbekend", NO_SUPPLIER_PLACEHOLDER):
            db.set_default(prod, supplier.supplier)

        en1090_required = should_require_en1090(prod, en1090_overrides)
        if not en1090_active:
            en1090_required = False

        items = []
        for row in rows:
            item = {
                "PartNumber": row.get("PartNumber", ""),
                "Description": row.get("Description", ""),
                "Materiaal": row.get("Materiaal", ""),
                "Lengte": row.get("Length profile", ""),
                "Aantal": _parse_qty(row.get("Aantal", "")),
                "Oppervlakte": row.get("Oppervlakte", ""),
                "Gewicht": row.get("Gewicht", ""),
            }
            item[_LINE_PRICE_KEY] = build_order_pricing_item_key(
                item,
                context_kind="Productie",
            )
            items.append(item)

        delivery = delivery_map.get(prod)
        order_remark = (remarks_map.get(prod, "") if remarks_map else "").strip()
        items, column_layout = _apply_order_pricing(
            items,
            pricing_map.get(prod),
            context_kind="Productie",
            vat_rate=vat_rate_map.get(prod, ""),
        )
        total_surface_m2, total_weight_kg = calculate_order_measure_totals(items)
        order_candidates.append(
            OrderDocumentCandidate(
                selection_key=make_production_selection_key(prod),
                context_label=prod,
                context_kind="Productie",
                filename_context=prod,
                target_dir=prod_folder,
                supplier=supplier,
                delivery=delivery,
                doc_type=doc_type,
                doc_num=doc_num,
                doc_num_display=doc_num_display,
                order_remark=order_remark or None,
                items=items,
                total_surface_m2=total_surface_m2,
                total_weight_kg=total_weight_kg,
                column_layout=column_layout,
                en1090_required=en1090_required,
            )
        )

        opticutter_order_items: List[Dict[str, object]] = []
        opticutter_total_weight: float | None = None
        opticutter_has_selection = bool(
            opticutter_prod is not None and opticutter_prod.selections
        )
        opticutter_allowed = (
            opticutter_export_filter.get(prod, True)
            if opticutter_export_filter
            else True
        )
        if opticutter_has_selection and opticutter_allowed:
            comp = opticutter_comp
            if comp is None:
                stats_map = opticutter_stats_map or {}
                comp = _compute_opticutter_order_exports(opticutter_prod, stats_map)

            settings_rows = [
                {"Parameter": "Exportdatum", "Waarde": today},
                {
                    "Parameter": "Zaagbreedte (kerf) [mm]",
                    "Waarde": opticutter_context.kerf_mm if opticutter_context else None,
                },
            ]
            if opticutter_context and opticutter_context.custom_stock_mm is not None:
                settings_rows.append(
                    {
                        "Parameter": "Aangepaste staaflengte (mm)",
                        "Waarde": opticutter_context.custom_stock_mm,
                    }
                )
            if comp.total_weight_kg is not None:
                settings_rows.append(
                    {
                        "Parameter": "Totaal brutogewicht (kg)",
                        "Waarde": round(comp.total_weight_kg, 2),
                    }
                )
            settings_rows.append({"Parameter": "Productie", "Waarde": prod})

            scenario_df = pd.DataFrame(comp.scenario_rows)
            pieces_df = pd.DataFrame(comp.piece_rows)
            settings_df = pd.DataFrame(settings_rows)
            order_df = pd.DataFrame(comp.order_rows)

            opticutter_requested = f"Opticutter_{prod}_{today}.xlsx"
            opticutter_filename = _fit_filename_within_path(
                prod_folder, opticutter_requested
            )
            _record_path_warning(
                prod_folder,
                opticutter_requested,
                opticutter_filename,
                context=f"Productie '{prod}' – Opticutter",
            )
            opticutter_path = os.path.join(prod_folder, opticutter_filename)
            with pd.ExcelWriter(opticutter_path) as writer:
                scenario_df.to_excel(writer, sheet_name="Scenario", index=False)
                pieces_df.to_excel(writer, sheet_name="Stukken", index=False)
                settings_df.to_excel(writer, sheet_name="Instellingen", index=False)
                if not order_df.empty:
                    order_df.to_excel(writer, sheet_name="Bestelling", index=False)
            _append_generated_document(
                opticutter_path,
                kind="opticutter_workbook",
                file_format="xlsx",
                selection_key=make_opticutter_selection_key(prod),
                context_kind="Brutemateriaal",
                context_label=prod,
            )

            order_overview_path: str | None = None
            should_write_order_overview = not order_df.empty
            if should_write_order_overview:
                order_requested = f"Bestelbon_brutemateriaal_{prod}_{today}.xlsx"
                order_filename = _fit_filename_within_path(
                    prod_folder, order_requested
                )
                _record_path_warning(
                    prod_folder,
                    order_requested,
                    order_filename,
                    context=f"Productie '{prod}' – Brutebestelling",
                )
                order_overview_path = os.path.join(prod_folder, order_filename)

            opticutter_order_items = list(comp.raw_items)
            opticutter_total_weight = comp.total_weight_kg

        if opticutter_has_selection and opticutter_allowed:
            opticutter_sel_key = make_opticutter_selection_key(prod)
            opticutter_supplier = pick_supplier_for_opticutter(
                prod, db, opticutter_override_map, suppliers_sorted=suppliers_sorted
            )
            chosen[opticutter_sel_key] = opticutter_supplier.supplier
            if remember_defaults and opticutter_supplier.supplier not in (
                "",
                "Onbekend",
                NO_SUPPLIER_PLACEHOLDER,
            ):
                db.set_default(
                    make_opticutter_default_key(prod), opticutter_supplier.supplier
                )

            opticutter_doc_type_raw = opticutter_doc_type_map.get(prod, "Bestelbon")
            opticutter_doc_type = (
                _to_str(opticutter_doc_type_raw).strip() or "Bestelbon"
            )
            opticutter_doc_num = _normalize_doc_number(
                opticutter_doc_num_map.get(prod, ""), opticutter_doc_type
            )
            opticutter_prefix = _prefix_for_doc_type(opticutter_doc_type)
            if (
                opticutter_doc_num
                and opticutter_prefix
                and opticutter_doc_num.upper() == opticutter_prefix.upper()
            ):
                opticutter_doc_num = ""
            opticutter_doc_num_display = format_document_number_for_display(
                opticutter_doc_num,
                opticutter_doc_type,
                compact=document_display_compact_doc_number,
            )
            opticutter_doc_lower = opticutter_doc_type.lower()
            opticutter_is_standaard = opticutter_doc_lower.startswith("standaard")

            opticutter_delivery = opticutter_delivery_map.get(prod)
            opticutter_remark_text = opticutter_remarks_map.get(prod, "")
            if opticutter_total_weight is not None:
                weight_line = f"Totaal brutogewicht: {opticutter_total_weight:.2f} kg"
                if opticutter_remark_text:
                    if weight_line not in opticutter_remark_text:
                        opticutter_remark_text = (
                            f"{opticutter_remark_text}\n{weight_line}"
                        )
                else:
                    opticutter_remark_text = weight_line

            opticutter_supplier_name = _to_str(opticutter_supplier.supplier).strip()
            supplier_for_opticutter_docs: Supplier | None = opticutter_supplier
            delivery_for_opticutter_docs = opticutter_delivery
            if opticutter_is_standaard and not opticutter_supplier_name:
                supplier_for_opticutter_docs = None
                delivery_for_opticutter_docs = None

            should_generate_opticutter_order = (
                bool(opticutter_order_items)
                and (opticutter_supplier_name or opticutter_is_standaard)
            )
            if should_generate_opticutter_order:
                should_write_order_overview = False
                opticutter_order_items, opticutter_column_layout = _apply_order_pricing(
                    opticutter_order_items,
                    opticutter_pricing_map.get(prod),
                    context_kind="Brutemateriaal",
                    vat_rate=opticutter_vat_rate_map.get(prod, ""),
                )

                opticutter_document_base = build_document_export_basename(
                    opticutter_doc_type,
                    opticutter_doc_num,
                    prod,
                    today,
                    profile=document_filename_profile,
                    show_doc_type=document_filename_show_doc_type,
                    show_doc_number=document_filename_show_doc_number,
                    show_context=document_filename_show_context,
                    show_date=document_filename_show_date,
                    compact_doc_number=document_filename_compact_doc_number,
                    separator=document_filename_separator,
                    extra_context_label="Brutemateriaal",
                )
                opticutter_excel_requested = f"{opticutter_document_base}.xlsx"
                opticutter_excel_filename = _fit_filename_within_path(
                    prod_folder, opticutter_excel_requested
                )
                _record_path_warning(
                    prod_folder,
                    opticutter_excel_requested,
                    opticutter_excel_filename,
                    context=f"Productie '{prod}' – Brutemateriaal {opticutter_doc_type}",
                )
                opticutter_excel_path = os.path.join(
                    prod_folder, opticutter_excel_filename
                )
                opticutter_en1090 = should_require_en1090(prod, en1090_overrides)
                if not en1090_active:
                    opticutter_en1090 = False

                write_order_excel(
                    opticutter_excel_path,
                    opticutter_order_items,
                    company,
                    supplier_for_opticutter_docs,
                    delivery_for_opticutter_docs,
                    opticutter_doc_type,
                    opticutter_doc_num_display or None,
                    project_number=project_number,
                    project_name=project_name,
                    context_label=prod,
                    context_kind="Brutemateriaal",
                    order_remark=opticutter_remark_text or None,
                    total_weight_kg=opticutter_total_weight,
                    en1090_required=opticutter_en1090,
                    en1090_note=en1090_note_text,
                    column_layout=opticutter_column_layout,
                )
                _append_generated_document(
                    opticutter_excel_path,
                    kind="order",
                    file_format="xlsx",
                    selection_key=opticutter_sel_key,
                    context_kind="Brutemateriaal",
                    context_label=prod,
                    doc_type=opticutter_doc_type,
                    doc_number=opticutter_doc_num_display or "",
                    supplier=opticutter_supplier_name,
                )

                opticutter_pdf_requested = f"{opticutter_document_base}.pdf"
                opticutter_pdf_filename = _fit_filename_within_path(
                    prod_folder, opticutter_pdf_requested
                )
                _record_path_warning(
                    prod_folder,
                    opticutter_pdf_requested,
                    opticutter_pdf_filename,
                    context=f"Productie '{prod}' – Brutemateriaal {opticutter_doc_type}",
                )
                opticutter_pdf_path = os.path.join(prod_folder, opticutter_pdf_filename)
                try:
                    generate_pdf_order_platypus(
                        opticutter_pdf_path,
                        company,
                        supplier_for_opticutter_docs,
                        prod,
                        opticutter_order_items,
                        doc_type=opticutter_doc_type,
                        doc_number=opticutter_doc_num_display or None,
                        footer_note=footer_note_text,
                        quote_footer_note=quote_footer_note_text,
                        delivery=delivery_for_opticutter_docs,
                        project_number=project_number,
                        project_name=project_name,
                        label_kind="brutemateriaal",
                        order_remark=opticutter_remark_text or None,
                        total_weight_kg=opticutter_total_weight,
                        en1090_required=opticutter_en1090,
                        en1090_note=en1090_note_text,
                        column_layout=opticutter_column_layout,
                    )
                    _append_generated_document(
                        opticutter_pdf_path,
                        kind="order",
                        file_format="pdf",
                        selection_key=opticutter_sel_key,
                        context_kind="Brutemateriaal",
                        context_label=prod,
                        doc_type=opticutter_doc_type,
                        doc_number=opticutter_doc_num_display or "",
                        supplier=opticutter_supplier_name,
                    )
                except Exception as exc:
                    print(
                        f"[WAARSCHUWING] PDF brutemateriaal mislukt voor {prod}: {exc}",
                        file=sys.stderr,
                    )

            if should_write_order_overview and order_overview_path:
                order_df.to_excel(order_overview_path, index=False)
                _append_generated_document(
                    order_overview_path,
                    kind="opticutter_order_overview",
                    file_format="xlsx",
                    selection_key=make_opticutter_selection_key(prod),
                    context_kind="Brutemateriaal",
                    context_label=prod,
                )

        packlist_items = step_entries.get(prod, [])
        if packlist_items and REPORTLAB_OK:
            try:
                with tempfile.TemporaryDirectory(prefix="previews_", dir=prod_folder) as preview_dir:
                    rendered_previews = step_previews.render_step_files(
                        packlist_items, preview_dir
                    )
                    if rendered_previews:
                        packlist_requested = f"Paklijst_{prod}_{today}.pdf"
                        packlist_filename = _fit_filename_within_path(
                            prod_folder, packlist_requested
                        )
                        _record_path_warning(
                            prod_folder,
                            packlist_requested,
                            packlist_filename,
                            context=f"Productie '{prod}' – Paklijst",
                        )
                        packlist_path = os.path.join(prod_folder, packlist_filename)
                        try:
                            if not generate_packlist_pdf(
                                packlist_path,
                                production=prod,
                                previews=rendered_previews,
                                doc_date=today,
                            ) and os.path.exists(packlist_path):
                                os.unlink(packlist_path)
                            elif os.path.exists(packlist_path):
                                _append_generated_document(
                                    packlist_path,
                                    kind="packlist",
                                    file_format="pdf",
                                    selection_key=make_production_selection_key(prod),
                                    context_kind="Productie",
                                    context_label=prod,
                                )
                        except Exception as exc:
                            print(
                                f"[WAARSCHUWING] Paklijst mislukt voor {prod}: {exc}",
                                file=sys.stderr,
                            )
            except Exception as exc:
                print(
                    f"[WAARSCHUWING] Previews genereren mislukt voor {prod}: {exc}",
                    file=sys.stderr,
                )

        _emit_progress(
            "productions",
            f"Productie verwerkt... {prod}",
            percent=_phase_percent(8, 45, prod_index, production_total),
            done=prod_index,
            total=production_total,
        )

    if finish_copy_items:
        _emit_progress(
            "finish_files",
            f"Afwerkingsbestanden verwerken... 0 van {finish_copy_total}",
            percent=46,
            done=0,
            total=finish_copy_total,
        )
        finish_seen: Dict[str, set[tuple[str, str]]] = defaultdict(set)
        for finish_index, (finish_key, info) in enumerate(finish_copy_items, start=1):
            finish_label = _to_str(info.get("label")) or finish_key
            _emit_progress(
                "finish_files",
                f"Afwerkingsbestanden verwerken... {finish_label}",
                percent=_phase_percent(46, 56, finish_index - 1, finish_copy_total),
                done=finish_index - 1,
                total=finish_copy_total,
            )
            if finish_export_filter and not finish_export_filter.get(finish_key, True):
                continue
            part_numbers = info.get("part_numbers") or set()
            if not part_numbers:
                continue
            folder_name = info.get("folder_name", finish_key)
            target_dir = os.path.join(dest, folder_name)
            os.makedirs(target_dir, exist_ok=True)
            seen_pairs = finish_seen[finish_key]
            zf = None
            if zip_finish_exports:
                zip_name = _fit_filename_within_path(
                    target_dir, f"{folder_name}.zip"
                )
                zip_path = os.path.join(target_dir, zip_name)
                try:
                    zf = zipfile.ZipFile(
                        zip_path,
                        "w",
                        compression=zipfile.ZIP_DEFLATED,
                        compresslevel=6,
                    )
                except TypeError:
                    zf = zipfile.ZipFile(
                        zip_path,
                        "w",
                        compression=zipfile.ZIP_DEFLATED,
                    )
                except (RuntimeError, NotImplementedError):
                    print(
                        "[WAARSCHUWING] ZIP_DEFLATED niet beschikbaar, val terug op ZIP_STORED",
                        file=sys.stderr,
                    )
                    zf = zipfile.ZipFile(
                        zip_path,
                        "w",
                        compression=zipfile.ZIP_STORED,
                    )
            for pn in sorted(part_numbers):
                files = file_index.get(pn, [])
                for src_file in files:
                    transformed = _transform_export_name(os.path.basename(src_file))
                    combo = (src_file, transformed)
                    if combo in seen_pairs:
                        continue
                    seen_pairs.add(combo)
                    ext = os.path.splitext(src_file)[1].lower()
                    if selected_exts_set and ext not in selected_exts_set:
                        continue
                    if zip_finish_exports:
                        if zf is not None:
                            zf.write(src_file, arcname=transformed)
                    else:
                        shutil.copy2(src_file, os.path.join(target_dir, transformed))
            if zf is not None:
                zf.close()
            _emit_progress(
                "finish_files",
                f"Afwerkingsbestanden verwerkt... {finish_label}",
                percent=_phase_percent(46, 56, finish_index, finish_copy_total),
                done=finish_index,
                total=finish_copy_total,
            )

    if finish_groups:
        _emit_progress(
            "finish_documents",
            f"Afwerkingsbonnen voorbereiden... 0 van {finish_doc_total}",
            percent=57,
            done=0,
            total=finish_doc_total,
        )
        for finish_index, (finish_key, info) in enumerate(finish_doc_items, start=1):
            finish_label = _to_str(info.get("label")) or finish_key
            _emit_progress(
                "finish_documents",
                f"Afwerkingsbon voorbereiden... {finish_label}",
                percent=_phase_percent(57, 65, finish_index - 1, finish_doc_total),
                done=finish_index - 1,
                total=finish_doc_total,
            )
            if finish_export_filter and not finish_export_filter.get(finish_key, True):
                label = _to_str(info.get("label")) or finish_key
                _append_document_status(
                    f"Afwerking '{label}' overgeslagen: export uitgeschakeld."
                )
                continue
            rows = list(info.get("rows", []))
            if not rows:
                label = _to_str(info.get("label")) or finish_key
                _append_document_status(
                    f"Afwerking '{label}' overgeslagen: geen BOM-rijen gevonden."
                )
                continue
            supplier = pick_supplier_for_finish(
                finish_key, db, finish_override_map, suppliers_sorted=suppliers_sorted
            )
            chosen[make_finish_selection_key(finish_key)] = supplier.supplier
            if remember_defaults and supplier.supplier not in ("", "Onbekend", NO_SUPPLIER_PLACEHOLDER):
                db.set_default_finish(finish_key, supplier.supplier)

            raw_doc_type = finish_doc_type_map.get(finish_key, "Bestelbon")
            doc_type = _to_str(raw_doc_type).strip() or "Bestelbon"
            doc_type_lower = doc_type.lower()
            is_standaard_doc = doc_type_lower.startswith("standaard")

            supplier_name_clean = _to_str(supplier.supplier).strip()

            doc_num = _normalize_doc_number(
                finish_doc_num_map.get(finish_key, ""), doc_type
            )
            prefix = _prefix_for_doc_type(doc_type)
            if doc_num and prefix and doc_num.upper() == prefix.upper():
                doc_num = ""
            elif doc_num and prefix and not doc_num.upper().startswith(prefix.upper()):
                doc_num = f"{prefix}{doc_num}"
            doc_num_display = format_document_number_for_display(
                doc_num,
                doc_type,
                compact=document_display_compact_doc_number,
            )
            folder_name = info.get("folder_name", finish_key)
            target_dir = os.path.join(dest, folder_name)
            os.makedirs(target_dir, exist_ok=True)

            items = []
            for row in rows:
                item = {
                    "PartNumber": row.get("PartNumber", ""),
                    "Description": row.get("Description", ""),
                    "Materiaal": row.get("Materiaal", ""),
                    "Lengte": row.get("Length profile", ""),
                    "Aantal": _parse_qty(row.get("Aantal", "")),
                    "Oppervlakte": row.get("Oppervlakte", ""),
                    "Gewicht": row.get("Gewicht", ""),
                }
                item[_LINE_PRICE_KEY] = build_order_pricing_item_key(
                    item,
                    context_kind="Afwerking",
                )
                items.append(item)

            label = _to_str(info.get("label")) or finish_key
            filename_component = info.get("filename_component") or finish_key
            delivery = finish_delivery_map.get(finish_key)
            finish_remark = (
                finish_remarks_map.get(finish_key, "") if finish_remarks_map else ""
            ).strip()
            items, column_layout = _apply_order_pricing(
                items,
                finish_pricing_map.get(finish_key),
                context_kind="Afwerking",
                vat_rate=finish_vat_rate_map.get(finish_key, ""),
            )
            total_surface_m2, total_weight_kg = calculate_order_measure_totals(items)
            order_candidates.append(
                OrderDocumentCandidate(
                    selection_key=make_finish_selection_key(finish_key),
                    context_label=label,
                    context_kind="Afwerking",
                    filename_context=filename_component,
                    target_dir=target_dir,
                    supplier=supplier,
                    delivery=delivery,
                    doc_type=doc_type,
                    doc_num=doc_num,
                    doc_num_display=doc_num_display,
                    order_remark=finish_remark or None,
                    items=items,
                    total_surface_m2=total_surface_m2,
                    total_weight_kg=total_weight_kg,
                    column_layout=column_layout,
                )
            )
            _emit_progress(
                "finish_documents",
                f"Afwerkingsbon voorbereid... {finish_label}",
                percent=_phase_percent(57, 65, finish_index, finish_doc_total),
                done=finish_index,
                total=finish_doc_total,
            )
            continue


    if spare_part_groups:
        spare_part_doc_total = len(spare_part_groups)
        _emit_progress(
            "spare_parts",
            f"Spare-partsdocumenten voorbereiden... 0 van {spare_part_doc_total}",
            percent=65,
            done=0,
            total=spare_part_doc_total,
        )
        for spare_index, group in enumerate(spare_part_groups, start=1):
            group_key = _to_str(group.get("key")).strip()
            group_label = _to_str(group.get("label")).strip() or group_key
            is_full_spare_list = bool(group.get("is_full_list"))
            display_label = (
                _to_str(group.get("display_label")).strip()
                or f"Spare Parts - {group_label}"
            )
            document_label = (
                _to_str(group.get("document_label")).strip()
                or ("Klaarleglijst" if is_full_spare_list else display_label)
            )
            progress_label = "Spare-partslijst" if is_full_spare_list else "Spare-partsbon"
            _emit_progress(
                "spare_parts",
                f"{progress_label} voorbereiden... {group_label}",
                percent=_phase_percent(65, 66, spare_index - 1, spare_part_doc_total),
                done=spare_index - 1,
                total=spare_part_doc_total,
            )
            if spare_part_export_filter and not spare_part_export_filter.get(group_key, True):
                _append_document_status(
                    f"Spare-partsgroep '{group_label}' overgeslagen: export uitgeschakeld."
                )
                continue

            raw_items = group.get("items") or []
            items: List[Dict[str, object]] = []
            for raw_item in raw_items:
                if not isinstance(raw_item, Mapping):
                    continue
                item = {
                    "PartNumber": raw_item.get("PartNumber", ""),
                    "Description": raw_item.get("Description", ""),
                    "Aantal": _parse_qty(raw_item.get("Aantal", raw_item.get("quantity", ""))),
                    "Materiaal": raw_item.get("Materiaal", raw_item.get("Material", "")),
                    "Supplier": raw_item.get("Supplier", ""),
                    "Supplier code": raw_item.get("Supplier code", ""),
                    "Manufacturer": raw_item.get("Manufacturer", ""),
                    "Manufacturer code": raw_item.get("Manufacturer code", ""),
                    "Bestelgroep": raw_item.get("Bestelgroep", ""),
                    "Status": raw_item.get("Status", ""),
                }
                line_key = _to_str(raw_item.get("key")).strip()
                if not line_key:
                    line_key = build_order_pricing_item_key(
                        item,
                        context_kind="Spare parts",
                    )
                item[_LINE_PRICE_KEY] = line_key
                items.append(item)

            if not items:
                _append_document_status(
                    f"Spare-partsgroep '{group_label}' overgeslagen: geen onderdelen gevonden."
                )
                continue

            if is_full_spare_list:
                supplier = Supplier(supplier="")
            else:
                supplier = pick_supplier_for_spare_part_group(
                    group_key,
                    _to_str(group.get("default_supplier")).strip(),
                    db,
                    spare_part_override_map,
                    suppliers_sorted=suppliers_sorted,
                )
            spare_sel_key = make_spare_part_selection_key(group_key)
            chosen[spare_sel_key] = supplier.supplier
            if (
                not is_full_spare_list
                and remember_defaults
                and supplier.supplier not in (
                    "",
                    "Onbekend",
                    NO_SUPPLIER_PLACEHOLDER,
                )
            ):
                db.set_default(make_spare_part_default_key(group_key), supplier.supplier)

            raw_doc_type = (
                "Standaard bon"
                if is_full_spare_list
                else spare_part_doc_type_map.get(
                    group_key,
                    group.get("default_doc_type") or "Bestelbon",
                )
            )
            doc_type = _to_str(raw_doc_type).strip() or "Bestelbon"
            doc_num = _normalize_doc_number(
                spare_part_doc_num_map.get(group_key, ""), doc_type
            )
            prefix = _prefix_for_doc_type(doc_type)
            if doc_num and prefix and doc_num.upper() == prefix.upper():
                doc_num = ""
            elif doc_num and prefix and not doc_num.upper().startswith(prefix.upper()):
                doc_num = f"{prefix}{doc_num}"
            doc_num_display = format_document_number_for_display(
                doc_num,
                doc_type,
                compact=document_display_compact_doc_number,
            )
            folder_name = (
                "Spare Parts"
                if is_full_spare_list
                else f"Spare Parts-{_normalize_finish_folder(group_label)}"
            )
            filename_context = (
                "Spare Parts klaarleglijst" if is_full_spare_list else folder_name
            )
            target_dir = os.path.join(dest, folder_name)
            os.makedirs(target_dir, exist_ok=True)
            order_remark = spare_part_remarks_map.get(group_key, "")
            priced_items, priced_layout = _apply_order_pricing(
                items,
                spare_part_pricing_map.get(group_key),
                context_kind="Spare parts",
                vat_rate=spare_part_vat_rate_map.get(group_key, ""),
            )
            if is_full_spare_list:
                base_layout = priced_layout or _spare_part_column_layout(
                    include_internal=False
                )
                column_layout = [
                    column
                    for column in base_layout
                    if _to_str(column.get("key")) not in {"Bestelgroep", "Status"}
                ]
            else:
                column_layout = priced_layout or _spare_part_column_layout()
            order_candidates.append(
                OrderDocumentCandidate(
                    selection_key=spare_sel_key,
                    context_label=document_label,
                    context_kind="Spare parts",
                    filename_context=filename_context,
                    target_dir=target_dir,
                    supplier=supplier,
                    delivery=(
                        None
                        if is_full_spare_list
                        else spare_part_delivery_map.get(group_key)
                    ),
                    doc_type=doc_type,
                    doc_num=doc_num,
                    doc_num_display=doc_num_display,
                    order_remark=order_remark or None,
                    items=priced_items,
                    column_layout=column_layout,
                )
            )
            _emit_progress(
                "spare_parts",
                f"{progress_label} voorbereid... {group_label}",
                percent=_phase_percent(65, 66, spare_index, spare_part_doc_total),
                done=spare_index,
                total=spare_part_doc_total,
            )

    if order_candidates:
        document_jobs = _build_grouped_document_jobs(
            order_candidates,
            document_group_map,
        )
        document_total = len(document_jobs)
        _emit_progress(
            "documents",
            f"Bestelbonnen maken... 0 van {document_total}",
            percent=66,
            done=0,
            total=document_total,
        )
        for document_index, job in enumerate(document_jobs, start=1):
            primary_section = job.sections[0]
            supplier_name_clean = (
                _to_str(job.supplier.supplier).strip() if job.supplier else ""
            )
            job_doc_type_lower = job.doc_type.lower()
            job_is_standaard = job_doc_type_lower.startswith("standaard")
            supplier_for_docs: Supplier | None = job.supplier
            delivery_for_docs = job.delivery
            if job_is_standaard and not supplier_name_clean:
                supplier_for_docs = None
                delivery_for_docs = None
            group_summary = _order_group_summary_text(job.sections)
            is_finish_job = primary_section.context_kind.strip().lower() == "afwerking"
            is_spare_part_job = primary_section.context_kind.strip().lower() == "spare parts"
            if not (supplier_name_clean or job_is_standaard):
                if is_finish_job:
                    if len(job.sections) > 1 and group_summary:
                        _append_document_status(
                            f"Samengestelde afwerking '{group_summary}' overgeslagen: geen leverancier gekozen."
                        )
                    else:
                        _append_document_status(
                            f"Afwerking '{primary_section.context_label}' overgeslagen: geen leverancier gekozen."
                        )
                elif is_spare_part_job:
                    _append_document_status(
                        f"Spare-partsgroep '{primary_section.context_label}' overgeslagen: geen leverancier gekozen."
                    )
                continue

            context_label = _to_str(primary_section.context_label).strip()
            context_kind = _to_str(primary_section.context_kind).strip() or "Productie"
            extra_context_label = "Groep" if len(job.sections) > 1 else ""
            context_for_warning = (
                f"Groep '{job.context_for_filename}'"
                if len(job.sections) > 1
                else _format_order_section_title(context_kind, context_label)
            )
            _emit_progress(
                "documents",
                f"Bestelbon maken... {context_for_warning}",
                percent=_phase_percent(66, 88, document_index - 1, document_total),
                done=document_index - 1,
                total=document_total,
            )

            document_base = build_document_export_basename(
                job.doc_type,
                job.doc_num,
                job.context_for_filename,
                today,
                profile=document_filename_profile,
                show_doc_type=document_filename_show_doc_type,
                show_doc_number=document_filename_show_doc_number,
                show_context=document_filename_show_context,
                show_date=document_filename_show_date,
                compact_doc_number=document_filename_compact_doc_number,
                separator=document_filename_separator,
                extra_context_label=extra_context_label,
            )

            excel_requested = f"{document_base}.xlsx"
            excel_filename = _fit_filename_within_path(job.target_dir, excel_requested)
            _record_path_warning(
                job.target_dir,
                excel_requested,
                excel_filename,
                context=f"{context_for_warning} - {job.doc_type}",
            )
            excel_path = os.path.join(job.target_dir, excel_filename)
            write_order_excel(
                excel_path,
                primary_section.items,
                company,
                supplier_for_docs,
                delivery_for_docs,
                job.doc_type,
                job.doc_num_display or None,
                project_number=project_number,
                project_name=project_name,
                context_label=context_label,
                context_kind=context_kind,
                order_remark=job.order_remark,
                en1090_required=job.en1090_required,
                en1090_note=en1090_note_text,
                sections=job.sections,
            )
            section_keys = [
                _to_str(section.selection_key).strip()
                for section in job.sections
                if _to_str(section.selection_key).strip()
            ]
            _append_generated_document(
                excel_path,
                kind="order",
                file_format="xlsx",
                selection_key=section_keys[0] if section_keys else "",
                selection_keys=section_keys,
                context_kind=context_kind,
                context_label=context_label,
                doc_type=job.doc_type,
                doc_number=job.doc_num_display or "",
                supplier=supplier_name_clean,
            )

            pdf_requested = f"{document_base}.pdf"
            pdf_filename = _fit_filename_within_path(job.target_dir, pdf_requested)
            _record_path_warning(
                job.target_dir,
                pdf_requested,
                pdf_filename,
                context=f"{context_for_warning} - {job.doc_type}",
            )
            pdf_path = os.path.join(job.target_dir, pdf_filename)
            try:
                pdf_created = True
                generate_pdf_order_platypus(
                    pdf_path,
                    company,
                    supplier_for_docs,
                    context_label,
                    primary_section.items,
                    doc_type=job.doc_type,
                    doc_number=job.doc_num_display or None,
                    footer_note=footer_note_text,
                    quote_footer_note=quote_footer_note_text,
                    delivery=delivery_for_docs,
                    project_number=project_number,
                    project_name=project_name,
                    label_kind=context_kind,
                    order_remark=job.order_remark,
                    en1090_required=job.en1090_required,
                    en1090_note=en1090_note_text,
                    sections=job.sections,
                )
                _append_generated_document(
                    pdf_path,
                    kind="order",
                    file_format="pdf",
                    selection_key=section_keys[0] if section_keys else "",
                    selection_keys=section_keys,
                    context_kind=context_kind,
                    context_label=context_label,
                    doc_type=job.doc_type,
                    doc_number=job.doc_num_display or "",
                    supplier=supplier_name_clean,
                )
            except Exception as exc:
                pdf_created = False
                print(
                    f"[WAARSCHUWING] PDF mislukt voor {context_for_warning}: {exc}",
                    file=sys.stderr,
                )
            if is_finish_job and len(job.sections) > 1:
                target_path = pdf_path if pdf_created else excel_path
                target_label = _report_export_path(target_path)
                summary_text = group_summary or context_label
                _append_document_status(
                    f"Samengestelde afwerkingsbon '{summary_text}' opgeslagen in {target_label}."
                )
            _emit_progress(
                "documents",
                f"Bestelbon gemaakt... {context_for_warning}",
                percent=_phase_percent(66, 88, document_index, document_total),
                done=document_index,
                total=document_total,
            )

    # Persist any (possibly unchanged) supplier defaults so that callers can rely on
    # the database reflecting the latest state on disk.
    if export_bom:
        _emit_progress("bom", "BOM exporteren...", percent=90)
        try:
            bom_filename = make_bom_export_filename(
                bom_source_path, today, _transform_export_name
            )
            if production_export_filter:
                bom_export_df = bom_df[
                    bom_df.apply(
                        lambda row: _production_enabled(row.get("Production")),
                        axis=1,
                    )
                ].copy()
            else:
                bom_export_df = bom_df
            bom_export_path = _export_bom_workbook(
                bom_export_df, dest, bom_filename
            )
            _append_generated_document(
                bom_export_path,
                kind="bom",
                file_format="xlsx",
                context_kind="BOM",
                context_label=_to_str(project_name or project_number).strip(),
            )
        except Exception as exc:  # pragma: no cover - unexpected
            raise RuntimeError(f"Kon BOM-export niet opslaan: {exc}") from exc

    if export_related_files and bom_source_path:
        related_exports = find_related_bom_exports(bom_source_path, file_index)
        related_total = len(related_exports)
        if related_total:
            _emit_progress(
                "related_files",
                f"Gerelateerde BOM-bestanden kopieren... 0 van {related_total}",
                percent=94,
                done=0,
                total=related_total,
            )
        for related_index, src_file in enumerate(related_exports, start=1):
            transformed = _transform_export_name(os.path.basename(src_file))
            target_path = os.path.join(dest, transformed)
            shutil.copy2(src_file, target_path)
            _append_generated_document(
                target_path,
                kind="related_bom_file",
                file_format=os.path.splitext(target_path)[1].lstrip(".").lower(),
                context_kind="BOM",
                context_label=_to_str(project_name or project_number).strip(),
            )
            count_copied += 1
            _emit_progress(
                "related_files",
                f"Gerelateerd BOM-bestand gekopieerd... {os.path.basename(src_file)}",
                percent=_phase_percent(94, 97, related_index, related_total),
                done=related_index,
                total=related_total,
            )

    _emit_progress("save", "Exportgegevens opslaan...", percent=98)
    save_to_storage = getattr(db, "save_to_storage", None)
    if callable(save_to_storage):
        save_to_storage()
    else:
        db.save(SUPPLIERS_DB_FILE)

    _emit_progress("done", "Export afgerond.", percent=100)
    return count_copied, chosen


def _append_pdf_production_group(
    plan: List[PdfWorkDossierPlanItem],
    *,
    section_name: str,
    production: str,
    drawing_files: Sequence[str],
    include_order_documents: bool,
    order_document_root: str | None,
    include_offers: bool,
) -> None:
    if include_order_documents:
        for path in _find_order_interleaf_pdfs(
            order_document_root,
            production,
            include_offers=include_offers,
        ):
            plan.append(
                PdfWorkDossierPlanItem(
                    path=path,
                    section_name=section_name,
                    production=production,
                    role="order",
                )
            )
    for path in sorted(_unique_pdf_paths(drawing_files), key=_natural_pdf_name_key):
        plan.append(
            PdfWorkDossierPlanItem(
                path=path,
                section_name=section_name,
                production=production,
                role="drawing",
            )
        )


def _generated_order_document_path(
    record: Mapping[str, object],
    order_document_root: str | None,
) -> str:
    raw_path = _to_str(record.get("path")).strip()
    if not raw_path:
        return ""
    if os.path.isabs(raw_path):
        return raw_path
    root = _to_str(order_document_root).strip()
    if root:
        return os.path.join(root, raw_path)
    return raw_path


def _generated_order_document_selection_pairs(
    record: Mapping[str, object],
) -> List[Tuple[str, str]]:
    keys: List[str] = []
    primary = _to_str(record.get("selection_key")).strip()
    if primary:
        keys.append(primary)
    extra_keys = record.get("selection_keys")
    if isinstance(extra_keys, (list, tuple, set)):
        keys.extend(_to_str(key).strip() for key in extra_keys if _to_str(key).strip())

    pairs: List[Tuple[str, str]] = []
    seen: set[str] = set()
    for key in keys:
        if key in seen:
            continue
        seen.add(key)
        pairs.append(parse_selection_key(key))
    return pairs


def _is_generated_order_pdf(
    record: Mapping[str, object],
    path: str,
    *,
    include_offers: bool,
) -> bool:
    if _to_str(record.get("kind")).strip().casefold() != "order":
        return False
    file_format = (
        _to_str(record.get("format")).strip()
        or _to_str(record.get("file_format")).strip()
    ).casefold()
    if file_format and file_format != "pdf":
        return False
    if not path.lower().endswith(".pdf"):
        return False
    doc_type = _to_str(record.get("doc_type")).strip().casefold()
    if doc_type.startswith("offerte") and not include_offers:
        return False
    return _is_order_interleaf_pdf(path, include_offers=include_offers)


def _is_generated_spare_part_full_list(
    record: Mapping[str, object],
    path: str = "",
) -> bool:
    pairs = _generated_order_document_selection_pairs(record)
    if any(
        kind == "sparepart" and identifier == SPARE_PARTS_FULL_LIST_KEY
        for kind, identifier in pairs
    ):
        return True

    context_kind = _to_str(record.get("context_kind")).strip().casefold()
    if not context_kind.startswith("spare"):
        return False

    hint_parts = [
        record.get("context_label"),
        record.get("document_label"),
        record.get("display_label"),
        os.path.basename(path),
    ]
    return any(
        "klaarleglijst" in _to_str(part).strip().casefold()
        for part in hint_parts
    )


def _order_document_sort_key(
    *,
    kind: str,
    label: str,
    path: str,
) -> Tuple[int, str, Tuple[Tuple[int, object], ...]]:
    rank = {"production": 0, "sparepart": 1, "finish": 2, "opticutter": 3}.get(kind, 4)
    return (rank, _pdf_match_text(label), _natural_pdf_name_key(path))


def _append_pdf_supplementary_order_documents(
    plan: List[PdfWorkDossierPlanItem],
    *,
    include_order_documents: bool,
    include_spare_part_list: bool,
    order_document_root: str | None,
    include_offers: bool,
    generated_order_documents: Sequence[Mapping[str, object]] | None,
    drawing_productions: Sequence[str],
    no_drawing_productions: Sequence[str],
) -> None:
    used_paths = {
        os.path.abspath(item.path).casefold()
        for item in plan
        if _to_str(item.path).strip()
    }
    drawing_keys = {_to_str(production).strip().casefold() for production in drawing_productions}
    candidates: List[
        Tuple[Tuple[int, str, Tuple[Tuple[int, object], ...]], str, str]
    ] = []
    candidate_paths: set[str] = set()

    def add_candidate(path: str, *, kind: str, label: str) -> None:
        clean_path = _to_str(path).strip()
        if not clean_path or not os.path.isfile(clean_path):
            return
        key = os.path.abspath(clean_path).casefold()
        if key in used_paths or key in candidate_paths:
            return
        candidate_paths.add(key)
        clean_label = _to_str(label).strip() or os.path.basename(clean_path)
        candidates.append(
            (
                _order_document_sort_key(
                    kind=kind,
                    label=clean_label,
                    path=clean_path,
                ),
                clean_path,
                clean_label,
            )
        )

    if include_order_documents:
        for production in sorted(
            {
                _to_str(value).strip()
                for value in no_drawing_productions
                if _to_str(value).strip()
                and not is_spare_parts_production(value)
            },
            key=lambda value: value.casefold(),
        ):
            for path in _find_order_interleaf_pdfs(
                order_document_root,
                production,
                include_offers=include_offers,
            ):
                add_candidate(path, kind="production", label=production)

    for record in generated_order_documents or []:
        if not isinstance(record, Mapping):
            continue
        path = _generated_order_document_path(record, order_document_root)
        if not _is_generated_order_pdf(record, path, include_offers=include_offers):
            continue
        is_spare_part_full_list = _is_generated_spare_part_full_list(record, path)
        if is_spare_part_full_list and not include_spare_part_list:
            continue
        if not is_spare_part_full_list and not include_order_documents:
            continue
        pairs = _generated_order_document_selection_pairs(record)
        production_ids = [
            identifier
            for kind, identifier in pairs
            if kind == "production" and _to_str(identifier).strip()
        ]
        if any(identifier.casefold() in drawing_keys for identifier in production_ids):
            continue
        primary_kind = pairs[0][0] if pairs else ""
        if is_spare_part_full_list and not primary_kind:
            primary_kind = "sparepart"
        primary_label = pairs[0][1] if pairs else ""
        if is_spare_part_full_list:
            label = SPARE_PARTS_FULL_LIST_DOCUMENT_LABEL
        else:
            label = (
                _to_str(record.get("context_label")).strip()
                or _to_str(record.get("context_kind")).strip()
                or primary_label
            )
        add_candidate(path, kind=primary_kind, label=label)

    for _sort_key, path, label in sorted(candidates, key=lambda item: item[0]):
        plan.append(
            PdfWorkDossierPlanItem(
                path=path,
                section_name=PDF_WORKDOSSIER_SUPPLEMENTARY_SECTION,
                production=label,
                role="order",
            )
        )
        used_paths.add(os.path.abspath(path).casefold())


def build_pdf_workdossier_plan(
    source: str,
    bom_df: pd.DataFrame,
    *,
    bom_source_path: str | None = None,
    preset: PdfWorkDossierPreset | None = None,
    include_order_documents: bool = False,
    include_spare_part_list: bool = False,
    order_document_root: str | None = None,
    include_offers: bool = False,
    generated_order_documents: Sequence[Mapping[str, object]] | None = None,
) -> List[PdfWorkDossierPlanItem]:
    """Return the ordered PDF files for a work dossier merge."""

    idx = _build_file_index(source, [".pdf"])

    related_bom_pdfs: List[str] = []
    if bom_source_path:
        related_bom_pdfs = [
            path
            for path in find_related_bom_exports(bom_source_path, idx)
            if path.lower().endswith(".pdf")
        ]
        related_bom_pdfs = _unique_pdf_paths(related_bom_pdfs)

    prod_to_files: Dict[str, List[str]] = defaultdict(list)
    for _, row in bom_df.iterrows():
        prod = _to_str(row.get("Production")).strip() or "_Onbekend"
        pn = _to_str(row.get("PartNumber")).strip()
        if not pn:
            continue
        prod_to_files[prod].extend(idx.get(pn, []))

    for prod, paths in list(prod_to_files.items()):
        prod_to_files[prod] = _unique_pdf_paths(paths)

    drawing_prod_to_files = {
        production: paths for production, paths in prod_to_files.items() if paths
    }
    no_drawing_productions = [
        production for production, paths in prod_to_files.items() if not paths
    ]

    plan: List[PdfWorkDossierPlanItem] = []
    if preset is None:
        if not include_order_documents and not include_spare_part_list:
            all_files: List[str] = []
            all_files.extend(related_bom_pdfs)
            for paths in prod_to_files.values():
                all_files.extend(paths)
            for path in sorted(_unique_pdf_paths(all_files), key=_natural_pdf_name_key):
                plan.append(PdfWorkDossierPlanItem(path=path, role="drawing"))
            return plan

        for path in sorted(related_bom_pdfs, key=_natural_pdf_name_key):
            plan.append(
                PdfWorkDossierPlanItem(
                    path=path,
                    section_name="Algemeen",
                    role="bom",
                )
            )
        production_order = sorted(
            drawing_prod_to_files.keys(),
            key=lambda prod: (
                _natural_pdf_name_key(drawing_prod_to_files[prod][0])
                if drawing_prod_to_files[prod]
                else _natural_pdf_name_key(prod)
            ),
        )
        for production in production_order:
            _append_pdf_production_group(
                plan,
                section_name="Alfabetisch",
                production=production,
                drawing_files=drawing_prod_to_files[production],
                include_order_documents=include_order_documents,
                order_document_root=order_document_root,
                include_offers=include_offers,
            )
        if include_order_documents or include_spare_part_list:
            _append_pdf_supplementary_order_documents(
                plan,
                include_order_documents=include_order_documents,
                include_spare_part_list=include_spare_part_list,
                order_document_root=order_document_root,
                include_offers=include_offers,
                generated_order_documents=generated_order_documents,
                drawing_productions=list(drawing_prod_to_files.keys()),
                no_drawing_productions=no_drawing_productions,
            )
        return plan

    sections = [section for section in preset.sections if section.enabled]
    assigned = _assign_pdf_sections(list(drawing_prod_to_files.keys()), sections)
    unmatched_inserted = False

    def unmatched_productions() -> List[str]:
        unmatched = [
            production
            for production in drawing_prod_to_files
            if production not in assigned
        ]
        unmatched.sort(
            key=lambda prod: (
                _natural_pdf_name_key(drawing_prod_to_files[prod][0])
                if drawing_prod_to_files[prod]
                else _natural_pdf_name_key(prod)
            )
        )
        return unmatched

    for section_index, section in enumerate(sections):
        if section.include_bom_pdf:
            for path in sorted(related_bom_pdfs, key=_natural_pdf_name_key):
                plan.append(
                    PdfWorkDossierPlanItem(
                        path=path,
                        section_name=section.name,
                        role="bom",
                    )
                )

        section_productions = [
            production
            for production, assigned_index in assigned.items()
            if assigned_index == section_index
        ]
        section_productions.sort(
            key=lambda prod: (
                _natural_pdf_name_key(drawing_prod_to_files[prod][0])
                if drawing_prod_to_files[prod]
                else _natural_pdf_name_key(prod)
            )
        )
        for production in section_productions:
            _append_pdf_production_group(
                plan,
                section_name=section.name,
                production=production,
                drawing_files=drawing_prod_to_files[production],
                include_order_documents=include_order_documents,
                order_document_root=order_document_root,
                include_offers=include_offers,
            )

        if getattr(section, "include_unmatched", False) and not unmatched_inserted:
            unmatched_inserted = True
            for production in unmatched_productions():
                _append_pdf_production_group(
                    plan,
                    section_name=section.name,
                    production=production,
                    drawing_files=drawing_prod_to_files[production],
                    include_order_documents=include_order_documents,
                    order_document_root=order_document_root,
                    include_offers=include_offers,
                )

    if preset.include_unmatched and not unmatched_inserted:
        for production in unmatched_productions():
            _append_pdf_production_group(
                plan,
                section_name=preset.unmatched_section_name or "Overige",
                production=production,
                drawing_files=drawing_prod_to_files[production],
                include_order_documents=include_order_documents,
                order_document_root=order_document_root,
                include_offers=include_offers,
            )

    if include_order_documents or include_spare_part_list:
        _append_pdf_supplementary_order_documents(
            plan,
            include_order_documents=include_order_documents,
            include_spare_part_list=include_spare_part_list,
            order_document_root=order_document_root,
            include_offers=include_offers,
            generated_order_documents=generated_order_documents,
            drawing_productions=list(drawing_prod_to_files.keys()),
            no_drawing_productions=no_drawing_productions,
        )

    return plan


def combine_workdossier_pdf_from_source(
    source: str,
    bom_df: pd.DataFrame,
    dest: str,
    date_str: str | None = None,
    *,
    project_number: str | None = None,
    project_name: str | None = None,
    timestamp: datetime.datetime | None = None,
    bom_source_path: str | None = None,
    preset: PdfWorkDossierPreset | None = None,
    include_order_documents: bool = False,
    include_spare_part_list: bool = False,
    order_document_root: str | None = None,
    include_offers: bool = False,
    generated_order_documents: Sequence[Mapping[str, object]] | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> CombinedPdfResult:
    """Combine drawing PDFs into one work dossier PDF."""

    if PdfMerger is None:
        raise ModuleNotFoundError(
            "PyPDF2 must be installed to combine PDF files"
        )

    date_str = date_str or datetime.date.today().strftime("%Y-%m-%d")
    out_dir = _create_combined_output_dir(
        dest,
        project_number,
        project_name,
        timestamp=timestamp,
    )

    plan = build_pdf_workdossier_plan(
        source,
        bom_df,
        bom_source_path=bom_source_path,
        preset=preset,
        include_order_documents=include_order_documents,
        include_spare_part_list=include_spare_part_list,
        order_document_root=order_document_root,
        include_offers=include_offers,
        generated_order_documents=generated_order_documents,
    )
    if not plan:
        if progress_callback is not None:
            progress_callback(0, 0, "")
        return CombinedPdfResult(count=0, output_dir=out_dir)

    merger = PdfMerger()
    output_path = ""
    appended = 0
    total_items = len(plan)
    try:
        for item in plan:
            merger.append(item.path)
            appended += 1
            if progress_callback is not None:
                progress_callback(appended, total_items, item.path)
        out_name = f"Werkdossier_{date_str}_combined.pdf"
        safe_name = _fit_filename_within_path(out_dir, out_name)
        output_path = os.path.join(out_dir, safe_name)
        merger.write(output_path)
    finally:
        merger.close()

    if appended <= 0 or not output_path:
        return CombinedPdfResult(count=0, output_dir=out_dir)
    return CombinedPdfResult(
        count=1,
        output_dir=out_dir,
        output_files=[output_path],
    )


def combine_pdfs_from_source(
    source: str,
    bom_df: pd.DataFrame,
    dest: str,
    date_str: str | None = None,
    *,
    project_number: str | None = None,
    project_name: str | None = None,
    timestamp: datetime.datetime | None = None,
    combine_per_production: bool = True,
    bom_source_path: str | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> CombinedPdfResult:
    """Combine PDF drawing files per production directly from ``source``.

    The BOM dataframe provides ``PartNumber`` to ``Production`` mappings.
    PDFs matching the part numbers are searched in ``source`` using
    :func:`_build_file_index` and merged per production when
    ``combine_per_production`` is :data:`True`. When the flag is :data:`False`,
    every matching PDF in the BOM is merged into a single export file. The
    resulting files are written to a newly created export directory inside
    ``dest`` whose name contains the project number, project name (slugified)
    and an ISO-like timestamp. Output filenames contain either the production
    name or ``BOM`` together with the current date. The returned
    :class:`CombinedPdfResult` provides the number of generated files and the
    absolute output directory path.
    """
    if PdfMerger is None:
        raise ModuleNotFoundError(
            "PyPDF2 must be installed to combine PDF files"
        )

    date_str = date_str or datetime.date.today().strftime("%Y-%m-%d")
    idx = _build_file_index(source, [".pdf"])

    related_bom_pdfs: List[str] = []
    if bom_source_path:
        seen_related: set[str] = set()
        for path in find_related_bom_exports(bom_source_path, idx):
            if not path.lower().endswith(".pdf"):
                continue
            if path in seen_related:
                continue
            related_bom_pdfs.append(path)
            seen_related.add(path)
    related_bom_pdfs.sort(key=lambda x: os.path.basename(x).lower())

    prod_to_files: Dict[str, List[str]] = defaultdict(list)
    for _, row in bom_df.iterrows():
        prod = (row.get("Production") or "").strip() or "_Onbekend"
        pn = str(row.get("PartNumber", ""))
        prod_to_files[prod].extend(idx.get(pn, []))

    out_dir = _create_combined_output_dir(
        dest,
        project_number,
        project_name,
        timestamp=timestamp,
    )
    count = 0
    output_files: List[str] = []
    progress_done = 0

    if combine_per_production:
        # When combining per production, copy related PDFs to output dir separately
        # instead of including them in each combined PDF
        for path in related_bom_pdfs:
            try:
                shutil.copy2(path, out_dir)
            except Exception:
                pass
        
        production_candidates: Dict[str, List[str]] = {}
        for prod, files in prod_to_files.items():
            candidates = _unique_pdf_paths(list(files))
            if not candidates:
                continue
            production_candidates[prod] = sorted(
                candidates, key=lambda x: os.path.basename(x).lower()
            )
        progress_total = sum(len(paths) for paths in production_candidates.values())

        if progress_total == 0 and progress_callback is not None:
            progress_callback(0, 0, "")

        for prod, candidates in production_candidates.items():
            merger = PdfMerger()
            appended: set[str] = set()
            for path in candidates:
                if path in appended:
                    continue
                merger.append(path)
                appended.add(path)
                progress_done += 1
                if progress_callback is not None:
                    progress_callback(progress_done, progress_total, path)
            if not appended:
                merger.close()
                continue
            out_name = f"{prod}_{date_str}_combined.pdf"
            safe_name = _fit_filename_within_path(out_dir, out_name)
            output_path = os.path.join(out_dir, safe_name)
            merger.write(output_path)
            merger.close()
            output_files.append(output_path)
            count += 1
    else:
        ordered_files: List[str] = []
        seen = set()
        for path in related_bom_pdfs:
            if path not in seen:
                ordered_files.append(path)
                seen.add(path)
        for files in prod_to_files.values():
            for path in files:
                if path not in seen:
                    ordered_files.append(path)
                    seen.add(path)
        if ordered_files:
            merger = PdfMerger()
            ordered_files.sort(key=lambda x: os.path.basename(x).lower())
            if related_bom_pdfs:
                related_sorted = sorted(related_bom_pdfs, key=lambda x: os.path.basename(x).lower())
                related_set = set(related_sorted)
                # Preserve related PDFs at the front by reordering ``ordered_files``.
                rest = [path for path in ordered_files if path not in related_set]
                ordered_files = related_sorted + rest
            progress_total = len(ordered_files)
            for path in ordered_files:
                merger.append(path)
                progress_done += 1
                if progress_callback is not None:
                    progress_callback(progress_done, progress_total, path)
            out_name = f"BOM_{date_str}_combined.pdf"
            safe_name = _fit_filename_within_path(out_dir, out_name)
            output_path = os.path.join(out_dir, safe_name)
            merger.write(output_path)
            merger.close()
            output_files.append(output_path)
            count = 1
        elif progress_callback is not None:
            progress_callback(0, 0, "")

    return CombinedPdfResult(count=count, output_dir=out_dir, output_files=output_files)


def combine_pdfs_per_production(
    dest: str,
    date_str: str | None = None,
    *,
    project_number: str | None = None,
    project_name: str | None = None,
    timestamp: datetime.datetime | None = None,
) -> CombinedPdfResult:
    """Combine PDF drawing files per production folder into single PDFs.

    The resulting files are written to a newly created export directory inside
    ``dest`` whose name contains the project number, project name (slugified)
    and an ISO-like timestamp. Output filenames contain the production name
    and current date. The returned :class:`CombinedPdfResult` provides the
    number of generated files and the absolute output directory path.
    """
    if PdfMerger is None:
        raise ModuleNotFoundError(
            "PyPDF2 must be installed to combine PDF files"
        )

    date_str = date_str or datetime.date.today().strftime("%Y-%m-%d")
    out_dir = _create_combined_output_dir(
        dest,
        project_number,
        project_name,
        timestamp=timestamp,
    )
    out_dir_name = os.path.basename(out_dir)
    count = 0
    output_files: List[str] = []
    for prod in sorted(os.listdir(dest)):
        prod_path = os.path.join(dest, prod)
        if not os.path.isdir(prod_path):
            continue
        if prod == out_dir_name or prod.lower().startswith("combined pdf"):
            continue
        pdfs = [
            f
            for f in os.listdir(prod_path)
            if f.lower().endswith(".pdf") and not f.startswith(("Bestelbon_", "Offerteaanvraag_"))
        ]
        if pdfs:
            merger = PdfMerger()
            pdfs.sort(key=lambda x: x.lower())
            for fname in pdfs:
                merger.append(os.path.join(prod_path, fname))
        else:
            zip_path = None
            prod_prefix = f"{prod}_"
            for fname in sorted(os.listdir(prod_path)):
                if not fname.lower().endswith(".zip"):
                    continue
                stem, _ = os.path.splitext(fname)
                if stem.startswith(prod_prefix):
                    zip_path = os.path.join(prod_path, fname)
                    break
            if zip_path is None:
                fallback_name = _fit_filename_within_path(prod_path, f"{prod}.zip")
                fallback = os.path.join(prod_path, fallback_name)
                if not os.path.isfile(fallback):
                    continue
                zip_path = fallback
            with zipfile.ZipFile(zip_path) as zf:
                zip_pdfs = [
                    name
                    for name in zf.namelist()
                    if name.lower().endswith(".pdf")
                    and not os.path.basename(name).startswith(("Bestelbon_", "Offerteaanvraag_"))
                ]
                if not zip_pdfs:
                    continue
                merger = PdfMerger()
                for name in sorted(
                    zip_pdfs, key=lambda x: os.path.basename(x).lower()
                ):
                    with zf.open(name) as fh:
                        merger.append(io.BytesIO(fh.read()))
        out_name = f"{prod}_{date_str}_combined.pdf"
        safe_name = _fit_filename_within_path(out_dir, out_name)
        output_path = os.path.join(out_dir, safe_name)
        merger.write(output_path)
        merger.close()
        output_files.append(output_path)
        count += 1
    return CombinedPdfResult(count=count, output_dir=out_dir, output_files=output_files)
