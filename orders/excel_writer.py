"""Excel writing functions for orders and BOM exports."""

import datetime
import re
from typing import Dict, List, Optional

import pandas as pd

from helpers import _to_str
from models import Supplier, DeliveryAddress
from en1090 import EN1090_NOTE_TEXT

from . import core

try:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception:
    Alignment = None
    Font = None
    get_column_letter = None


def _order_column_export_label(column: Dict[str, object]) -> str:
    key = _to_str(column.get("key")).strip().lower()
    label = _to_str(column.get("label") or column.get("key") or "").strip()
    label_compact = (
        label.lower()
        .replace("€", "")
        .replace("(euro)", "")
        .replace("(", "")
        .replace(")", "")
        .replace(" ", "")
    )

    if key == "oppervlakte" or label.lower() == "oppervlakte":
        return "m\u00b2"
    if key == "gewicht" or label.lower() in {"gewicht", "gewicht (kg)"}:
        return "kg"
    if key == "eenheidsprijs" or label_compact in {"eenheidsprijs", "unitprice"}:
        return "Prijs/st."
    if key == "totaalprijs" or label_compact in {"totaalprijs", "totalprice"}:
        return "Totaal"
    return label or _to_str(column.get("key")).strip()


def _export_bom_workbook(bom_df: pd.DataFrame, dest: str, filename: str) -> str:
    """Write the processed BOM dataframe to an Excel workbook."""
    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"
    import os
    target_path = os.path.join(dest, filename)
    export_df = bom_df.reset_index(drop=True).copy()
    # Drop status-related columns that are only useful inside the app.
    to_drop = [col for col in core._BOM_STATUS_COLUMNS if col in export_df.columns]
    if to_drop:
        export_df = export_df.drop(columns=to_drop)

    # Normalise quantity column naming to ``QTY.`` and drop aliases.
    qty_aliases = ("QTY.", "Qty.", "Qty", "Quantity", "Aantal")
    qty_columns = [col for col in qty_aliases if col in export_df.columns]
    if "QTY." not in export_df.columns:
        if qty_columns:
            export_df = export_df.rename(columns={qty_columns[0]: "QTY."})
        else:
            export_df["QTY."] = ""
    for alias in qty_columns:
        if alias == "QTY." or alias not in export_df.columns:
            continue
        source = export_df[alias]
        destination = export_df["QTY."]
        dest_str = destination.astype(str)
        missing_mask = dest_str.str.strip().isin(("", "nan"))
        export_df.loc[missing_mask, "QTY."] = source[missing_mask]
        export_df = export_df.drop(columns=alias)

    # Ensure all primary BOM columns are present and appear first.
    for column in core._BOM_EXPORT_BASE_COLUMNS:
        if column not in export_df.columns:
            export_df[column] = ""
    ordered_columns = [c for c in core._BOM_EXPORT_BASE_COLUMNS if c in export_df.columns]
    remaining_columns = [c for c in export_df.columns if c not in ordered_columns]
    export_df = export_df[ordered_columns + remaining_columns]

    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="BOM")
        if Alignment is not None and get_column_letter is not None:
            ws = writer.sheets["BOM"]
            alignment = Alignment(wrap_text=False, vertical="top")
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment
            for col_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[column_letter]:
                    value = cell.value
                    cell_length = len(str(value)) if value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                ws.column_dimensions[column_letter].width = min(max(12, max_length + 2), 80)

    return target_path


def make_bom_export_filename(
    bom_source_path: Optional[str],
    date_iso: str,
    transform,
) -> str:
    """Return a normalized filename for exporting the processed BOM workbook."""
    from pathlib import Path
    source_stem = ""
    if bom_source_path:
        source_stem = Path(bom_source_path).stem
        match = re.search(r"(.*?\bBOM\b)", source_stem, flags=re.IGNORECASE)
        if match:
            source_stem = match.group(1)
        source_stem = source_stem.rstrip(" -_.")
    stem = source_stem or "BOM-FileHopper-Export"
    stem_with_date = f"{stem}-{date_iso}"
    filename = f"{stem_with_date}.xlsx"
    return transform(filename)


def find_related_bom_exports(
    bom_source_path: Optional[str],
    file_index,
) -> List[str]:
    """Return export files whose stem appears in the BOM filename."""
    from pathlib import Path
    if not bom_source_path:
        return []
    stem = Path(bom_source_path).stem.lower()
    if not stem:
        return []
    matches: List[str] = []
    seen: set = set()
    for key in sorted(file_index.keys(), key=len, reverse=True):
        if not key:
            continue
        key_lower = key.lower()
        if len(key_lower) < 4 and not any(ch.isdigit() for ch in key_lower):
            continue
        idx = stem.find(key_lower)
        if idx == -1:
            continue
        if idx > 0 and stem[idx - 1].isalnum():
            continue
        end = idx + len(key_lower)
        if end < len(stem) and stem[end].isalnum():
            continue
        for src_file in file_index.get(key, []):
            if src_file not in seen:
                matches.append(src_file)
                seen.add(src_file)
    return matches


def write_order_excel(
    path: str,
    items: List[Dict[str, object]],
    company_info: Dict[str, str] | None = None,
    supplier: Supplier | None = None,
    delivery: DeliveryAddress | None = None,
    doc_type: str = "Bestelbon",
    doc_number: str | None = None,
    project_number: str | None = None,
    project_name: str | None = None,
    context_label: str | None = None,
    context_kind: str = "productie",
    order_remark: str | None = None,
    total_surface_m2: float | None = None,
    total_weight_kg: float | None = None,
    en1090_required: bool = False,
    en1090_note: Optional[str] = None,
    column_layout: Optional[List[Dict[str, object]]] = None,
) -> None:
    """Write order information to an Excel file with header info."""
    context_kind_clean = (_to_str(context_kind) or "productie").strip() or "productie"
    is_raw_material_order = context_kind_clean.lower().startswith("brutemateriaal")
    column_layout = [dict(col) for col in column_layout] if column_layout else []
    custom_layout = bool(column_layout)
    if custom_layout:
        headers: List[str] = []
        for column in column_layout:
            header = _order_column_export_label(column)
            if not header:
                header = column.get("key", "")
            column["label"] = header
            headers.append(header)
        rows: List[Dict[str, object]] = []
        for item in items:
            row: Dict[str, object] = {}
            for column, header in zip(column_layout, headers):
                key = column.get("key")
                value = item.get(key, "") if key else ""
                if column.get("integer"):
                    value = core._coerce_integer_like(value)
                row[header] = value
            rows.append(row)
        df = pd.DataFrame(rows, columns=headers)
        surface_header: str | None = None
        weight_header: str | None = None
        for column in column_layout:
            if surface_header is None and core.is_order_surface_column(column):
                surface_header = column["label"]
            if weight_header is None and core.is_order_weight_column(column):
                weight_header = column["label"]
        if (
            (surface_header and total_surface_m2 is not None)
            or (weight_header and total_weight_kg is not None)
        ):
            total_row = {header: "" for header in headers}
            if headers:
                total_row[headers[0]] = "Totaal"
            if surface_header and total_surface_m2 is not None:
                total_row[surface_header] = core._format_weight_kg(total_surface_m2)
            if weight_header and total_weight_kg is not None:
                total_row[weight_header] = core._format_weight_kg(total_weight_kg)
            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    else:
        if is_raw_material_order:
            df_columns = ["Profiel", "Materiaal", "Lengte", "St.", "kg"]
        else:
            df_columns = ["PartNumber", "Description", "Materiaal", "Aantal", "Oppervlakte", "Gewicht"]
        df = pd.DataFrame(items, columns=df_columns)
        if is_raw_material_order:
            if total_weight_kg is not None:
                total_row = {
                    "Profiel": "Totaal",
                    "Materiaal": "",
                    "Lengte": "",
                    "St.": "",
                    "kg": core._format_weight_kg(total_weight_kg),
                }
                df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        elif total_surface_m2 is not None or total_weight_kg is not None:
            total_row = {
                "PartNumber": "Totaal",
                "Description": "",
                "Materiaal": "",
                "Aantal": "",
                "Oppervlakte": (
                    core._format_weight_kg(total_surface_m2)
                    if total_surface_m2 is not None
                    else ""
                ),
                "Gewicht": (
                    core._format_weight_kg(total_weight_kg)
                    if total_weight_kg is not None
                    else ""
                ),
            }
            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    note_text = EN1090_NOTE_TEXT if en1090_note is None else _to_str(en1090_note)

    append_note_to_df = en1090_required and note_text and (
        Alignment is None or not hasattr(pd, "ExcelWriter")
    )
    if append_note_to_df:
        blank_row = {col: "" for col in df.columns}
        note_row = {col: "" for col in df.columns}
        if df.columns.tolist():
            note_row[df.columns[0]] = note_text
        df = pd.concat(
            [df, pd.DataFrame([blank_row, note_row])], ignore_index=True
        )

    doc_type_text = (_to_str(doc_type).strip() or "Bestelbon")
    doc_type_text_lower = doc_type_text.lower()
    doc_type_text_slug = re.sub(r"[^0-9a-z]+", "", doc_type_text_lower)
    is_standaard_doc = doc_type_text_lower.startswith("standaard")
    order_remark_text = _to_str(order_remark) if order_remark is not None else ""
    order_remark_has_content = bool(order_remark_text.strip())
    place_remark_in_delivery_block = core._should_place_remark_in_delivery_block(
        order_remark_has_content=order_remark_has_content,
        doc_type_text_slug=doc_type_text_slug,
        is_standaard_doc=is_standaard_doc,
        delivery=delivery,
    )

    header_lines: List[tuple] = []
    today = datetime.date.today().strftime("%Y-%m-%d")
    if doc_number:
        header_lines.append(("Nummer", str(doc_number)))
    header_lines.append(("Datum", today))
    if context_label:
        header_lines.append((context_kind_clean.capitalize(), context_label))
    if project_number:
        header_lines.append(("Projectnummer", project_number))
    if project_name:
        header_lines.append(("Projectnaam", project_name))
    if order_remark_has_content and not place_remark_in_delivery_block:
        header_lines.append(("Opmerking", order_remark_text))
    header_lines.append(("", ""))
    if company_info:
        header_lines.extend(
            [
                ("Bedrijf", company_info.get("name", "")),
                ("Adres", company_info.get("address", "")),
                ("BTW", company_info.get("vat", "")),
                ("E-mail", company_info.get("email", "")),
                ("", ""),
            ]
        )
    supplier_name = _to_str(supplier.supplier).strip() if supplier else ""
    include_supplier_block = supplier is not None and (
        not is_standaard_doc or bool(supplier_name)
    )
    if include_supplier_block:
        full_addr = core.format_supplier_address(supplier)
        header_lines.extend(
            [
                ("Leverancier", supplier.supplier),
                ("Adres", full_addr),
                ("BTW", supplier.btw or ""),
                ("E-mail", supplier.sales_email or ""),
                ("Tel", supplier.phone or ""),
                ("", ""),
            ]
        )
    include_delivery_block = False
    if delivery is not None:
        delivery_has_content = any(
            _to_str(value).strip()
            for value in (delivery.name, delivery.address, delivery.remarks)
        )
        include_delivery_block = (
            not is_standaard_doc
            or delivery_has_content
            or place_remark_in_delivery_block
        )
    elif place_remark_in_delivery_block and not is_standaard_doc:
        include_delivery_block = True
    if include_delivery_block:
        if delivery:
            header_lines.extend(
                [
                    ("Leveradres", ""),
                    ("", delivery.name),
                    ("Adres", delivery.address or ""),
                    ("Opmerking", delivery.remarks or ""),
                ]
            )
        if place_remark_in_delivery_block and order_remark_has_content:
            header_lines.append(("Opmerking", order_remark_text))
        header_lines.append(("", ""))

    startrow = len(header_lines)
    # Try writing a proper Excel file using openpyxl (via pandas). If that
    # fails (missing engine or other issue), fall back to a plain CSV so the
    # user still gets data instead of an empty/absent file.
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, startrow=startrow)
            ws = writer.sheets[list(writer.sheets.keys())[0]]
            for r, (label, value) in enumerate(header_lines, start=1):
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=value)

            if custom_layout:
                left_cols = {
                    _to_str(column.get("label") or column.get("key") or "").strip()
                    or column.get("key", "")
                    for column in column_layout
                    if _to_str(column.get("justify") or "left").strip().lower() != "right"
                }
                wrap_cols = {
                    _to_str(column.get("label") or column.get("key") or "").strip()
                    or column.get("key", "")
                    for column in column_layout
                    if bool(column.get("wrap"))
                }
            else:
                if is_raw_material_order:
                    left_cols = {"Profiel", "Materiaal"}
                    wrap_cols = {"Profiel", "Materiaal"}
                else:
                    left_cols = {"PartNumber", "Description"}
                    wrap_cols = {"PartNumber", "Description"}

            # Apply cell alignment and optional width styling when openpyxl
            # style helpers were imported successfully.
            if Alignment is not None:
                for col_idx, col_name in enumerate(df.columns, start=1):
                    align = Alignment(
                        horizontal="left" if col_name in left_cols else "right",
                        wrap_text=col_name in wrap_cols,
                    )
                    if (
                        not custom_layout
                        and col_name in {"PartNumber", "Profiel"}
                        and get_column_letter is not None
                    ):
                        column_letter = get_column_letter(col_idx)
                        ws.column_dimensions[column_letter].width = 25
                    for row in range(startrow + 1, startrow + len(df) + 2):
                        ws.cell(row=row, column=col_idx).alignment = align

            if en1090_required and note_text and not append_note_to_df:
                note_row = ws.max_row + 2
                cell = ws.cell(row=note_row, column=1, value=note_text)
                if Font is not None:
                    cell.font = Font(bold=True)
                if Alignment is not None:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)
    except Exception:
        # Best-effort fallback: write a CSV next to the requested path so the
        # data isn't lost. If the user really needs an .xlsx they should
        # ensure `openpyxl` is installed in the environment.
        try:
            csv_path = re.sub(r"\.xlsx?$", ".csv", path, flags=re.IGNORECASE)
            df.to_csv(csv_path, index=False)
        except Exception:
            # As a last resort, try the simplest write which may still fail.
            df.to_csv(path + ".csv", index=False)
